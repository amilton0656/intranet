"""
Gera planilha Excel com as unidades do BLISS PEDRO DEMORO
extraídas do Quadro II - NBR 12.721 (25/jan/2015)
Incorporadora: COTA Empreendimentos Imobiliários Ltda
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Dados extraídos do Quadro II ──────────────────────────────────────────────

LOJA = [
    {'bloco': 'Loja',    'numero': '19',   'tipo': 'loja',       'tipologia': '',
     'localizacao': 'Térreo', 'area_priv': 499.290, 'coef': 0.05437301},
]

TORRE1 = []
_t1_tipos = [
    ([101,201,301,401,501,601,701], 79.140, 0.00851840),
    ([102,202,302,402,502,602,702], 76.580, 0.00833961),
    ([103,203,303,403,503,603,703], 71.730, 0.00781440),
    ([104,204,304,404,504,604,704], 73.000, 0.00794975),
    ([105,205,305,405,505,605,705], 73.690, 0.00802489),
    ([106,206,306,406,506,606,706], 75.160, 0.00818497),
]
for nums, area, coef in _t1_tipos:
    for n in nums:
        andar = f'{n // 100}º andar'
        TORRE1.append({'bloco': 'Torre 1', 'numero': str(n), 'tipo': 'apartamento',
                       'tipologia': '', 'localizacao': andar,
                       'area_priv': area, 'coef': coef})

# Coberturas Torre 1
TORRE1 += [
    {'bloco': 'Torre 1', 'numero': '801', 'tipo': 'apartamento', 'tipologia': 'Cobertura',
     'localizacao': '8º andar', 'area_priv': 120.978, 'coef': 0.01317458},
    {'bloco': 'Torre 1', 'numero': '802', 'tipo': 'apartamento', 'tipologia': 'Cobertura',
     'localizacao': '8º andar', 'area_priv': 141.593, 'coef': 0.01541957},
]

TORRE2 = []
_t2_tipos = [
    ([101,201,301,401,501,601,701], 73.800, 0.00803687),
    ([102,202,302,402,502,602,702], 72.510, 0.00789639),
    ([103,203,303,403,503,603,703], 72.550, 0.00790074),
    ([104,204,304,404,504,604,704], 75.420, 0.00821329),
    ([105,205,305,405,505,605,705], 70.390, 0.00766853),
]
for nums, area, coef in _t2_tipos:
    for n in nums:
        andar = f'{n // 100}º andar'
        TORRE2.append({'bloco': 'Torre 2', 'numero': str(n), 'tipo': 'apartamento',
                       'tipologia': '', 'localizacao': andar,
                       'area_priv': area, 'coef': coef})

TORRE2 += [
    {'bloco': 'Torre 2', 'numero': '801', 'tipo': 'apartamento', 'tipologia': 'Cobertura',
     'localizacao': '8º andar', 'area_priv': 103.250, 'coef': 0.01124399},
    {'bloco': 'Torre 2', 'numero': '802', 'tipo': 'apartamento', 'tipologia': 'Cobertura',
     'localizacao': '8º andar', 'area_priv': 108.823, 'coef': 0.01185009},
]

# ── Vagas ─────────────────────────────────────────────────────────────────────
VAGAS = []

def add_vagas(nums, tipo, tipologia, localizacao, area, coef):
    for n in nums:
        VAGAS.append({'bloco': 'Vagas', 'numero': str(n), 'tipo': tipo,
                      'tipologia': tipologia, 'localizacao': localizacao,
                      'area_priv': area, 'coef': coef})

# Vagas cobertas (x50) — área 16.760 m²
add_vagas([1,2,3,4,5,6,7,8,13,14,22,23,26,35,36,37],
          'garagem','coberta','Subsolo', 16.760, 0.00182518)
add_vagas([39,46,47,48,49,50,'52A','53A'],
          'garagem','coberta','Térreo',  16.760, 0.00182518)
add_vagas([54,55,56,57,58,59,60,61,62,63,64,65,66,67,'67A'],
          'garagem','coberta','G1/SL',   16.760, 0.00182518)
add_vagas([76,79,'79A',80,'80A',81,82,84,85,86,'86A'],
          'garagem','coberta','G1/SL',   16.760, 0.00182518)

# Vagas especiais cobertas (x35) — área 30.167 m²
add_vagas([9,10,11,12,15,16,17,18,19,20,21,24,25,27,28,29,30,31,32,33,34],
          'garagem','especial coberta','Subsolo', 30.167, 0.00328520)
add_vagas([68,69,70,71,72,73,74,75],
          'garagem','especial coberta','G1/SL',   30.167, 0.00328520)
# Nota: total listado = 29; 6 não identificados na legenda legível

# Vaga PNE 1 (x1)
add_vagas([40], 'garagem','PNE coberta','Térreo', 30.168, 0.00328531)

# Vaga PNE 2 (x1)
add_vagas([51], 'garagem','PNE coberta','Térreo', 31.425, 0.00342219)

# Vagas descobertas (x7) — área 11.173 m²
add_vagas([52,53],           'garagem','descoberta','Térreo',  11.173, 0.00121675)
add_vagas(['76A',77,78,'82A',83], 'garagem','descoberta','Pilotis', 11.173, 0.00121675)

# Vagas especiais descobertas (x4) — área 20.112 m²
add_vagas([87,88,93,94], 'garagem','especial descoberta','Pilotis', 20.112, 0.00219021)

# Vagas parcialmente descobertas (x1 cada)
add_vagas([89], 'garagem','parc. descoberta','Pilotis', 29.905, 0.00325667)
add_vagas([90], 'garagem','parc. descoberta','Pilotis', 28.152, 0.00306577)
add_vagas([91], 'garagem','parc. descoberta','Pilotis', 26.476, 0.00288325)
add_vagas([92], 'garagem','parc. descoberta','Pilotis', 24.558, 0.00267438)

# ── Boxes ─────────────────────────────────────────────────────────────────────
BOXES = []

def add_boxes(nums, tipo_box, localizacao, area, coef):
    for n in nums:
        BOXES.append({'bloco': 'Boxes', 'numero': f'Box {n}', 'tipo': 'hobby_box',
                      'tipologia': tipo_box, 'localizacao': localizacao,
                      'area_priv': area, 'coef': coef})

# BOX A (x10) — área 2.835 m²
add_boxes([6,24,25,26,27],      'BOX A','Subsolo',  2.835, 0.00030873)
add_boxes([36,42,43],           'BOX A','Térreo',   2.835, 0.00030873)
add_boxes([55],                 'BOX A','G1/SL',    2.835, 0.00030873)
add_boxes([71],                 'BOX A','Pilotis',  2.835, 0.00030873)

# BOX B (x26) — área 3.059 m²
add_boxes([2,3,4,5,7,8,11,12],  'BOX B','Subsolo',  3.059, 0.00033313)
add_boxes([35,40,44],           'BOX B','Térreo',   3.059, 0.00033313)
add_boxes([47,48,49,50,51,52,53,54,61], 'BOX B','G1/SL', 3.059, 0.00033313)
add_boxes([66,67,68,69,70,77],  'BOX B','Pilotis',  3.059, 0.00033313)

# BOX C (x14) — área 3.282 m²
add_boxes([39,41],              'BOX C','Subsolo',  3.282, 0.00035741)
add_boxes([21,23],              'BOX C','Térreo',   3.282, 0.00035741)  # boxes 21,22,23 Térreo? legenda parcial
add_boxes([58,59,60,62],        'BOX C','G1/SL',    3.282, 0.00035741)
add_boxes([74,75,76,78],        'BOX C','Pilotis',  3.282, 0.00035741)

# BOX D (x4) — área 1.992 m²
add_boxes([16,37],              'BOX D','Subsolo',  1.992, 0.00021689)
add_boxes([56],                 'BOX D','Pilotis',  1.992, 0.00021689)
add_boxes([72],                 'BOX D','G1/SL',    1.992, 0.00021689)

# BOX E (x5) — área 3.771 m²
add_boxes([13,14,15],           'BOX E','Subsolo',  3.771, 0.00041066)
add_boxes([45,46],              'BOX E','G1/SL',    3.771, 0.00041066)

# BOX F (x3) — área 4.190 m²  (legenda parcialmente legível)
add_boxes([9,10,17,18,19],      'BOX F','Subsolo',  4.190, 0.00045629)
add_boxes([80,81],              'BOX F','Pilotis',  4.190, 0.00045629)
add_boxes([64,65],              'BOX F','G1/SL',    4.190, 0.00045629)

# BOX G (x10? legenda parcial) — área 4.469 m²
add_boxes([29],                 'BOX G','Subsolo',  4.469, 0.00048668)
add_boxes([63],                 'BOX G','G1/SL',    4.469, 0.00048668)
add_boxes([79],                 'BOX G','Pilotis',  4.469, 0.00048668)

# BOX H (x3? legenda parcial) — área 4.804 m²
add_boxes([32,33,34],           'BOX H','Subsolo',  4.804, 0.00052316)  # legenda parcial
add_boxes([38],                 'BOX H','Térreo',   4.804, 0.00052316)
add_boxes([57],                 'BOX H','G1/SL',    4.804, 0.00052316)
add_boxes([73],                 'BOX H','Pilotis',  4.804, 0.00052316)

# BOX I (x1) — área 5.922 m²
add_boxes([30],                 'BOX I','Subsolo',  5.922, 0.00064491)

# BOX J (x1) — área 6.257 m²
add_boxes([31],                 'BOX J','Subsolo',  6.257, 0.00068139)

# ── Monta lista completa ──────────────────────────────────────────────────────
TODAS = LOJA + TORRE1 + TORRE2 + VAGAS + BOXES

# ── Gera Excel ────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

COR_HEADER = 'A7A3AB'
COR_SUBTIT = 'D6D3D9'
COR_TOTAL  = '1A1A2E'

hdr_font   = Font(bold=True, color='FFFFFF', size=10)
sub_font   = Font(bold=True, color='1A1A2E', size=9)
hdr_fill   = PatternFill('solid', fgColor=COR_HEADER)
sub_fill   = PatternFill('solid', fgColor=COR_SUBTIT)
tot_fill   = PatternFill('solid', fgColor=COR_TOTAL)
tot_font   = Font(bold=True, color='FFFFFF', size=10)
center     = Alignment(horizontal='center', vertical='center')
right      = Alignment(horizontal='right',  vertical='center')
left_align = Alignment(horizontal='left',   vertical='center')

thin = Side(style='thin', color='CCCCCC')
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

def set_header(ws, row, cols):
    for col, (title, width) in enumerate(cols, 1):
        c = ws.cell(row=row, column=col, value=title)
        c.font  = hdr_font
        c.fill  = hdr_fill
        c.alignment = center
        c.border = brd
        ws.column_dimensions[get_column_letter(col)].width = width

# ── Sheet 1: Unidades completas ───────────────────────────────────────────────
ws1 = wb.active
ws1.title = 'Unidades'
ws1.freeze_panes = 'A2'

cols = [
    ('Bloco',30), ('Número',12), ('Tipo',16), ('Tipologia',20),
    ('Localização',18), ('Área Priv. (m²)',16), ('Coef. Proporcionalidade',24),
    ('Valor Tabela (R$)',18), ('Status',14),
]
set_header(ws1, 1, cols)

grupos = {
    'Loja':    LOJA,
    'Torre 1': TORRE1,
    'Torre 2': TORRE2,
    'Vagas':   VAGAS,
    'Boxes':   BOXES,
}

row = 2
for grupo_nome, unidades in grupos.items():
    # Linha de subtítulo do grupo
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(cols))
    c = ws1.cell(row=row, column=1, value=f'  {grupo_nome}  ({len(unidades)} unidades)')
    c.font  = sub_font
    c.fill  = sub_fill
    c.alignment = left_align
    row += 1

    for u in unidades:
        data = [
            u['bloco'], u['numero'], u['tipo'], u['tipologia'],
            u['localizacao'], u['area_priv'], u['coef'],
            0, 'disponivel',
        ]
        for col, val in enumerate(data, 1):
            c = ws1.cell(row=row, column=col, value=val)
            c.border = brd
            c.font = Font(size=9)
            if col in (6, 7, 8):
                c.alignment = right
                if col == 6:
                    c.number_format = '#,##0.000'
                elif col == 7:
                    c.number_format = '0.00000000'
                elif col == 8:
                    c.number_format = '#,##0.00'
            else:
                c.alignment = left_align
        row += 1

# Linha total
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
c = ws1.cell(row=row, column=1, value=f'TOTAL — {len(TODAS)} unidades')
c.font = tot_font; c.fill = tot_fill; c.alignment = center

total_coef = sum(u['coef'] for u in TODAS)
c6 = ws1.cell(row=row, column=7, value=total_coef)
c6.font = tot_font; c6.fill = tot_fill
c6.alignment = right; c6.number_format = '0.00000000'

# ── Sheet 2: Para Importar (formato CSV do sistema) ───────────────────────────
ws2 = wb.create_sheet('Para Importar')
ws2.freeze_panes = 'A2'

cols2 = [
    ('bloco',28), ('numero',12), ('tipo',16), ('tipologia',20),
    ('localizacao',18), ('area_privativa',16), ('area_privativa_acessoria',22),
    ('area_comum',14), ('fracao_ideal',20), ('valor_tabela',16),
    ('status',14), ('descricao1',30), ('descricao2',30), ('descricao3',30),
]
set_header(ws2, 1, cols2)

for i, u in enumerate(TODAS, 2):
    vals = [
        u['bloco'], u['numero'], u['tipo'], u['tipologia'],
        u['localizacao'], u['area_priv'], 0, 0,
        u['coef'], 0, 'disponivel', '', '', '',
    ]
    for col, val in enumerate(vals, 1):
        c = ws2.cell(row=i, column=col, value=val)
        c.font = Font(size=9)
        c.border = brd
        if col in (6,7,8,9,10):
            c.alignment = right
        else:
            c.alignment = left_align

ws2.auto_filter.ref = f'A1:{get_column_letter(len(cols2))}1'

# ── Sheet 3: Resumo por tipo ──────────────────────────────────────────────────
ws3 = wb.create_sheet('Resumo')

resumo = [
    ('LOJA',              'Loja',                  1,     499.290, 0.05437301),
    ('TORRE 1',           'Apto tipo 01 (1-7)',     7,    79.140,  0.00851840),
    ('TORRE 1',           'Apto tipo 02 (1-7)',     7,    76.580,  0.00833961),
    ('TORRE 1',           'Apto tipo 03 (1-7)',     7,    71.730,  0.00781440),
    ('TORRE 1',           'Apto tipo 04 (1-7)',     7,    73.000,  0.00794975),
    ('TORRE 1',           'Apto tipo 05 (1-7)',     7,    73.690,  0.00802489),
    ('TORRE 1',           'Apto tipo 06 (1-7)',     7,    75.160,  0.00818497),
    ('TORRE 1',           'Cobertura 801',          1,   120.978,  0.01317458),
    ('TORRE 1',           'Cobertura 802',          1,   141.593,  0.01541957),
    ('TORRE 2',           'Apto tipo 01 (1-7)',     7,    73.800,  0.00803687),
    ('TORRE 2',           'Apto tipo 02 (1-7)',     7,    72.510,  0.00789639),
    ('TORRE 2',           'Apto tipo 03 (1-7)',     7,    72.550,  0.00790074),
    ('TORRE 2',           'Apto tipo 04 (1-7)',     7,    75.420,  0.00821329),
    ('TORRE 2',           'Apto tipo 05 (1-7)',     7,    70.390,  0.00766853),
    ('TORRE 2',           'Cobertura 801',          1,   103.250,  0.01124399),
    ('TORRE 2',           'Cobertura 802',          1,   108.823,  0.01185009),
    ('VAGAS',             'Vaga coberta',          50,    16.760,  0.00182518),
    ('VAGAS',             'Vaga especial coberta', 35,    30.167,  0.00328520),
    ('VAGAS',             'Vaga PNE coberta',       2,    30.168,  0.00328531),
    ('VAGAS',             'Vaga descoberta',        7,    11.173,  0.00121675),
    ('VAGAS',             'Vaga especial descoberta',4,   20.112,  0.00219021),
    ('VAGAS',             'Vaga 89 parc. desc.',    1,    29.905,  0.00325667),
    ('VAGAS',             'Vaga 90 parc. desc.',    1,    28.152,  0.00306577),
    ('VAGAS',             'Vaga 91 parc. desc.',    1,    26.476,  0.00288325),
    ('VAGAS',             'Vaga 92 parc. desc.',    1,    24.558,  0.00267438),
    ('BOXES',             'BOX A',                 10,     2.835,  0.00030873),
    ('BOXES',             'BOX B',                 26,     3.059,  0.00033313),
    ('BOXES',             'BOX C',                 14,     3.282,  0.00035741),
    ('BOXES',             'BOX D',                  4,     1.992,  0.00021689),
    ('BOXES',             'BOX E',                  5,     3.771,  0.00041066),
    ('BOXES',             'BOX F',                  9,     4.190,  0.00045629),
    ('BOXES',             'BOX G',                  3,     4.469,  0.00048668),
    ('BOXES',             'BOX H',                  4,     4.804,  0.00052316),
    ('BOXES',             'BOX I',                  1,     5.922,  0.00064491),
    ('BOXES',             'BOX J',                  1,     6.257,  0.00068139),
]

cols3 = [
    ('Bloco/Grupo',22), ('Descrição',30), ('Qtde',8),
    ('Área Unit. (m²)',16), ('Coef/Unidade',20), ('Coef Total',20),
]
set_header(ws3, 1, cols3)

for i, (bloco, desc, qty, area, coef) in enumerate(resumo, 2):
    ws3.cell(row=i, column=1, value=bloco).font   = Font(size=9)
    ws3.cell(row=i, column=2, value=desc).font    = Font(size=9)
    ws3.cell(row=i, column=3, value=qty).font     = Font(size=9)
    ws3.cell(row=i, column=4, value=area).font    = Font(size=9)
    ws3.cell(row=i, column=5, value=coef).font    = Font(size=9)
    ws3.cell(row=i, column=6, value=qty * coef).font = Font(size=9)
    ws3.cell(row=i, column=3).alignment = center
    ws3.cell(row=i, column=4).number_format = '#,##0.000'
    ws3.cell(row=i, column=5).number_format = '0.00000000'
    ws3.cell(row=i, column=6).number_format = '0.00000000'
    ws3.cell(row=i, column=4).alignment = right
    ws3.cell(row=i, column=5).alignment = right
    ws3.cell(row=i, column=6).alignment = right
    for col in range(1, 7):
        ws3.cell(row=i, column=col).border = brd

# Total resumo
tot_row = len(resumo) + 2
total_units = sum(r[2] for r in resumo)
total_coef_r = sum(r[2] * r[4] for r in resumo)
ws3.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=3)
c = ws3.cell(row=tot_row, column=1, value=f'TOTAL — {total_units} unidades')
c.font = tot_font; c.fill = tot_fill; c.alignment = center
c6 = ws3.cell(row=tot_row, column=6, value=total_coef_r)
c6.font = tot_font; c6.fill = tot_fill; c6.alignment = right
c6.number_format = '0.00000000'

# ── Salva ─────────────────────────────────────────────────────────────────────
output = r'C:\newway\intranet\bliss_pedro_demoro_unidades.xlsx'
wb.save(output)
print(f'Arquivo gerado: {output}')
print(f'Total de unidades: {len(TODAS)}')
print(f'Soma coeficientes: {sum(u["coef"] for u in TODAS):.8f}')
