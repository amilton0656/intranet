import csv
import hashlib
import io
import json
import logging
import re
from datetime import datetime, date

logger = logging.getLogger(__name__)
from collections import defaultdict, OrderedDict

from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import F, Sum, Count, Q
from django.db.models.functions import ExtractYear, ExtractMonth

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm

from .models import (
    ImportLog, Tabela, Permuta, Vinculo, Venda,
    Unidade, FluxoContrato, FluxoParcela, Comissao, SerieContrato, Parcela, ComissaoObs,
)
from apps.indices.models import IndiceData

# ---------------------------------------------------------------------------
# Helpers de formatação
# ---------------------------------------------------------------------------

def _parse_float(s):
    if not s or s.strip() in ('', '-', '—'):
        return 0.0
    cleaned = s.strip().replace('.', '').replace(',', '.')
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _parse_date(s):
    if not s:
        return None
    for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(s.strip(), fmt)
        except ValueError:
            continue
    return None


def _require_float(s, campo, linha, erros):
    """Como _parse_float, mas registra erro se a string não é vazia e falha no parse."""
    if not s or s.strip() in ('', '-', '—'):
        return 0.0
    cleaned = s.strip().replace('.', '').replace(',', '.')
    try:
        return float(cleaned)
    except ValueError:
        erros.append(f'Linha {linha}: "{campo}" com valor inválido → {s!r}')
        return 0.0


def _require_date(s, campo, linha, erros):
    """Como _parse_date, mas registra erro se a string não é vazia e falha no parse."""
    result = _parse_date(s)
    if s and s.strip() and result is None:
        erros.append(f'Linha {linha}: "{campo}" com data inválida → {s!r}')
    return result


def _erros_msg(erros):
    msg = f'{len(erros)} erro(s) de parse encontrado(s):\n' + '\n'.join(erros[:10])
    if len(erros) > 10:
        msg += f'\n...e mais {len(erros) - 10} erro(s).'
    return msg


def _fmt_brl(value):
    if value == 0:
        return 'R$ 0,00'
    formatted = f'{value:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
    return f'R$ {formatted}'


def _fmt_num(value):
    """Formata número no padrão BR sem prefixo R$."""
    if not value:
        return '0,00'
    formatted = f'{abs(value):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
    return formatted


def _get_tipo_serie(serie_name):
    """Mapeia nome da série para código de tipo: AT, PM, RA, PE, CH, FI."""
    s = serie_name.lower()
    if 'financiamento' in s:
        return 'FI'
    if 'ato' in s:
        return 'AT'
    if 'refor' in s:
        return 'RA'
    if 'permuta' in s:
        return 'PE'
    if 'chave' in s:
        return 'CH'
    return 'PM'


def _fmt_m2(value):
    if value == 0:
        return '0,00 m²'
    formatted = f'{value:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
    return f'{formatted} m²'


def _file_sha256(file_obj):
    """Calcula SHA-256 do arquivo e devolve o ponteiro ao início."""
    file_obj.seek(0)
    digest = hashlib.sha256(file_obj.read()).hexdigest()
    file_obj.seek(0)
    return digest


def _xlsx_val(val):
    """Previne injeção de fórmula Excel: strings iniciadas com = + - @ são prefixadas com '."""
    if isinstance(val, str) and val and val[0] in ('=', '+', '-', '@', '\t', '\r'):
        return "'" + val
    return val


def _add_months(dt, n):
    month = dt.month - 1 + n
    year = dt.year + month // 12
    month = month % 12 + 1
    return datetime(year, month, 1)


def _parse_tabela_m2(s):
    return _parse_float(s.replace('m²', '').replace('m2', ''))


def _parse_tabela_brl(s):
    return _parse_float(s.replace('R$', ''))


def _open_csv(file_obj):
    raw = file_obj.read()
    if raw.startswith(b'\xef\xbb\xbf'):
        raw = raw[3:]
    return io.StringIO(raw.decode('utf-8'))


# ---------------------------------------------------------------------------
# Leitura de dados — ORM
# ---------------------------------------------------------------------------

def _latest_competencia():
    from django.db.models import Max
    return Tabela.objects.aggregate(latest=Max('competencia'))['latest']


def _tabela_qs(competencia=None):
    lc = competencia or _latest_competencia()
    if lc is None:
        return Tabela.objects.none()
    return Tabela.objects.filter(competencia=lc)


def _load_tabela():
    return {t.unidade: t.valor_total for t in _tabela_qs()}


def _load_permutas():
    return set(Permuta.objects.values_list('unidade', flat=True))


def _load_vinculos():
    return {
        v.unidade: {'garagens': v.garagens, 'hb': v.hb}
        for v in Vinculo.objects.all()
    }


def _load_vendas():
    return [
        {
            'numero':       v.numero,
            'situacao':     v.situacao,
            'unidade':      v.unidade,
            'm2':           v.m2,
            'cliente':      v.cliente,
            'imobiliaria':  v.imobiliaria,
            'espacos':      v.espacos,
        }
        for v in Venda.objects.all()
    ]


def _load_fluxo():
    rows = []
    for c in FluxoContrato.objects.prefetch_related('parcelas').all():
        monthly = [0.0] * 45
        for p in c.parcelas.all():
            if 0 <= p.mes_idx < 45:
                monthly[p.mes_idx] = p.valor
        rows.append({
            'id':              c.id_contrato,
            'cliente':         c.cliente,
            'unidade':         c.unidade,
            'empreendimento':  c.empreendimento,
            'vgv':             c.vgv,
            'pv':              c.pv,
            'primeira_parcela': datetime(c.primeira_parcela.year,
                                         c.primeira_parcela.month,
                                         c.primeira_parcela.day),
            'ultima_parcela':  datetime(c.ultima_parcela.year,
                                        c.ultima_parcela.month,
                                        c.ultima_parcela.day)
                               if c.ultima_parcela else None,
            'monthly':         monthly,
            'imobiliaria':     c.imobiliaria,
            'corretor':        c.corretor,
        })
    return rows


def _build_monthly_totals(fluxo_rows):
    totals = defaultdict(float)
    for c in fluxo_rows:
        base = c['primeira_parcela'].replace(day=1)
        for i, amt in enumerate(c['monthly']):
            if amt:
                key = _add_months(base, i).strftime('%m/%Y')
                totals[key] += amt
    return OrderedDict(
        sorted(totals.items(), key=lambda x: datetime.strptime(x[0], '%m/%Y'))
    )


def _build_monthly_tipo_breakdown():
    """
    Retorna OrderedDict {mes_key: {tipo: valor}} com tipos AT/PM/RA/PE/CH/FI.
    Usa SerieContrato para identificar o tipo de cada parcela de FluxoParcela.
    Poupança = AT + PM + RA + PE + CH (coluna calculada, não armazenada).
    """
    from collections import Counter

    # Agrupa séries por reserva e tipo
    series_by_reserva = defaultdict(lambda: defaultdict(float))
    for s in SerieContrato.objects.all():
        tipo = _get_tipo_serie(s.serie)
        series_by_reserva[s.reserva][tipo] += s.total

    monthly_breakdown = defaultdict(lambda: defaultdict(float))

    for c in FluxoContrato.objects.prefetch_related('parcelas').all():
        tipos = series_by_reserva.get(c.id_contrato, {})
        base = datetime(c.primeira_parcela.year, c.primeira_parcela.month, 1)

        parcelas = sorted(c.parcelas.all(), key=lambda p: p.mes_idx)
        if not parcelas:
            continue

        first_idx = parcelas[0].mes_idx
        last_idx  = parcelas[-1].mes_idx

        # PM base = valor mais frequente nas parcelas (pagamento mensal padrão)
        vals = [round(p.valor, 2) for p in parcelas]
        pm_base = Counter(vals).most_common(1)[0][0]

        # Detecta Ato parcelado: quando AT_total > mes0_valor, o AT installment = mes0_valor
        at_total      = tipos.get('AT', 0)
        at_installment = 0.0
        num_at_parcelas = 0
        if at_total > 0 and parcelas:
            mes0_val = parcelas[0].valor
            if at_total > mes0_val + 1:
                ratio = at_total / mes0_val
                if abs(ratio - round(ratio)) < 0.01:
                    at_installment  = mes0_val
                    num_at_parcelas = round(ratio)
            else:
                at_installment  = at_total
                num_at_parcelas = 1

        at_paid_count = 0  # controla quantas parcelas de AT já foram contabilizadas

        for p in parcelas:
            month_key = _add_months(base, p.mes_idx).strftime('%m/%Y')
            remaining = p.valor

            # Primeira parcela: pode ser AT puro, ou PE+PM misturado
            if p.mes_idx == first_idx:
                pe_total = tipos.get('PE', 0)
                if at_installment > 0:
                    if abs(at_installment - p.valor) < 1:
                        # AT puro no primeiro mês
                        monthly_breakdown[month_key]['AT'] += p.valor
                        at_paid_count += 1
                        continue
                    elif at_installment < p.valor:
                        # AT + PM no primeiro mês
                        monthly_breakdown[month_key]['AT'] += at_installment
                        remaining -= at_installment
                        at_paid_count += 1
                elif pe_total > 0 and p.valor > pm_base + 1:
                    pe_first = p.valor - pm_base
                    monthly_breakdown[month_key]['PE'] += pe_first
                    remaining = pm_base

            # Parcelas de AT parcelado (além do mês 0)
            elif at_paid_count > 0 and at_paid_count < num_at_parcelas:
                if abs(p.valor - at_installment - pm_base) < 1:
                    monthly_breakdown[month_key]['AT'] += at_installment
                    remaining = pm_base
                    at_paid_count += 1

            # Última parcela: FI, CH ou PE (segunda parcela de permuta)
            elif p.mes_idx == last_idx:
                fi_total = tipos.get('FI', 0)
                ch_total = tipos.get('CH', 0)
                pe_total = tipos.get('PE', 0)
                if fi_total > 0:
                    monthly_breakdown[month_key]['FI'] += p.valor
                    continue
                elif ch_total > 0 and abs(ch_total - p.valor) < 1:
                    monthly_breakdown[month_key]['CH'] += p.valor
                    continue
                elif pe_total > 0 and abs(p.valor - pm_base) > 1:
                    monthly_breakdown[month_key]['PE'] += p.valor
                    continue

            # Penúltima parcela: quando FI é a última, CH pode estar aqui (PM + CH misturados)
            elif p.mes_idx == last_idx - 1 and tipos.get('FI', 0) > 0 and tipos.get('CH', 0) > 0:
                ch_total = tipos['CH']
                if abs(remaining - pm_base - ch_total) < 1:
                    monthly_breakdown[month_key]['CH'] += ch_total
                    remaining = pm_base

            # Reforço anual: qualquer mês (após o primeiro e antes do último)
            # com excedente acima do PM base, quando série RA existe
            if remaining > 0 and p.mes_idx != first_idx and p.mes_idx != last_idx and 'RA' in tipos:
                if remaining > pm_base + 1:
                    monthly_breakdown[month_key]['RA'] += remaining - pm_base
                    remaining = pm_base

            if remaining > 0.01:
                monthly_breakdown[month_key]['PM'] += remaining

    return OrderedDict(
        sorted(monthly_breakdown.items(), key=lambda x: datetime.strptime(x[0], '%m/%Y'))
    )


# ---------------------------------------------------------------------------
# Resumos — ORM
# ---------------------------------------------------------------------------

def _compute_resumos_tabela():
    sit_vt = defaultdict(float)
    sit_ap = defaultdict(float)
    sit_n  = defaultdict(int)
    tip_vt     = defaultdict(float)
    tip_ap     = defaultdict(float)
    tip_n      = defaultdict(int)
    tip_est_vt = defaultdict(float)
    tip_est_ap = defaultdict(float)
    tip_est_n  = defaultdict(int)
    tip_vnd_vt = defaultdict(float)
    tip_vnd_ap = defaultdict(float)
    tip_vnd_n  = defaultdict(int)
    tip_perm_n = defaultdict(int)

    permutas = _load_permutas()
    unidade_tip_map = {}  # unidade -> (tipologia, area_privativa)

    for t in _tabela_qs():
        sit = 'Permuta' if t.unidade in permutas else t.situacao
        tip = t.tipologia
        sit_vt[sit] += t.valor_total
        sit_ap[sit] += t.area_privativa
        sit_n[sit]  += 1
        if tip:
            tip_vt[tip] += t.valor_total
            tip_ap[tip] += t.area_privativa
            tip_n[tip]  += 1
            if sit not in ('Vendida', 'Permuta'):
                tip_est_vt[tip] += t.valor_total
                tip_est_ap[tip] += t.area_privativa
                tip_est_n[tip]  += 1
            if sit == 'Permuta':
                tip_perm_n[tip] += 1
            unidade_tip_map[t.unidade] = (tip, t.area_privativa)

    for v in Venda.objects.all():
        entry = unidade_tip_map.get(v.unidade)
        if entry:
            tip, ap = entry
            tip_vnd_vt[tip] += v.valor_contrato
            tip_vnd_ap[tip] += ap
            tip_vnd_n[tip]  += 1

    total_vt = sum(sit_vt.values()) or 1
    total_ap = sum(sit_ap.values()) or 1
    total_n  = sum(sit_n.values())  or 1

    SIT_ORDER = ['Disponível', 'Reservada', 'Bloqueada', 'Vendida', 'Permuta', 'QA']
    all_sits = sorted(sit_n.keys(), key=lambda s: SIT_ORDER.index(s) if s in SIT_ORDER else 99)

    resumo_sit = []
    for s in all_sits:
        resumo_sit.append({
            'situacao': s,
            'vt_fmt':   _fmt_brl(sit_vt[s]),
            'pct_vt':   f"{sit_vt[s]/total_vt*100:.2f}%",
            'ap_fmt':   _fmt_m2(sit_ap[s]),
            'pct_ap':   f"{sit_ap[s]/total_ap*100:.2f}%",
            'n':        sit_n[s],
            'pct_n':    f"{sit_n[s]/total_n*100:.2f}%",
        })
    resumo_sit.append({
        'situacao': 'Total Geral',
        'vt_fmt':   _fmt_brl(sum(sit_vt.values())),
        'pct_vt':   '100,00%',
        'ap_fmt':   _fmt_m2(sum(sit_ap.values())),
        'pct_ap':   '100,00%',
        'n':        total_n,
        'pct_n':    '100,00%',
        'is_total': True,
    })

    LIQUIDO_SITS = ['Disponível', 'Reservada', 'Bloqueada', 'Vendida']
    liq_vt = sum(sit_vt.get(s, 0.0) for s in LIQUIDO_SITS)
    liq_ap = sum(sit_ap.get(s, 0.0) for s in LIQUIDO_SITS)
    liq_n  = sum(sit_n.get(s, 0)    for s in LIQUIDO_SITS)
    resumo_sit_liquido = []
    for s in LIQUIDO_SITS:
        if s not in sit_n:
            continue
        resumo_sit_liquido.append({
            'situacao': s,
            'vt_fmt':   _fmt_brl(sit_vt[s]),
            'pct_vt':   f"{sit_vt[s]/liq_vt*100:.2f}%" if liq_vt else '0,00%',
            'ap_fmt':   _fmt_m2(sit_ap[s]),
            'pct_ap':   f"{sit_ap[s]/liq_ap*100:.2f}%" if liq_ap else '0,00%',
            'n':        sit_n[s],
            'pct_n':    f"{sit_n[s]/liq_n*100:.2f}%" if liq_n else '0,00%',
        })
    resumo_sit_liquido.append({
        'situacao': 'Total Geral Líquido',
        'vt_fmt':   _fmt_brl(liq_vt),
        'pct_vt':   '100,00%',
        'ap_fmt':   _fmt_m2(liq_ap),
        'pct_ap':   '100,00%',
        'n':        liq_n,
        'pct_n':    '100,00%',
        'is_total': True,
    })

    def _grupo(t):
        tl = t.lower()
        if 'studio' in tl: return 'Studio'
        if 'loja'   in tl: return 'Loja'
        return '2D'

    grp_vt = defaultdict(float)
    grp_ap = defaultdict(float)
    grp_n    = defaultdict(int)
    grp_perm = defaultdict(int)
    for t in tip_n:
        g = _grupo(t)
        grp_vt[g]   += tip_vt[t]
        grp_ap[g]   += tip_ap[t]
        grp_n[g]    += tip_n[t]
        grp_perm[g] += tip_perm_n[t]

    total_tip_vt = sum(grp_vt.values()) or 1
    total_tip_ap = sum(grp_ap.values()) or 1
    total_tip_n  = sum(grp_n.values())
    preco_medio  = total_tip_vt / total_tip_n if total_tip_n else 0

    GRP_ORDER = ['Studio', '2D', 'Loja']
    all_grps = sorted(grp_n.keys(), key=lambda g: GRP_ORDER.index(g) if g in GRP_ORDER else 99)

    resumo_tip = []
    for g in all_grps:
        rsm2 = grp_vt[g] / grp_ap[g] if grp_ap[g] else 0
        resumo_tip.append({
            'n':      grp_n[g],
            'tipo':   g,
            'ap_fmt': f"{grp_ap[g]:,.2f}".replace(',','X').replace('.',',').replace('X','.'),
            'vt_fmt': _fmt_brl(grp_vt[g]),
            'rsm2':   f"{rsm2:,.2f}".replace(',','X').replace('.',',').replace('X','.'),
            'perm_n': grp_perm[g],
        })
    resumo_tip.append({
        'n':      total_tip_n,
        'tipo':   'Total',
        'ap_fmt': f"{total_tip_ap:,.2f}".replace(',','X').replace('.',',').replace('X','.'),
        'vt_fmt': _fmt_brl(total_tip_vt),
        'rsm2':   f"{total_tip_vt/total_tip_ap:,.2f}".replace(',','X').replace('.',',').replace('X','.')
                  if total_tip_ap else '0,00',
        'perm_n': sum(grp_perm.values()),
        'is_total': True,
    })

    est_grp_vt = defaultdict(float)
    est_grp_ap = defaultdict(float)
    est_grp_n  = defaultdict(int)
    for t in tip_est_n:
        g = _grupo(t)
        est_grp_vt[g] += tip_est_vt[t]
        est_grp_ap[g] += tip_est_ap[t]
        est_grp_n[g]  += tip_est_n[t]

    total_est_vt = sum(est_grp_vt.values()) or 1
    total_est_ap = sum(est_grp_ap.values()) or 1
    total_est_n  = sum(est_grp_n.values())
    preco_medio_estoque = total_est_vt / total_est_n if total_est_n else 0

    all_est_grps = sorted(est_grp_n.keys(), key=lambda g: GRP_ORDER.index(g) if g in GRP_ORDER else 99)

    resumo_tip_estoque = []
    for g in all_est_grps:
        rsm2 = est_grp_vt[g] / est_grp_ap[g] if est_grp_ap[g] else 0
        resumo_tip_estoque.append({
            'n':      est_grp_n[g],
            'tipo':   g,
            'ap_fmt': f"{est_grp_ap[g]:,.2f}".replace(',','X').replace('.',',').replace('X','.'),
            'vt_fmt': _fmt_brl(est_grp_vt[g]),
            'rsm2':   f"{rsm2:,.2f}".replace(',','X').replace('.',',').replace('X','.'),
        })
    resumo_tip_estoque.append({
        'n':      total_est_n,
        'tipo':   'Total',
        'ap_fmt': f"{total_est_ap:,.2f}".replace(',','X').replace('.',',').replace('X','.'),
        'vt_fmt': _fmt_brl(total_est_vt),
        'rsm2':   f"{total_est_vt/total_est_ap:,.2f}".replace(',','X').replace('.',',').replace('X','.')
                  if total_est_ap else '0,00',
        'is_total': True,
    })

    # ── Agrupa vendidas por grupo de tipologia ────────────────────────────────
    vnd_grp_vt = defaultdict(float)
    vnd_grp_ap = defaultdict(float)
    vnd_grp_n  = defaultdict(int)
    for t in tip_vnd_n:
        g = _grupo(t)
        vnd_grp_vt[g] += tip_vnd_vt[t]
        vnd_grp_ap[g] += tip_vnd_ap[t]
        vnd_grp_n[g]  += tip_vnd_n[t]

    # ── Consolidado Total / Vendido / Estoque por grupo ──────────────────────
    all_grps_union = sorted(
        set(all_grps) | set(all_est_grps),
        key=lambda g: GRP_ORDER.index(g) if g in GRP_ORDER else 99,
    )
    resumo_tip_total = []
    for g in all_grps_union:
        tot_n  = grp_n.get(g, 0)
        tot_ap = grp_ap.get(g, 0.0)
        tot_vt = grp_vt.get(g, 0.0)
        vnd_n  = vnd_grp_n.get(g, 0)
        vnd_ap = vnd_grp_ap.get(g, 0.0)
        vnd_vt = vnd_grp_vt.get(g, 0.0)
        est_n  = est_grp_n.get(g, 0)
        est_ap = est_grp_ap.get(g, 0.0)
        est_vt = est_grp_vt.get(g, 0.0)
        pct    = vnd_vt / tot_vt * 100 if tot_vt else 0.0
        resumo_tip_total.append({
            'tipo':   g,
            'tot_n':  tot_n,  'tot_ap': tot_ap,  'tot_vt': tot_vt,
            'vnd_n':  vnd_n,  'vnd_ap': vnd_ap,  'vnd_vt': vnd_vt,  'pct': pct,
            'est_n':  est_n,  'est_ap': est_ap,  'est_vt': est_vt,
            'perm_n': grp_perm.get(g, 0),
        })
    _real_tot_ap  = sum(grp_ap.values())
    _real_tot_vt  = sum(grp_vt.values())
    _real_vnd_n   = sum(vnd_grp_n.values())
    _real_vnd_ap  = sum(vnd_grp_ap.values())
    _real_vnd_vt  = sum(vnd_grp_vt.values())
    _real_est_ap  = sum(est_grp_ap.values())
    _real_est_vt  = sum(est_grp_vt.values())
    _real_est_n   = sum(est_grp_n.values())
    resumo_tip_total.append({
        'tipo':     'Total',
        'tot_n':    total_tip_n,   'tot_ap': _real_tot_ap,  'tot_vt': _real_tot_vt,
        'vnd_n':    _real_vnd_n,   'vnd_ap': _real_vnd_ap,  'vnd_vt': _real_vnd_vt,
        'pct':      _real_vnd_vt / _real_tot_vt * 100 if _real_tot_vt else 0.0,
        'est_n':    _real_est_n,   'est_ap': _real_est_ap,  'est_vt': _real_est_vt,
        'perm_n':   sum(grp_perm.values()),
        'is_total': True,
    })

    return (
        resumo_sit,
        resumo_sit_liquido,
        resumo_tip,
        resumo_tip_estoque,
        resumo_tip_total,
        _fmt_brl(preco_medio),
        _fmt_brl(preco_medio_estoque),
        sum(sit_vt.values()),
        sit_vt.get('Permuta',    0.0),
        sit_vt.get('Disponível', 0.0),
        sit_vt.get('Reservada',  0.0),
        sit_vt.get('Vendida',    0.0),
    )


# ---------------------------------------------------------------------------
# Unidades — ORM
# ---------------------------------------------------------------------------

UNIDADE_HEADERS = [
    'Unidade', 'Tipo', 'Complemento do Tipo',
    'Area Privativa', 'Area Priv. Acessoria', 'Area Comum', 'Fracao Ideal',
    'Garagens', 'HB', 'Preço Tabela', 'Status', 'Preço de Venda',
]


def _load_unidades():
    sold_units  = set(Venda.objects.values_list('unidade', flat=True))
    vgv_by_unit = {c.unidade: c.vgv for c in FluxoContrato.objects.all()}
    vinculos    = _load_vinculos()
    tabela      = _load_tabela()
    permutas    = _load_permutas()

    data = []
    for u in Unidade.objects.all():
        un = u.unidade
        vinc = vinculos.get(un, {})
        if un in permutas:
            status = 'Permuta'
        elif un in sold_units:
            status = 'Vendido'
        else:
            status = 'Disponível'
        data.append([
            un,
            u.tipo,
            u.complemento_tipo,
            u.area_privativa,
            u.area_priv_acessoria,
            u.area_comum,
            u.fracao_ideal,
            vinc.get('garagens', ''),
            vinc.get('hb', ''),
            tabela.get(un, ''),
            status,
            vgv_by_unit.get(un, ''),
        ])
    return UNIDADE_HEADERS, data


def _compute_areas():
    qs = Unidade.objects.all()
    area_priv       = sum(u.area_privativa      for u in qs)
    area_priv_acess = sum(u.area_priv_acessoria for u in qs)
    area_comum      = sum(u.area_comum          for u in qs)
    total_priv      = area_priv + area_priv_acess
    area_total      = total_priv + area_comum
    return area_priv, area_priv_acess, total_priv, area_comum, area_total


# ---------------------------------------------------------------------------
# Importação de CSVs
# ---------------------------------------------------------------------------

def _col(*names):
    def getter(row):
        for n in names:
            v = row.get(n)
            if v is not None:
                return v
        row_lower = {k.lower(): v for k, v in row.items()}
        for n in names:
            v = row_lower.get(n.lower())
            if v is not None:
                return v
        return ''
    return getter


@transaction.atomic
def _import_tabela(file_obj, nome, sha256='', competencia=None):
    if competencia is None:
        raise ValueError('Selecione o mês/ano da competência antes de importar a tabela de preços.')
    f = _open_csv(file_obj)
    objs = []
    erros = []
    get_unidade  = _col('UNIDADE')
    get_tipo     = _col('TIPOLOGIA')
    get_sit      = _col('SITUAÇÃO', 'SITUACAO', 'Situação', 'Situacao')
    get_area     = _col('ÁREA PRIVATIVA', 'AREA PRIVATIVA', 'Área Privativa', 'Area Privativa')
    get_valor    = _col('VALOR TOTAL', 'Valor Total')
    for linha, r in enumerate(csv.DictReader(f, delimiter=';'), start=2):
        u = get_unidade(r).strip()
        if not u:
            continue
        objs.append(Tabela(
            unidade        = u,
            competencia    = competencia,
            tipologia      = get_tipo(r).strip(),
            situacao       = get_sit(r).strip(),
            area_privativa = _parse_tabela_m2(get_area(r)),
            valor_total    = _require_float(get_valor(r).replace('R$', ''), 'VALOR TOTAL', linha, erros),
        ))
    if not objs:
        raise ValueError('Nenhuma linha válida encontrada. Verifique o delimitador (;) e o cabeçalho.')
    if erros:
        raise ValueError(_erros_msg(erros))
    Tabela.objects.filter(competencia=competencia).delete()
    Tabela.objects.bulk_create(objs)
    ImportLog.objects.create(tipo='tabela', total_registros=len(objs), nome_arquivo=nome, sha256=sha256)
    vgv = sum(o.valor_total for o in objs)
    return len(objs), {'Competência': competencia.strftime('%m/%Y'), 'VGV total': _fmt_brl(vgv)}


@transaction.atomic
def _import_permutas(file_obj, nome, sha256=''):
    f = _open_csv(file_obj)
    unidades = []
    for row in csv.reader(f):
        u = row[0].strip() if row else ''
        if u:
            unidades.append(u)
    if not unidades:
        raise ValueError('Arquivo vazio ou sem unidades válidas.')
    Permuta.objects.all().delete()
    Permuta.objects.bulk_create([Permuta(unidade=u) for u in unidades])
    ImportLog.objects.create(tipo='permutas', total_registros=len(unidades), nome_arquivo=nome, sha256=sha256)
    return len(unidades), {}


@transaction.atomic
def _import_vinculos(file_obj, nome, sha256=''):
    f = _open_csv(file_obj)
    objs = []
    for r in csv.DictReader(f, delimiter=';'):
        u = r.get('Unidade', '').strip()
        if u:
            objs.append(Vinculo(
                unidade  = u,
                garagens = r.get('Garagens', '').strip(),
                hb       = r.get('HB', '').strip(),
            ))
    if not objs:
        raise ValueError('Nenhuma linha válida encontrada.')
    Vinculo.objects.all().delete()
    Vinculo.objects.bulk_create(objs)
    ImportLog.objects.create(tipo='vinculo', total_registros=len(objs), nome_arquivo=nome, sha256=sha256)
    return len(objs), {}


@transaction.atomic
def _import_vendas(file_obj, nome, sha256=''):
    f = _open_csv(file_obj)
    objs = []
    get_sit     = _col('Situação', 'Situacao', 'SITUAÇÃO')
    get_m2      = _col('M² da unidade', 'M2 da unidade', 'M da unidade')
    get_imob    = _col('Imobiliária', 'Imobiliaria', 'IMOBILIÁRIA')
    get_espacos = _col('Espaços complementares', 'Espacos complementares')
    get_valor   = _col('Valor do contrato', 'Valor Contrato', 'VALOR DO CONTRATO')
    get_data    = _col('Data de Venda', 'Data Venda', 'DATA DE VENDA', 'DATA VENDA')
    for row in csv.DictReader(f, delimiter=';'):
        reserva_raw = row.get('Reserva', '')
        if not reserva_raw.strip():
            continue
        if 'HIPERLINK' in reserva_raw:
            m = re.search(r'"(\d+)"\)', reserva_raw)
            reserva = m.group(1) if m else ''
        else:
            reserva = reserva_raw.strip()
        if not reserva.isdigit():
            continue
        cliente_raw = row.get('Cliente', '')
        if 'HIPERLINK' in cliente_raw:
            m = re.search(r';"([^"]+)"\)', cliente_raw)
            cliente = m.group(1) if m else cliente_raw.strip()
        else:
            cliente = cliente_raw.strip()
        data_venda_dt = _parse_date(get_data(row))
        data_venda = date(data_venda_dt.year, data_venda_dt.month, data_venda_dt.day) if data_venda_dt else None
        objs.append(Venda(
            numero          = reserva,
            situacao        = get_sit(row).strip(),
            unidade         = row.get('Unidade', '').strip(),
            m2              = get_m2(row).strip(),
            cliente         = cliente,
            imobiliaria     = get_imob(row).strip(),
            espacos         = get_espacos(row).strip(),
            valor_contrato  = _parse_float(get_valor(row).replace('R$', '')),
            data_venda      = data_venda,
        ))
    if not objs:
        raise ValueError('Nenhuma reserva válida encontrada.')
    Venda.objects.all().delete()
    Venda.objects.bulk_create(objs)
    ImportLog.objects.create(tipo='vendas', total_registros=len(objs), nome_arquivo=nome, sha256=sha256)
    n_com_valor = sum(1 for o in objs if o.valor_contrato)
    return len(objs), {'Com valor': n_com_valor}


@transaction.atomic
def _import_fluxo(file_obj, nome, sha256=''):
    f = _open_csv(file_obj)
    get_empr  = _col('Empreendimento', 'EMPREENDIMENTO')
    get_imob  = _col('Imobiliária', 'Imobiliaria', 'Imob. Coordenação', 'IMOBILIÁRIA')
    get_corr  = _col('Corretor', 'CORRETOR')
    get_ult   = _col('Última parcela', 'Ultima parcela', 'ÚLTIMA PARCELA')
    contratos = []
    erros = []
    for linha, row in enumerate(csv.DictReader(f, delimiter=';'), start=2):
        primeira_raw = row.get('Primeira parcela', '')
        primeira = _require_date(primeira_raw, 'Primeira parcela', linha, erros)
        if not primeira:
            continue
        ultima_dt = _require_date(get_ult(row), 'Última parcela', linha, erros)
        c = FluxoContrato(
            id_contrato      = row.get('Id.', '').strip(),
            cliente          = row.get('Cliente', '').strip(),
            unidade          = row.get('Unidade', '').strip(),
            empreendimento   = get_empr(row).strip(),
            vgv              = _require_float(row.get('VGV', ''), 'VGV', linha, erros),
            pv               = _parse_float(row.get('PV', '0')),
            primeira_parcela = date(primeira.year, primeira.month, primeira.day),
            ultima_parcela   = date(ultima_dt.year, ultima_dt.month, ultima_dt.day)
                               if ultima_dt else None,
            imobiliaria      = get_imob(row).strip(),
            corretor         = get_corr(row).strip(),
        )
        monthly_raw = [_parse_float(row.get(f'Mês {i}', '0')) for i in range(45)]
        contratos.append((c, monthly_raw))
    if not contratos:
        raise ValueError('Nenhum contrato com data de primeira parcela encontrado.')
    if erros:
        raise ValueError(_erros_msg(erros))
    FluxoContrato.objects.all().delete()
    saved = FluxoContrato.objects.bulk_create([c for c, _ in contratos])
    parcelas_bulk = []
    for contrato_obj, monthly in zip(saved, [m for _, m in contratos]):
        for i, val in enumerate(monthly):
            if val:
                parcelas_bulk.append(FluxoParcela(contrato=contrato_obj, mes_idx=i, valor=val))
    FluxoParcela.objects.bulk_create(parcelas_bulk)
    ImportLog.objects.create(tipo='fluxo', total_registros=len(saved), nome_arquivo=nome, sha256=sha256)
    vgv = sum(c.vgv for c, _ in contratos)
    return len(saved), {'VGV total': _fmt_brl(vgv), 'Parcelas': len(parcelas_bulk)}


@transaction.atomic
def _import_unidades(file_obj, nome, sha256=''):
    f = _open_csv(file_obj)
    get_ap  = _col('Area Privativa', 'Área Privativa', 'ÁREA PRIVATIVA', 'AREA PRIVATIVA')
    get_apa = _col('Area Priv. Acessoria', 'Área Priv. Acessoria', 'Area Priv Acessoria')
    get_ac  = _col('Area Comum', 'Área Comum', 'ÁREA COMUM')
    get_fi  = _col('Fracao Ideal', 'Fração Ideal', 'FRAÇÃO IDEAL')
    objs = []
    for r in csv.DictReader(f, delimiter=';'):
        u = r.get('Unidade', '').strip()
        if not u:
            continue
        objs.append(Unidade(
            unidade             = u,
            tipo                = r.get('Tipo', '').strip(),
            complemento_tipo    = r.get('Complemento do Tipo', '').strip(),
            area_privativa      = _parse_float(get_ap(r)),
            area_priv_acessoria = _parse_float(get_apa(r)),
            area_comum          = _parse_float(get_ac(r)),
            fracao_ideal        = get_fi(r).strip(),
        ))
    if not objs:
        raise ValueError('Nenhuma unidade válida encontrada.')
    Unidade.objects.all().delete()
    Unidade.objects.bulk_create(objs)
    ImportLog.objects.create(tipo='unidades', total_registros=len(objs), nome_arquivo=nome, sha256=sha256)
    area = sum(o.area_privativa for o in objs)
    return len(objs), {'Área privativa total': _fmt_m2(area)}


def _import_comissoes(f, nome, sha256=''):
    def _brl(s, campo, linha, erros):
        s = str(s).strip().strip('"').replace('.', '').replace(',', '.')
        if not s:
            return 0.0
        try:
            return float(s)
        except ValueError:
            erros.append(f'Linha {linha}: "{campo}" com valor inválido → {s!r}')
            return 0.0

    def _col(row, *keys):
        for k in keys:
            v = row.get(k, '')
            if v.strip():
                return v.strip().strip('"')
        return ''

    for enc in ('utf-8-sig', 'latin-1'):
        try:
            raw = f.read().decode(enc)
            f.seek(0)
            break
        except Exception:
            f.seek(0)

    reader = csv.DictReader(raw.splitlines(), delimiter=';')

    # --- Fase 1: parse sem tocar no banco ---
    erros = []
    parsed = []
    for linha, row in enumerate(reader, start=2):
        num = _col(row, 'Número', 'NÃºmero', 'Numero')
        if not num or not num.isdigit():
            continue

        beneficiario  = _col(row, 'Beneficiário', 'BeneficiÃ¡rio')
        tipo_comissao = _col(row, 'Tipo da comissão', 'Tipo da comissÃ£o')

        dados = {
            'reserva':              _col(row, 'Reserva'),
            'corretor':             _col(row, 'Corretor'),
            'imobiliaria':          _col(row, 'Imobiliária', 'ImobiliÃ¡ria'),
            'unidade':              _col(row, 'Unidade'),
            'cliente':              _col(row, 'Cliente'),
            'valor_contrato':       _brl(_col(row, 'Valor do contrato'), 'Valor do contrato', linha, erros),
            'valor_comissao_pagar': _brl(_col(row, 'Valor Comissão a pagar', 'Valor ComissÃ£o a pagar'), 'Valor Comissão a pagar', linha, erros),
            'valor_comissao':       _brl(_col(row, 'Valor da Comissão do Beneficiário', 'Valor da ComissÃ£o do BeneficiÃ¡rio'), 'Valor da Comissão', linha, erros),
            'pct_comissao':         _brl(_col(row, 'Porcentagem da Comissão do Beneficiário', 'Porcentagem da ComissÃ£o do BeneficiÃ¡rio'), 'Porcentagem da Comissão', linha, erros),
        }
        parsed.append((num, beneficiario, tipo_comissao, dados))

    if not parsed:
        raise ValueError('Nenhuma comissão válida encontrada.')
    if erros:
        raise ValueError(_erros_msg(erros))

    # --- Fase 2: persistência ---
    csv_keys = set()
    count = 0
    with transaction.atomic():
        for num, beneficiario, tipo_comissao, dados in parsed:
            csv_keys.add((num, beneficiario, tipo_comissao))
            try:
                obj = Comissao.objects.get(
                    numero=num, beneficiario=beneficiario, tipo_comissao=tipo_comissao)
                if obj.data_prevista or obj.data_pagamento:
                    continue  # protegido — tem data, não atualiza
                for k, v in dados.items():
                    setattr(obj, k, v)
                obj.save()
            except Comissao.DoesNotExist:
                Comissao.objects.create(
                    numero=num,
                    beneficiario=beneficiario,
                    tipo_comissao=tipo_comissao,
                    **dados,
                )
                count += 1

        # Remove apenas registros SEM datas que não estão no CSV
        pks_deletar = [
            obj.pk
            for obj in Comissao.objects.filter(
                data_prevista__isnull=True, data_pagamento__isnull=True)
            if (obj.numero, obj.beneficiario, obj.tipo_comissao) not in csv_keys
        ]
        if pks_deletar:
            Comissao.objects.filter(pk__in=pks_deletar).delete()

        ImportLog.objects.create(tipo='comissoes', total_registros=count, nome_arquivo=nome, sha256=sha256)

    total = sum(c.valor_comissao for c in Comissao.objects.all())
    return count, {'Valor total comissões': _fmt_brl(total)}


def _import_series(file_obj, nome):
    raw = file_obj.read()
    for enc in ('utf-8-sig', 'latin-1', 'utf-8'):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    reader = csv.DictReader(io.StringIO(text), delimiter=';')

    def _parse_brl(s):
        if not s:
            return 0.0
        s = s.strip().replace('R$', '').strip().replace('.', '').replace(',', '.')
        try:
            return float(s)
        except ValueError:
            return 0.0

    objs = []
    for row in reader:
        serie   = (row.get('Série') or row.get('Serie') or row.get('SÃ©rie') or '').strip().strip('"')
        reserva = (row.get('Reserva') or '').strip()
        if not serie or not reserva:
            continue
        total_sc = _parse_brl(row.get('Total Sem Comissão') or row.get('Total Sem ComissÃ£o') or '')
        total    = _parse_brl(row.get('Total') or '')
        objs.append(SerieContrato(
            serie=serie, reserva=reserva,
            total_sem_comissao=total_sc, total=total,
        ))

    if not objs:
        raise ValueError('Nenhuma linha válida encontrada.')
    SerieContrato.objects.all().delete()
    SerieContrato.objects.bulk_create(objs)
    ImportLog.objects.create(tipo='series', total_registros=len(objs), nome_arquivo=nome)
    total_fin = sum(o.total for o in objs if 'financiamento' in o.serie.lower())
    total_poup = sum(o.total for o in objs if 'financiamento' not in o.serie.lower())
    return len(objs), {
        'Financiamento': _fmt_brl(total_fin),
        'Poupança':      _fmt_brl(total_poup),
    }


def _read_csv_text(file_obj):
    raw = file_obj.read()
    for enc in ('utf-8-sig', 'latin-1', 'utf-8'):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode('latin-1', errors='replace')


def _reconcile_parcelas():
    """
    Baixa parcial: quando a mesma chave (titulo+parcela+tipo+vencimento) aparece em
    a_receber e recebidas, o saldo pendente = valor_original - soma_recebida.
    Incluir vencimento na chave evita matches falsos em contratos renegociados,
    onde a numeração de parcelas é reiniciada mas as datas diferem.
    """
    from collections import defaultdict
    paid_by_key = defaultdict(float)
    for p in Parcela.objects.filter(data_pagamento__isnull=False).values('titulo', 'parcela', 'tipo', 'vencimento', 'valor'):
        paid_by_key[(p['titulo'], p['parcela'], p['tipo'], p['vencimento'])] += p['valor']

    if not paid_by_key:
        return

    to_update = []
    to_delete = []
    for p in Parcela.objects.filter(data_pagamento__isnull=True):
        key = (p.titulo, p.parcela, p.tipo, p.vencimento)
        paid = paid_by_key.get(key, 0.0)
        if paid <= 0:
            continue
        remaining = round(p.valor_original - paid, 2)
        if remaining <= 0:
            to_delete.append(p.pk)
        else:
            p.valor = remaining
            to_update.append(p)

    if to_delete:
        Parcela.objects.filter(pk__in=to_delete).delete()
    if to_update:
        Parcela.objects.bulk_update(to_update, ['valor'])


@transaction.atomic
def _import_a_receber(file_obj, nome, sha256=''):
    reader = csv.DictReader(io.StringIO(_read_csv_text(file_obj)), delimiter=';')
    objs = []
    erros = []
    for linha, row in enumerate(reader, start=2):
        titulo = (row.get('nuTitulo') or '').strip()
        if not titulo:
            continue
        vencimento = _require_date(row.get('dtVencto') or '', 'dtVencto', linha, erros)
        v = _require_float(row.get('vlTotal') or '', 'vlTotal', linha, erros)
        objs.append(Parcela(
            titulo=titulo,
            parcela=(row.get('nuParcelaApresentacao') or '').strip(),
            tipo=(row.get('cdTipoCondicao') or '').strip(),
            vencimento=vencimento.date() if vencimento else None,
            data_pagamento=None,
            valor=v,
            valor_original=v,
            cliente=(row.get('nmCliente') or '').strip(),
        ))
    if not objs:
        raise ValueError('Nenhuma linha válida encontrada.')
    if erros:
        raise ValueError(_erros_msg(erros))
    Parcela.objects.filter(data_pagamento__isnull=True).delete()
    Parcela.objects.bulk_create(objs)
    _reconcile_parcelas()
    ImportLog.objects.create(tipo='a_receber', total_registros=len(objs), nome_arquivo=nome, sha256=sha256)
    return len(objs), {'Total': _fmt_brl(sum(o.valor for o in objs))}


@transaction.atomic
def _import_recebidas(file_obj, nome, sha256=''):
    reader = csv.DictReader(io.StringIO(_read_csv_text(file_obj)), delimiter=';')
    objs = []
    erros = []
    for linha, row in enumerate(reader, start=2):
        titulo = (row.get('NumeroDoTitulo') or '').strip()
        if not titulo:
            continue
        vencimento     = _require_date(row.get('DataDeVencimento') or '', 'DataDeVencimento', linha, erros)
        data_pagamento = _require_date(row.get('DataDaBaixa') or '', 'DataDaBaixa', linha, erros)
        objs.append(Parcela(
            titulo=titulo,
            parcela=(row.get('NumeroDaParcela') or '').strip(),
            tipo=(row.get('CodigoDoTipoDeCondicao') or '').strip(),
            vencimento=vencimento.date() if vencimento else None,
            data_pagamento=data_pagamento.date() if data_pagamento else None,
            valor=_require_float(row.get('ValorDaBaixa') or '', 'ValorDaBaixa', linha, erros),
            valor_original=0,
            cliente=(row.get('NomeDoCliente') or '').strip(),
        ))
    if not objs:
        raise ValueError('Nenhuma linha válida encontrada.')
    if erros:
        raise ValueError(_erros_msg(erros))
    Parcela.objects.filter(data_pagamento__isnull=False).delete()
    Parcela.objects.bulk_create(objs)
    _reconcile_parcelas()
    ImportLog.objects.create(tipo='recebidas', total_registros=len(objs), nome_arquivo=nome, sha256=sha256)
    return len(objs), {'Total': _fmt_brl(sum(o.valor for o in objs))}


_IMPORTERS = {
    'tabela':     _import_tabela,
    'permutas':   _import_permutas,
    'vinculo':    _import_vinculos,
    'vendas':     _import_vendas,
    'fluxo':      _import_fluxo,
    'unidades':   _import_unidades,
    'comissoes':  _import_comissoes,
    'series':     _import_series,
    'a_receber':  _import_a_receber,
    'recebidas':  _import_recebidas,
}

_LABELS = {
    'tabela':     'Tabela de Preços',
    'permutas':   'Permutas',
    'vinculo':    'Vínculos',
    'vendas':     'Vendas',
    'fluxo':      'Fluxo de Caixa',
    'unidades':   'Unidades',
    'comissoes':  'Comissões',
    'series':     'Séries de Contratos',
    'a_receber':  'Parcelas a Receber',
    'recebidas':  'Parcelas Recebidas',
}


def comissoes_cadastro(request):
    def _parse_date(s):
        s = (s or '').strip()
        if not s:
            return None
        for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    if request.method == 'POST':
        reserva = request.POST.get('reserva', '').strip()
        n = Comissao.objects.filter(reserva=reserva).update(
            data_prevista  = _parse_date(request.POST.get('data_prevista')),
            data_pagamento = _parse_date(request.POST.get('data_pagamento')),
        )
        if n:
            messages.success(request, f'Datas da reserva {reserva} atualizadas.')
        else:
            messages.error(request, f'Reserva {reserva} não encontrada.')
        return redirect('cota365:comissoes_cadastro')

    # Agrupa por reserva
    obs_map = {o.reserva: o.observacao for o in ComissaoObs.objects.all()}
    groups = defaultdict(list)
    for c in Comissao.objects.order_by('unidade', 'reserva', 'beneficiario'):
        groups[c.reserva].append(c)

    lista = []
    for reserva, records in groups.items():
        first  = records[0]
        extra  = len(records) - 1
        tem_premio = any(
            'premio' in r.tipo_comissao.lower() or 'prêmio' in r.tipo_comissao.lower()
            for r in records
        )
        outros_str = f'+{extra}{"P" if tem_premio else ""}' if extra > 0 else ''
        is_cota = 'COTA EMPREENDIMENTOS' in first.imobiliaria.upper()

        total_comissao = sum(r.valor_comissao for r in records)
        data_prevista  = next((r.data_prevista  for r in records if r.data_prevista),  None)
        data_pagamento = next((r.data_pagamento for r in records if r.data_pagamento), None)

        parceiros = [
            {'beneficiario': r.beneficiario,
             'tipo':         r.tipo_comissao,
             'valor':        _fmt_brl(r.valor_comissao)}
            for r in records
        ]

        lista.append({
            'reserva':            reserva,
            'unidade':            first.unidade,
            'cliente':            first.cliente,
            'imobiliaria':        first.imobiliaria,
            'total_fmt':          _fmt_brl(total_comissao),
            'outros_str':         outros_str,
            'tem_premio':         tem_premio,
            'is_cota':            is_cota,
            'data_prevista_iso':  data_prevista.strftime('%Y-%m-%d')  if data_prevista  else '',
            'data_pagamento_iso': data_pagamento.strftime('%Y-%m-%d') if data_pagamento else '',
            'pago':               bool(data_pagamento),
            'tem_data':           bool(data_prevista or data_pagamento),
            'parceiros_json':     json.dumps(parceiros, ensure_ascii=False),
            'total_json':         _fmt_brl(total_comissao),
            'observacao':         obs_map.get(reserva, ''),
        })

    sort     = request.GET.get('sort', 'unidade')
    sort_dir = request.GET.get('dir', 'asc')
    _SORT_KEYS = {
        'unidade':     lambda x: (x['unidade'] or '').lower(),
        'reserva':     lambda x: (x['reserva'] or '').lower(),
        'cliente':     lambda x: (x['cliente'] or '').lower(),
        'imobiliaria': lambda x: (x['imobiliaria'] or '').lower(),
    }
    lista.sort(key=_SORT_KEYS.get(sort, _SORT_KEYS['unidade']), reverse=(sort_dir == 'desc'))

    return render(request, 'cota365/comissoes_cadastro.html', {
        'lista':      lista,
        'total_n':    len(lista),
        'sort':       sort,
        'sort_dir':   sort_dir,
        'sort_colums': [
            ('unidade',     'Unidade'),
            ('reserva',     'Reserva'),
            ('cliente',     'Cliente'),
            ('imobiliaria', 'Imobiliária'),
        ],
    })


def salvar_obs_reserva(request, reserva):
    from django.http import JsonResponse
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    obs = request.POST.get('observacao', '').strip()
    ComissaoObs.objects.update_or_create(reserva=reserva, defaults={'observacao': obs})
    return JsonResponse({'ok': True})


def delete_reserva(request, reserva):
    if request.method == 'POST':
        n, _ = Comissao.objects.filter(reserva=reserva).delete()
        if n:
            messages.success(request, f'Reserva {reserva} excluída.')
        else:
            messages.error(request, f'Reserva {reserva} não encontrada.')
    return redirect('cota365:comissoes_cadastro')


def export_cadastro_pdf(request):
    if not Comissao.objects.exists():
        return HttpResponse('Sem dados.', status=404)

    cliente_filter = request.GET.get('cliente', '').strip()
    imob_filter    = request.GET.get('imob', '').strip()
    status_filter  = request.GET.get('status', 'todas')

    qs = Comissao.objects.order_by('unidade', 'reserva', 'beneficiario')
    if cliente_filter:
        qs = qs.filter(cliente__icontains=cliente_filter)
    if imob_filter:
        qs = qs.filter(imobiliaria__icontains=imob_filter)

    groups = defaultdict(list)
    for c in qs:
        groups[c.reserva].append(c)

    lista = []
    for reserva, records in groups.items():
        first      = records[0]
        extra      = len(records) - 1
        tem_premio = any('premio' in r.tipo_comissao.lower() or
                         'prêmio' in r.tipo_comissao.lower() for r in records)
        extra_str  = f'+{extra}{"P" if tem_premio else ""}' if extra > 0 else ''
        is_cota    = 'COTA EMPREENDIMENTOS' in first.imobiliaria.upper()
        outros_str = ('COTA ' + extra_str if extra_str else 'COTA') if is_cota else extra_str
        total      = sum(r.valor_comissao for r in records)
        dp_prev    = next((r.data_prevista  for r in records if r.data_prevista),  None)
        dp_pago    = next((r.data_pagamento for r in records if r.data_pagamento), None)
        if status_filter == 'nao-pagas' and dp_pago:
            continue
        lista.append({
            'unidade':    first.unidade,
            'reserva':    reserva,
            'cliente':    first.cliente,
            'imobiliaria':first.imobiliaria,
            'total':      total,
            'outros_str': outros_str,
            'tem_premio': tem_premio,
            'dp_prev':    dp_prev,
            'dp_pago':    dp_pago,
        })
    lista.sort(key=lambda x: x['unidade'])

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    W = doc.width
    styles = getSampleStyleSheet()

    NAVY   = colors.HexColor('#1a1a2e')
    BORDER = colors.HexColor('#dee2e6')
    GREEN  = colors.HexColor('#d1e7dd')
    YELLOW = colors.HexColor('#fff3cd')
    ORANGE = colors.HexColor('#fff3cd')

    def ps(name, **kw):
        return ParagraphStyle(name, parent=styles['Normal'], **kw)

    title_s = ps('CPT', fontSize=14, fontName='Helvetica-Bold', textColor=NAVY, spaceAfter=2)
    sub_s   = ps('CPS', fontSize=8,  textColor=colors.HexColor('#6c757d'), spaceAfter=10)

    def th(txt):
        return Paragraph(f'<b><font color="white">{txt}</font></b>',
                         ps(f'CPTH{txt}', fontSize=7, alignment=1))
    def td(txt):
        return Paragraph(str(txt), ps(f'CPTD{txt}', fontSize=7))
    def tdc(txt):
        return Paragraph(str(txt), ps(f'CPTC{txt}', fontSize=7, alignment=1))
    def tdr(txt):
        return Paragraph(str(txt), ps(f'CPTR{txt}', fontSize=7, alignment=2))
    def tdr_blue(txt):
        return Paragraph(f'<b><font color="#0d6efd">{txt}</font></b>',
                         ps(f'CPLB{txt}', fontSize=7, alignment=2))
    def tdr_green(txt):
        return Paragraph(f'<b><font color="#198754">{txt}</font></b>',
                         ps(f'CPLG{txt}', fontSize=7, alignment=2))

    story = []
    story.append(Paragraph('Cota 365 — Cadastro de Comissões', title_s))
    story.append(Paragraph(
        f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}  |  {len(lista)} reservas',
        sub_s))

    rows = [[th('UNID.'), th('RESERVA'), th('CLIENTE'), th('IMOBILIÁRIA'),
             th('COMISSÃO'), th('OUTRAS'), th('PREVISTA'), th('PAGAMENTO')]]

    row_cmds = [
        ('BACKGROUND',    (0, 0), (-1, 0), NAVY),
        ('GRID',          (0, 0), (-1, -1), 0.3, BORDER),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('TOPPADDING',    (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING',   (0, 0), (-1, -1), 4),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
    ]

    for i, r in enumerate(lista, 1):
        if r['dp_pago']:
            val_cell = tdr_green(_fmt_brl(r['total']))
        elif r['dp_prev']:
            val_cell = tdr_blue(_fmt_brl(r['total']))
        else:
            val_cell = tdr(_fmt_brl(r['total']))

        rows.append([
            td(r['unidade']),
            td(r['reserva']),
            td(r['cliente'][:30]     if r['cliente']     else ''),
            td(r['imobiliaria'][:28] if r['imobiliaria'] else ''),
            val_cell,
            tdc(r['outros_str']),
            td(r['dp_prev'].strftime('%d/%m/%Y') if r['dp_prev'] else '—'),
            td(r['dp_pago'].strftime('%d/%m/%Y') if r['dp_pago'] else '—'),
        ])
        if r['dp_pago']:
            row_cmds.append(('BACKGROUND', (0, i), (-1, i), GREEN))
        elif r['dp_prev']:
            row_cmds.append(('BACKGROUND', (0, i), (-1, i), YELLOW))

    t = Table(rows,
              colWidths=[W*0.07, W*0.07, W*0.22, W*0.24, W*0.14, W*0.07, W*0.10, W*0.09],
              repeatRows=1)
    t.setStyle(TableStyle(row_cmds))
    story.append(t)

    doc.build(story)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/pdf')
    resp['Content-Disposition'] = 'attachment; filename="cadastro_comissoes.pdf"'
    return resp


def comissoes(request):
    sort     = request.GET.get('sort', 'unidade')
    sort_dir = request.GET.get('dir', 'asc')

    qs = list(Comissao.objects.all())

    # KPIs — valor_contrato é o mesmo para todos os beneficiários da mesma reserva
    reservas_vt = {}
    for c in qs:
        if c.reserva not in reservas_vt:
            reservas_vt[c.reserva] = c.valor_contrato
    total_vendas   = len(reservas_vt)
    total_contrato = sum(reservas_vt.values())
    total_comissao = sum(c.valor_comissao        for c in qs)
    total_pagar    = sum(c.valor_comissao_pagar  for c in qs)

    # Resumo por Imobiliária (qtde = reservas únicas)
    imob_map      = defaultdict(lambda: {'comissao': 0.0, 'pagar': 0.0})
    imob_reservas = defaultdict(set)
    for c in qs:
        k = c.imobiliaria or '(sem imobiliária)'
        imob_map[k]['comissao'] += c.valor_comissao
        imob_map[k]['pagar']    += c.valor_comissao_pagar
        imob_reservas[k].add(c.reserva)
    resumo_imob = sorted(
        [{'imobiliaria': k, 'n': len(imob_reservas[k]), **v,
          'comissao_fmt': _fmt_brl(v['comissao']),
          'pagar_fmt':    _fmt_brl(v['pagar'])}
         for k, v in imob_map.items()],
        key=lambda x: x['imobiliaria'].lower(),
    )

    # Resumo por Beneficiário
    benef_map = defaultdict(lambda: {'n': 0, 'comissao': 0.0, 'pagar': 0.0})
    for c in qs:
        k = c.beneficiario or '(sem beneficiário)'
        benef_map[k]['n']        += 1
        benef_map[k]['comissao'] += c.valor_comissao
        benef_map[k]['pagar']    += c.valor_comissao_pagar
    resumo_benef = sorted(
        [{'beneficiario': k, **v,
          'comissao_fmt': _fmt_brl(v['comissao']),
          'pagar_fmt':    _fmt_brl(v['pagar'])}
         for k, v in benef_map.items()],
        key=lambda x: x['beneficiario'].lower(),
    )

    # Lista completa
    lista = [{
        'numero':             c.numero,
        'reserva':            c.reserva,
        'unidade':            c.unidade,
        'cliente':            c.cliente,
        'imobiliaria':        c.imobiliaria,
        'beneficiario':       c.beneficiario,
        'tipo_comissao':      c.tipo_comissao,
        'valor_contrato_fmt': _fmt_brl(c.valor_contrato),
        'pct_comissao':       f"{c.pct_comissao:.2f}%".replace('.', ','),
        'valor_comissao_fmt': _fmt_brl(c.valor_comissao),
        'valor_pagar_fmt':    _fmt_brl(c.valor_comissao_pagar),
    } for c in qs]

    _SORT_KEYS = {
        'unidade':     lambda x: (x['unidade'] or '').lower(),
        'reserva':     lambda x: (x['reserva'] or '').lower(),
        'cliente':     lambda x: (x['cliente'] or '').lower(),
        'imobiliaria': lambda x: (x['imobiliaria'] or '').lower(),
    }
    lista.sort(key=_SORT_KEYS.get(sort, _SORT_KEYS['unidade']), reverse=(sort_dir == 'desc'))

    paginator_lista = Paginator(lista, 50)
    page_lista      = paginator_lista.get_page(request.GET.get('page', 1))

    params = request.GET.copy()
    params.pop('page', None)
    query_string = params.urlencode()

    context = {
        'total_n':            total_vendas,
        'total_linhas':       len(lista),
        'total_contrato_fmt': _fmt_brl(total_contrato),
        'total_comissao_fmt': _fmt_brl(total_comissao),
        'total_pagar_fmt':    _fmt_brl(total_pagar),
        'resumo_imob':        resumo_imob,
        'resumo_benef':       resumo_benef,
        'lista':              page_lista,
        'page_lista':         page_lista,
        'sort':               sort,
        'sort_dir':           sort_dir,
        'query_string':       query_string,
        'sort_colums': [
            ('unidade',     'Unidade'),
            ('reserva',     'Reserva'),
            ('cliente',     'Cliente'),
            ('imobiliaria', 'Imobiliária'),
        ],
    }
    return render(request, 'cota365/comissoes.html', context)


def importar(request):
    if request.method == 'POST':
        tipo = request.POST.get('tipo', '')
        arquivo = request.FILES.get('arquivo')
        if not tipo or tipo not in _IMPORTERS:
            messages.error(request, 'Tipo de arquivo inválido.')
            return redirect('cota365:importar')
        if not arquivo:
            messages.error(request, 'Nenhum arquivo selecionado.')
            return redirect('cota365:importar')
        try:
            sha256 = _file_sha256(arquivo)
            if tipo == 'tabela':
                comp_str = request.POST.get('competencia', '').strip()
                if not comp_str:
                    raise ValueError('Selecione o mês/ano da competência antes de importar a tabela de preços.')
                try:
                    ano, mes = comp_str.split('-')
                    from datetime import date as _date
                    competencia = _date(int(ano), int(mes), 1)
                except (ValueError, AttributeError):
                    raise ValueError('Mês/ano de competência inválido.')
                n, stats = _import_tabela(arquivo, arquivo.name, sha256, competencia=competencia)
            else:
                n, stats = _IMPORTERS[tipo](arquivo, arquivo.name, sha256)
            extras = '  |  '.join(f'{k}: {v}' for k, v in stats.items())
            msg = f'{_LABELS[tipo]} importado — {n} registros.'
            if extras:
                msg += f'  ({extras})'
            messages.success(request, msg)
            logger.info('cota365 import ok | tipo=%s arquivo=%s sha256=%s registros=%d',
                        tipo, arquivo.name, sha256, n)
        except Exception as e:
            logger.exception('cota365 import erro | tipo=%s arquivo=%s', tipo, arquivo.name)
            messages.error(request, f'Erro ao importar {_LABELS[tipo]}: {e}')
        return redirect('cota365:importar')

    # GET — monta status de cada tipo
    ultimos = {}
    for log in ImportLog.objects.all():
        if log.tipo not in ultimos:
            ultimos[log.tipo] = log

    arquivos = []
    for tipo, label in _LABELS.items():
        log = ultimos.get(tipo)
        arquivos.append({
            'tipo':       tipo,
            'label':      label,
            'ultimo':     log.importado_em.strftime('%d/%m/%Y %H:%M') if log else None,
            'registros':  log.total_registros if log else 0,
            'arquivo':    log.nome_arquivo if log else '—',
        })

    from .models import FluxoContrato as FC, Tabela as Tab, Unidade as Un, Venda as Ve
    from django.db.models import Count, Max, Sum
    lc = Tab.objects.aggregate(latest=Max('competencia'))['latest']
    vgv_tab = Tab.objects.filter(competencia=lc).aggregate(total=Sum('valor_total'))['total'] or 0 if lc else 0
    competencias_tabela = list(
        Tab.objects.values('competencia')
        .annotate(total=Count('id'), vgv=Sum('valor_total'))
        .order_by('-competencia')
    )
    resumo = {
        'vgv_tabela':  _fmt_brl(vgv_tab),
        'n_contratos': FC.objects.count(),
        'n_unidades':  Un.objects.count(),
        'n_vendas':    Ve.objects.count(),
    }

    from apps.intranet.context_processors import COTA365_TABELAS
    return render(request, 'cota365/importar.html', {
        'arquivos':            arquivos,
        'resumo':              resumo,
        'cota365_tabelas':     COTA365_TABELAS,
        'competencias_tabela': competencias_tabela,
        'competencia_atual':   lc,
    })


# ---------------------------------------------------------------------------
# Views principais
# ---------------------------------------------------------------------------

def index(request):
    return render(request, 'cota365/index.html')


def dashboard(request):
    # ── Fluxo mensal via Parcela ──────────────────────────────────────────────
    monthly_rec  = {}
    monthly_pend = {}
    for r in (Parcela.objects
              .filter(vencimento__isnull=False)
              .annotate(ano=ExtractYear('vencimento'), mes=ExtractMonth('vencimento'))
              .values('ano', 'mes')
              .annotate(
                  rec=Sum('valor', filter=Q(data_pagamento__isnull=False)),
                  pend=Sum('valor', filter=Q(data_pagamento__isnull=True)),
              )
              .order_by('ano', 'mes')):
        key = (r['ano'], r['mes'])
        monthly_rec[key]  = r['rec']  or 0.0
        monthly_pend[key] = r['pend'] or 0.0

    all_keys    = sorted(set(monthly_rec) | set(monthly_pend))
    total_fluxo = sum(monthly_rec[k] + monthly_pend[k] for k in all_keys)

    acumulado = 0.0
    fluxo_mensal_rows = []
    for key in all_keys:
        yr, mo  = key
        rec     = monthly_rec[key]
        pend    = monthly_pend[key]
        total   = rec + pend
        acumulado += total
        fluxo_mensal_rows.append({
            'mes':           f'{mo:02d}/{yr}',
            'recebido_fmt':  _fmt_brl(rec),
            'a_receber_fmt': _fmt_brl(pend),
            'total_fmt':     _fmt_brl(total),
            'acumulado_fmt': _fmt_brl(acumulado),
        })

    ano_totals = defaultdict(float)
    for key in all_keys:
        yr, mo = key
        ano_totals[str(yr)] += monthly_rec[key] + monthly_pend[key]
    receita_por_ano = [
        {
            'ano':       ano,
            'total_fmt': _fmt_brl(val),
            'pct':       f'{val / total_fluxo * 100:.1f}%' if total_fluxo else '0%',
        }
        for ano, val in sorted(ano_totals.items())
    ]

    # ── KPIs de contrato via Venda + Tabela ───────────────────────────────────
    resumo_sit, resumo_sit_liquido, resumo_tip, resumo_tip_estoque, resumo_tip_total, \
        preco_medio_tipo, preco_medio_estoque, \
        vgv_tabela, vgv_permuta, vgv_disponivel, vgv_reservada, vgv_vendida = _compute_resumos_tabela()
    area_priv, area_priv_acess, total_priv, area_comum, area_total = _compute_areas()

    n_contratos   = Venda.objects.count()
    total_vendido = total_fluxo
    ticket_medio  = total_vendido / n_contratos if n_contratos else 0
    vgv_liquido   = vgv_tabela - vgv_permuta

    # ── Ranking por imobiliária ───────────────────────────────────────────────
    venda_por_imob = {
        r['imobiliaria']: {'cnt': r['cnt'], 'vgv': r['vgv'] or 0.0}
        for r in Venda.objects
            .exclude(imobiliaria='')
            .values('imobiliaria')
            .annotate(cnt=Count('id'), vgv=Sum('valor_contrato'))
    }
    com_por_imob = {
        r['imobiliaria']: r['total'] or 0.0
        for r in Comissao.objects
            .exclude(imobiliaria='')
            .values('imobiliaria')
            .annotate(total=Sum('valor_comissao'))
    }
    all_imobs = set(venda_por_imob)
    total_vgv_imob = sum(v['vgv'] for v in venda_por_imob.values()) or 1.0
    ranking_imobiliaria = sorted(
        [
            {
                'imobiliaria': imob,
                'n_vendas':    venda_por_imob[imob]['cnt'],
                'vgv':         venda_por_imob[imob]['vgv'],
                'vgv_fmt':     _fmt_brl(venda_por_imob[imob]['vgv']),
                'pct':         f"{venda_por_imob[imob]['vgv'] / total_vgv_imob * 100:.1f}%",
                'com_fmt':     _fmt_brl(com_por_imob.get(imob, 0.0)),
                'com_pct':     f"{com_por_imob.get(imob, 0.0) / venda_por_imob[imob]['vgv'] * 100:.1f}%" if venda_por_imob[imob]['vgv'] else '—',
            }
            for imob in all_imobs
        ],
        key=lambda x: x['vgv'],
        reverse=True,
    )
    _tot_vgv = sum(r['vgv'] for r in ranking_imobiliaria)
    _tot_com = sum(com_por_imob.get(imob, 0.0) for imob in all_imobs)
    ranking_total = {
        'n_vendas': sum(r['n_vendas'] for r in ranking_imobiliaria),
        'vgv_fmt':  _fmt_brl(_tot_vgv),
        'com_fmt':  _fmt_brl(_tot_com),
        'com_pct':  f"{_tot_com / _tot_vgv * 100:.1f}%" if _tot_vgv else '—',
    }

    context = {
        'total_geral':          _fmt_brl(vgv_tabela),
        'vgv_liquido':          _fmt_brl(vgv_liquido),
        'n_contratos':          n_contratos,
        'ticket_medio':         _fmt_brl(ticket_medio),
        'total_vendido':        _fmt_brl(total_vendido),
        'area_priv':            _fmt_m2(area_priv),
        'area_priv_acess':      _fmt_m2(area_priv_acess),
        'total_priv':           _fmt_m2(total_priv),
        'area_comum':           _fmt_m2(area_comum),
        'area_total':           _fmt_m2(area_total),
        'resumo_sit':           resumo_sit,
        'resumo_sit_liquido':   resumo_sit_liquido,
        'resumo_tip':           resumo_tip,
        'resumo_tip_estoque':   resumo_tip_estoque,
        'preco_medio_tipo':     preco_medio_tipo,
        'preco_medio_estoque':  preco_medio_estoque,
        'receita_por_ano':      receita_por_ano,
        'total_fluxo_fmt':      _fmt_brl(total_fluxo),
        'fluxo_mensal_rows':    fluxo_mensal_rows,
        'ranking_imobiliaria':  ranking_imobiliaria,
        'ranking_total':        ranking_total,
    }
    return render(request, 'cota365/dashboard.html', context)


def export_dashboard(request):
    from reportlab.platypus import HRFlowable

    # ── Fluxo via Parcela ─────────────────────────────────────────────────────
    exp_monthly_rec  = defaultdict(float)
    exp_monthly_pend = defaultdict(float)
    poupanca_total = 0.0
    fi_total = 0.0
    tipo_totals = defaultdict(float)
    total_recebido  = 0.0
    total_a_receber = 0.0
    for p in Parcela.objects.all():
        if p.data_pagamento:
            key_rec = (p.data_pagamento.year, p.data_pagamento.month)
            exp_monthly_rec[key_rec] += p.valor
            total_recebido           += p.valor
        elif p.vencimento:
            key_pend = (p.vencimento.year, p.vencimento.month)
            exp_monthly_pend[key_pend] += p.valor
            total_a_receber            += p.valor
        tipo_totals[p.tipo] += p.valor
        if p.tipo == 'FI':
            fi_total += p.valor
        else:
            poupanca_total += p.valor

    exp_all_keys = sorted(set(exp_monthly_rec) | set(exp_monthly_pend))
    total_fluxo  = sum(exp_monthly_rec[k] + exp_monthly_pend[k] for k in exp_all_keys)

    ano_totals = defaultdict(float)
    for key in exp_all_keys:
        yr, mo = key
        ano_totals[str(yr)] += exp_monthly_rec[key] + exp_monthly_pend[key]

    # ── KPIs via Venda + Tabela ───────────────────────────────────────────────
    resumo_sit, resumo_sit_liquido, resumo_tip, resumo_tip_estoque, resumo_tip_total, \
        preco_medio_tipo, preco_medio_estoque, \
        vgv_tabela, vgv_permuta, vgv_disponivel, vgv_reservada, vgv_vendida = _compute_resumos_tabela()
    area_priv, area_priv_acess, total_priv, area_comum, area_total = _compute_areas()

    n_contratos   = Venda.objects.count()
    total_vendido = vgv_vendida
    ticket_medio  = total_vendido / n_contratos if n_contratos else 0

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    W = doc.width

    styles = getSampleStyleSheet()
    NAVY   = colors.HexColor('#1a1a2e')
    LIGHT  = colors.HexColor('#f8f9fa')
    TOTAL_BG = colors.HexColor('#e9ecef')
    BORDER = colors.HexColor('#dee2e6')

    def ps(name, **kw):
        return ParagraphStyle(name, parent=styles['Normal'], **kw)

    title_s = ps('T',  fontSize=16, textColor=NAVY, fontName='Helvetica-Bold', spaceAfter=6)
    sub_s   = ps('S',  fontSize=8,  textColor=colors.HexColor('#6c757d'), spaceAfter=14)
    sec_s   = ps('H',  fontSize=9,  textColor=NAVY, fontName='Helvetica-Bold', spaceBefore=12, spaceAfter=5)
    cell_s  = ps('C',  fontSize=8,  leading=11)
    cell_r  = ps('CR', fontSize=8,  leading=11, alignment=2)
    cell_b  = ps('CB', fontSize=8,  leading=11, fontName='Helvetica-Bold')
    cell_rb = ps('CRB',fontSize=8,  leading=11, fontName='Helvetica-Bold', alignment=2)
    cell_c  = ps('CC', fontSize=8,  leading=11, alignment=1)

    def th(txt):
        return Paragraph(txt, ps('TH', fontSize=8, fontName='Helvetica-Bold',
                                  textColor=colors.white, alignment=1))
    def td(txt):   return Paragraph(str(txt), cell_s)
    def tdr(txt):  return Paragraph(str(txt), cell_r)
    def tdb(txt):  return Paragraph(str(txt), cell_b)
    def tdrb(txt): return Paragraph(str(txt), cell_rb)
    def tdc(txt):  return Paragraph(str(txt), cell_c)

    def tbl(data, col_widths, total_last=False):
        t = Table(data, colWidths=col_widths, repeatRows=1)
        cmds = [
            ('BACKGROUND',    (0, 0), (-1, 0), NAVY),
            ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, LIGHT]),
            ('GRID',          (0, 0), (-1, -1), 0.4, BORDER),
            ('TOPPADDING',    (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING',   (0, 0), (-1, -1), 6),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 6),
        ]
        if total_last:
            cmds += [
                ('BACKGROUND', (0, -1), (-1, -1), TOTAL_BG),
                ('FONTNAME',   (0, -1), (-1, -1), 'Helvetica-Bold'),
            ]
        t.setStyle(TableStyle(cmds))
        return t

    story = []
    story.append(Paragraph('Cota 365 — Resumo Gerencial', title_s))
    story.append(Paragraph(
        f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}  |  {n_contratos} contratos', sub_s))
    story.append(HRFlowable(width='100%', thickness=1, color=BORDER, spaceAfter=10))

    story.append(Paragraph('Áreas', sec_s))
    story.append(tbl([[
        th('ÁREA PRIVATIVA'), th('ÁREA PRIV. ACESSÓRIA'), th('TOTAL ÁREA PRIVATIVA'),
        th('ÁREA COMUM'), th('ÁREA TOTAL'),
    ], [
        tdc(_fmt_m2(area_priv)), tdc(_fmt_m2(area_priv_acess)), tdc(_fmt_m2(total_priv)),
        tdc(_fmt_m2(area_comum)), tdc(_fmt_m2(area_total)),
    ]], [W/5]*5))
    story.append(Spacer(1, 6))

    vgv_liquido    = vgv_tabela - vgv_permuta
    total_real_a   = vgv_disponivel + vgv_reservada + total_vendido + vgv_permuta
    total_real_b   = vgv_disponivel + vgv_reservada + total_fluxo

    def _lbl(txt):
        return Paragraph(txt, ps('LB', fontSize=7, textColor=colors.HexColor('#6c757d'),
                                  fontName='Helvetica-Oblique', alignment=1))

    story.append(Paragraph('Indicadores Gerais', sec_s))
    kpi_table = Table([
        [th('VGV TOTAL'), th('VGV LÍQUIDO'), th('CONTRATOS'), th('TICKET MÉDIO'), th('VENDIDO')],
        [tdrb(_fmt_brl(vgv_tabela)), tdrb(_fmt_brl(vgv_liquido)),
         tdrb(str(n_contratos)),     tdrb(_fmt_brl(ticket_medio)), tdrb(_fmt_brl(total_fluxo))],
        [tdrb(_fmt_brl(total_real_a)), tdrb(_fmt_brl(total_real_b)),
         _lbl('—'), _lbl('—'), _lbl('—')],
    ], colWidths=[W/5]*5)
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4ff')]),
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#dee2e6')),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 6),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 6),
        ('BACKGROUND',    (0, 2), (-1, 2), colors.HexColor('#e8f0fe')),
        ('FONTNAME',      (0, 2), (-1, 2), 'Helvetica-Bold'),
    ]))
    story.append(kpi_table)

    legend_style = ps('LEG', fontSize=7, textColor=colors.HexColor('#6c757d'), spaceBefore=3, spaceAfter=6)
    story.append(Paragraph(
        '<b>Linha 1:</b> valores de tabela de vendas  |  <b>Linha 2:</b> valores reais de vendas (contratos)',
        legend_style,
    ))
    story.append(Spacer(1, 4))

    desconto_val = total_vendido - total_fluxo
    _, _, _, tot_desconto_cubs, tot_cubs, cub_atual, _, _, _, _ = _get_descontos_rows()
    desconto_cub_val = tot_cubs * cub_atual if cub_atual else 0
    tot_cubs_brl = _fmt_brl(desconto_cub_val) if cub_atual else '—'
    tot_cubs_fmt = f'{tot_cubs:.2f}'.replace('.', ',')
    desconto_pct = f'{desconto_cub_val / total_vendido * 100:.1f}%' if total_vendido else '0%'
    story.append(Paragraph('Descontos', sec_s))
    _dw = W / 5
    desc_table = Table([
        [th('VALOR TABELA'), th('VALOR CONTRATO'), th('DESCONTO CORRIGIDO PELO CUB'), th('% DESCONTO MÉDIO'), th('DESCONTOS EM CUBs')],
        [tdrb(_fmt_brl(total_vendido)), tdrb(_fmt_brl(total_fluxo)), tdrb(tot_cubs_brl), tdrb(desconto_pct), tdrb(tot_cubs_fmt)],
    ], colWidths=[_dw, _dw, _dw, _dw, _dw])
    desc_table.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white]),
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#dee2e6')),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 6),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 6),
    ]))
    story.append(desc_table)
    story.append(Spacer(1, 6))

    story.append(Paragraph('Resumo por Situação (Com permutas)', sec_s))
    sit_header = [[
        th('SITUAÇÃO'), th('VALOR TABELA'), th('% VALOR'),
        th('ÁREA PRIV.'), th('% ÁREA'), th('UNIDADES'), th('% UNID.'),
    ]]
    sit_rows = [
        [td(r['situacao']), tdr(r['vt_fmt']), tdr(r['pct_vt']),
         tdr(r['ap_fmt']),  tdr(r['pct_ap']), tdr(str(r['n'])), tdr(r['pct_n'])]
        for r in resumo_sit
    ]
    story.append(tbl(sit_header + sit_rows,
                     [2.7*cm, 3.5*cm, 1.8*cm, 3*cm, 1.8*cm, 1.98*cm, 1.8*cm], total_last=True))
    story.append(Spacer(1, 6))

    story.append(Paragraph('Resumo por Situação (Sem permutas)', sec_s))
    liq_header = [[
        th('SITUAÇÃO'), th('VALOR TABELA'), th('% VALOR'),
        th('ÁREA PRIV.'), th('% ÁREA'), th('UNIDADES'), th('% UNID.'),
    ]]
    liq_rows = [
        [td(r['situacao']), tdr(r['vt_fmt']), tdr(r['pct_vt']),
         tdr(r['ap_fmt']),  tdr(r['pct_ap']), tdr(str(r['n'])), tdr(r['pct_n'])]
        for r in resumo_sit_liquido
    ]
    story.append(tbl(liq_header + liq_rows,
                     [2.7*cm, 3.5*cm, 1.8*cm, 3*cm, 1.8*cm, 1.98*cm, 1.8*cm], total_last=True))
    story.append(PageBreak())

    # -- Resumo por Tipo (oculto) --
    # story.append(Paragraph('Resumo por Tipo', sec_s))
    # tip_header = [[th('QTDE'), th('TIPO'), th('M² PRIV.'), th('VALOR TABELA'), th('R$/M²'), th('PERM.')]]
    # tip_rows = [
    #     [tdr(str(r['n'])), td(r['tipo']), tdr(r['ap_fmt']), tdr(r['vt_fmt']), tdr(r['rsm2']), tdr(str(r['perm_n']))]
    #     for r in resumo_tip
    # ]
    # story.append(tbl(tip_header + tip_rows, [1.5*cm, 3.2*cm, 3.0*cm, 4.5*cm, 3.0*cm, 1.7*cm], total_last=True))
    # story.append(Paragraph(
    #     f'Preço médio por unidade: {preco_medio_tipo}',
    #     ps('PM', fontSize=8, textColor=NAVY, fontName='Helvetica-Bold', spaceBefore=4, spaceAfter=6),
    # ))

    # -- Resumo por Tipo (Estoque) (oculto) --
    # story.append(Spacer(1, 6))
    # story.append(Paragraph('Resumo por Tipo (Estoque)', sec_s))
    # est_header = [[th('QTDE'), th('TIPO'), th('M² PRIV.'), th('VALOR TABELA'), th('R$/M²')]]
    # est_rows = [
    #     [tdr(str(r['n'])), td(r['tipo']), tdr(r['ap_fmt']), tdr(r['vt_fmt']), tdr(r['rsm2'])]
    #     for r in resumo_tip_estoque
    # ]
    # story.append(tbl(est_header + est_rows, [1.5*cm, 3.5*cm, 3.0*cm, 4.5*cm, 4.4*cm], total_last=True))
    # story.append(Paragraph(
    #     f'Preço médio por unidade (estoque): {preco_medio_estoque}',
    #     ps('PM', fontSize=8, textColor=NAVY, fontName='Helvetica-Bold', spaceBefore=4, spaceAfter=6),
    # ))

    # ── Resumo por Tipo (Todos) ───────────────────────────────────────────────
    story.append(Spacer(1, 4))
    story.append(Paragraph('Resumo por Tipo (Todos)', sec_s))

    SUBHDR_BG = colors.HexColor('#495057')

    def _fmt_ap2(v):
        return f'{v:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')

    def _fmt_pct(v):
        return f'{v:.1f}'.replace('.', ',')

    COL_W = [
        1.5*cm,   # Tipo
        0.8*cm,   # tot_n
        1.5*cm,   # tot_ap
        2.3*cm,   # tot_vt
        0.8*cm,   # vnd_n
        1.5*cm,   # vnd_ap
        2.3*cm,   # vnd_vt
        0.85*cm,  # pct
        0.8*cm,   # est_n
        1.5*cm,   # est_ap
        2.3*cm,   # est_vt
        0.75*cm,  # perm
    ]

    tot_hdr = [
        th('Tipo'),
        th('Total'), th(''), th(''),
        th('Vendido'), th(''), th(''), th(''),
        th('Estoque'), th(''), th(''),
        th('P'),
    ]
    sub_hdr = [
        th(''),
        th('Qt.'), th('A.Priv'), th('Valor (R$)'),
        th('Qt.'), th('A.Priv'), th('Valor (R$)'), th('%'),
        th('Qt.'), th('A.Priv'), th('Valor (R$)'),
        th(''),
    ]

    def _row_tot(r):
        _b = r.get('is_total', False)
        _f = tdrb if _b else tdr
        _fl = tdb if _b else td
        return [
            _fl(r['tipo']),
            _f(str(r['tot_n'])),
            _f(_fmt_ap2(r['tot_ap'])),
            _f(_fmt_ap2(r['tot_vt'])),
            _f(str(r['vnd_n'])),
            _f(_fmt_ap2(r['vnd_ap'])),
            _f(_fmt_ap2(r['vnd_vt'])),
            _f(_fmt_pct(r['pct'])),
            _f(str(r['est_n'])),
            _f(_fmt_ap2(r['est_ap'])),
            _f(_fmt_ap2(r['est_vt'])),
            _f(str(r['perm_n'])),
        ]

    tot_data = [tot_hdr, sub_hdr] + [_row_tot(r) for r in resumo_tip_total]

    t_tot = Table(tot_data, colWidths=COL_W, repeatRows=2)
    n_tot_data = len(tot_data)
    t_tot.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0), NAVY),
        ('BACKGROUND',    (0, 1), (-1, 1), SUBHDR_BG),
        ('ROWBACKGROUNDS',(0, 2), (-1, n_tot_data - 2), [colors.white, LIGHT]),
        ('BACKGROUND',    (0, -1), (-1, -1), TOTAL_BG),
        ('FONTNAME',      (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID',          (0, 0), (-1, -1), 0.4, BORDER),
        ('TOPPADDING',    (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING',   (0, 0), (-1, -1), 4),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
        ('VALIGN',        (0, 0), (-1, 1), 'MIDDLE'),
        ('SPAN',          (0, 0), (0, 1)),    # Tipo abrange as 2 linhas de header
        ('SPAN',          (1, 0), (3, 0)),    # Total
        ('SPAN',          (4, 0), (7, 0)),    # Vendido
        ('SPAN',          (8, 0), (10, 0)),   # Estoque
        ('SPAN',          (11, 0), (11, 1)),  # P abrange as 2 linhas de header
    ]))
    story.append(t_tot)
    story.append(Spacer(1, 6))

    # ── Preço Médio ───────────────────────────────────────────────────────────
    story.append(Paragraph('Preço Médio', sec_s))

    def _rsm2(vt, ap):
        return _fmt_ap2(vt / ap) if ap else '—'

    def _pm(vt, n):
        return _fmt_ap2(vt / n) if n else '—'

    PM_COL_W = [1.5*cm, 2.57*cm, 2.57*cm, 2.57*cm, 2.57*cm, 2.57*cm, 2.55*cm]

    pm_hdr0 = [
        th('Tipo'),
        th('Tabela'), th(''),
        th('Vendido'), th(''),
        th('Estoque'), th(''),
    ]
    pm_hdr1 = [
        th(''),
        th('R$/m2'), th('Preço médio'),
        th('R$/m2'), th('Preço médio'),
        th('R$/m2'), th('Preço médio'),
    ]

    def _row_pm(r):
        _b = r.get('is_total', False)
        _f = tdrb if _b else tdr
        _fl = tdb if _b else td
        return [
            _fl(r['tipo']),
            _f(_rsm2(r['tot_vt'], r['tot_ap'])),
            _f(_pm(r['tot_vt'],   r['tot_n'])),
            _f(_rsm2(r['vnd_vt'], r['vnd_ap'])),
            _f(_pm(r['vnd_vt'],   r['vnd_n'])),
            _f(_rsm2(r['est_vt'], r['est_ap'])),
            _f(_pm(r['est_vt'],   r['est_n'])),
        ]

    pm_data = [pm_hdr0, pm_hdr1] + [_row_pm(r) for r in resumo_tip_total]
    n_pm = len(pm_data)

    t_pm = Table(pm_data, colWidths=PM_COL_W, repeatRows=2)
    t_pm.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0), NAVY),
        ('BACKGROUND',    (0, 1), (-1, 1), SUBHDR_BG),
        ('ROWBACKGROUNDS',(0, 2), (-1, n_pm - 2), [colors.white, LIGHT]),
        ('BACKGROUND',    (0, -1), (-1, -1), TOTAL_BG),
        ('FONTNAME',      (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID',          (0, 0), (-1, -1), 0.4, BORDER),
        ('TOPPADDING',    (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING',   (0, 0), (-1, -1), 4),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
        ('VALIGN',        (0, 0), (-1, 1), 'MIDDLE'),
        ('SPAN',          (0, 0), (0, 1)),   # Tipo
        ('SPAN',          (1, 0), (2, 0)),   # Tabela
        ('SPAN',          (3, 0), (4, 0)),   # Vendido
        ('SPAN',          (5, 0), (6, 0)),   # Estoque
    ]))
    story.append(t_pm)
    story.append(Spacer(1, 6))

    story.append(Paragraph('Receita por Ano', sec_s))
    ano_header = [[th('ANO'), th('RECEITA PREVISTA'), th('% DO TOTAL')]]
    ano_rows = [
        [td(ano), tdr(_fmt_brl(val)),
         tdr(f'{val/total_fluxo*100:.1f}%' if total_fluxo else '0%')]
        for ano, val in sorted(ano_totals.items())
    ]
    ano_rows.append([tdb('TOTAL'), tdrb(_fmt_brl(total_fluxo)), tdrb('100%')])
    story.append(tbl(ano_header + ano_rows, [3*cm, 7*cm, 4*cm], total_last=True))
    story.append(Spacer(1, 6))

    story.append(Paragraph('Fluxo de Caixa Mensal', sec_s))
    fm_header = [[th('MÊS'), th('RECEBIDO'), th('A RECEBER'), th('TOTAL'), th('ACUMULADO')]]
    fm_rows = []
    acumulado = 0.0
    total_rec_acc  = 0.0
    total_pend_acc = 0.0
    for key in exp_all_keys:
        yr, mo  = key
        rec     = exp_monthly_rec[key]
        pend    = exp_monthly_pend[key]
        total   = rec + pend
        acumulado      += total
        total_rec_acc  += rec
        total_pend_acc += pend
        fm_rows.append([
            td(f'{mo:02d}/{yr}'),
            tdr(_fmt_brl(rec)),
            tdr(_fmt_brl(pend)),
            tdr(_fmt_brl(total)),
            tdr(_fmt_brl(acumulado)),
        ])
    fm_rows.append([
        tdb('TOTAL'),
        tdrb(_fmt_brl(total_rec_acc)),
        tdrb(_fmt_brl(total_pend_acc)),
        tdrb(_fmt_brl(total_fluxo)),
        tdrb(''),
    ])
    story.append(tbl(fm_header + fm_rows, [2.5*cm, 4*cm, 4*cm, 4*cm, 4*cm], total_last=True))

    # ── Resumo por Tipo (igual ao Fluxo Mensal) ───────────────────────────────
    if total_fluxo:
        story.append(Paragraph('Resumo por Tipo de Parcela', sec_s))

        NAVY   = colors.HexColor('#1a1a2e')
        POUPA  = colors.HexColor('#fffde7')
        FIN    = colors.HexColor('#fff3e0')
        LIGHT  = colors.HexColor('#f8f9fa')
        BORDER = colors.HexColor('#dee2e6')

        def pct_fmt(v):
            return f'{v / total_fluxo * 100:.2f}%'.replace('.', ',') if total_fluxo else '0,00%'

        TIPO_LABELS = [
            ('AT', 'Ato'),
            ('PM', 'Mensais'),
            ('RA', 'Ref. Anuais'),
            ('PE', 'Permuta'),
            ('CH', 'Chaves'),
        ]

        def sr(label, val, bold=False, bg=None):
            fn = 'Helvetica-Bold' if bold else 'Helvetica'
            return (
                Paragraph(f'<b>{label}</b>' if bold else label,
                          ParagraphStyle('sl', fontName=fn, fontSize=8)),
                Paragraph(f'<b>{_fmt_num(val)}</b>' if bold else _fmt_num(val),
                          ParagraphStyle('sv', fontName=fn, fontSize=8, alignment=2)),
                Paragraph(f'<b>{pct_fmt(val)}</b>' if bold else pct_fmt(val),
                          ParagraphStyle('sp', fontName=fn, fontSize=8, alignment=2)),
            )

        def shdr(txt):
            return Paragraph(f'<b><font color="white">{txt}</font></b>',
                             ParagraphStyle('sh', fontName='Helvetica-Bold', fontSize=8, alignment=1))

        sum_data = [[shdr('Tipo'), shdr('Total'), shdr('%')]]
        for code, label in TIPO_LABELS:
            sum_data.append(list(sr(label, tipo_totals.get(code, 0))))
        sum_data.append(list(sr('Poupança',      poupanca_total, bold=True)))
        sum_data.append(list(sr('Financiamento', fi_total,       bold=True)))
        sum_data.append(list(sr('Total',         total_fluxo,    bold=True)))
        sum_data.append(list(sr('Recebido',      total_recebido)))
        sum_data.append(list(sr('A receber',     total_a_receber)))

        # índices no sum_data (0 = header)
        IDX_POUPANCA = 6   # linha Poupança
        IDX_FI       = 7   # linha Financiamento
        IDX_TOTAL    = 8   # linha Total

        sum_table = Table(sum_data, colWidths=[4.5*cm, 4.5*cm, 2.5*cm])
        sum_table.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0),           (-1, 0),           NAVY),
            ('VALIGN',        (0, 0),           (-1, -1),          'MIDDLE'),
            ('ALIGN',         (1, 1),           (-1, -1),          'RIGHT'),
            ('ROWBACKGROUNDS',(0, 1),           (-1, IDX_POUPANCA - 1), [colors.white, LIGHT]),
            ('BACKGROUND',    (0, IDX_POUPANCA),(-1, IDX_POUPANCA),POUPA),
            ('BACKGROUND',    (0, IDX_FI),      (-1, IDX_FI),      FIN),
            ('LINEABOVE',     (0, IDX_TOTAL),   (-1, IDX_TOTAL),   1, NAVY),
            ('GRID',          (0, 0),           (-1, -1),          0.4, BORDER),
            ('TOPPADDING',    (0, 0),           (-1, -1),          4),
            ('BOTTOMPADDING', (0, 0),           (-1, -1),          4),
            ('LEFTPADDING',   (0, 0),           (-1, -1),          4),
            ('RIGHTPADDING',  (0, 0),           (-1, -1),          4),
        ]))
        story.append(sum_table)

    # ── Ranking por imobiliária ───────────────────────────────────────────────
    venda_por_imob_pdf = {
        r['imobiliaria']: {'cnt': r['cnt'], 'vgv': r['vgv'] or 0.0}
        for r in Venda.objects
            .exclude(imobiliaria='')
            .values('imobiliaria')
            .annotate(cnt=Count('id'), vgv=Sum('valor_contrato'))
    }
    com_por_imob = {
        r['imobiliaria']: r['total'] or 0.0
        for r in Comissao.objects
            .exclude(imobiliaria='')
            .values('imobiliaria')
            .annotate(total=Sum('valor_comissao'))
    }
    all_imobs = set(venda_por_imob_pdf)
    total_vgv_imob = sum(v['vgv'] for v in venda_por_imob_pdf.values()) or 1.0
    ranking_imob = sorted(
        [
            {
                'imobiliaria': imob,
                'n_vendas':    venda_por_imob_pdf[imob]['cnt'],
                'vgv':         venda_por_imob_pdf[imob]['vgv'],
                'pct':         f"{venda_por_imob_pdf[imob]['vgv'] / total_vgv_imob * 100:.1f}%",
                'com':         com_por_imob.get(imob, 0.0),
                'com_pct':     f"{com_por_imob.get(imob, 0.0) / venda_por_imob_pdf[imob]['vgv'] * 100:.1f}%" if venda_por_imob_pdf[imob]['vgv'] else '—',
            }
            for imob in all_imobs
        ],
        key=lambda x: x['vgv'],
        reverse=True,
    )
    _tot_vgv = sum(r['vgv'] for r in ranking_imob)
    _tot_com = sum(r['com'] for r in ranking_imob)

    if ranking_imob:
        story.append(PageBreak())
        story.append(Paragraph('Ranking de Vendas por Imobiliária', sec_s))
        rank_header = [[th('IMOBILIÁRIA'), th('VENDAS'), th('VGV'), th('%'), th('COMISSÕES'), th('% COM.')]]
        rank_rows = [
            [td(r['imobiliaria']), tdr(str(r['n_vendas'])),
             tdrb(_fmt_brl(r['vgv'])), tdr(r['pct']),
             tdr(_fmt_brl(r['com'])), tdr(r['com_pct'])]
            for r in ranking_imob
        ]
        _tot_com_pct = f"{_tot_com / _tot_vgv * 100:.1f}%" if _tot_vgv else '—'
        rank_rows.append([
            tdb('TOTAL'),
            tdrb(str(sum(r['n_vendas'] for r in ranking_imob))),
            tdrb(_fmt_brl(_tot_vgv)),
            tdrb('100%'),
            tdrb(_fmt_brl(_tot_com)),
            tdrb(_tot_com_pct),
        ])
        story.append(tbl(rank_header + rank_rows,
                         [6.5*cm, 1.8*cm, 3.5*cm, 1.8*cm, 3.5*cm, 1.8*cm], total_last=True))

    doc.build(story)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/pdf')
    resp['Content-Disposition'] = 'attachment; filename="resumo_cota365.pdf"'
    return resp


def unidades(request):
    headers, rows = _load_unidades()

    status_col = next((i for i, h in enumerate(headers) if h.lower() == 'status'), None)
    tabela_col = next((i for i, h in enumerate(headers) if 'tabela' in h.lower()), None)
    venda_col  = next((i for i, h in enumerate(headers) if 'venda'  in h.lower()), None)

    def _col(name):
        return next((i for i, h in enumerate(headers)
                     if h.lower().replace('.', '').replace(' ', '') ==
                     name.lower().replace('.', '').replace(' ', '')), None)

    ap_col  = _col('Area Privativa')
    apa_col = _col('Area Priv Acessoria')
    ac_col  = _col('Area Comum')

    def _area(row, idx):
        if idx is None: return 0.0
        try: return float(row[idx]) if row[idx] not in ('', None) else 0.0
        except (ValueError, TypeError): return 0.0

    total       = len(rows)
    vendidos    = sum(1 for r in rows if status_col is not None and str(r[status_col]).lower() == 'vendido')
    disponiveis = sum(1 for r in rows if status_col is not None and str(r[status_col]).lower() == 'disponível')
    permutas_n  = sum(1 for r in rows if status_col is not None and str(r[status_col]).lower() == 'permuta')
    reservados  = total - vendidos - disponiveis - permutas_n

    def _to_float(v):
        try: return float(v)
        except (ValueError, TypeError): return 0.0

    vgv_vendido = sum(
        _to_float(r[venda_col]) for r in rows
        if venda_col is not None and status_col is not None
        and str(r[status_col]).lower() == 'vendido'
    ) if venda_col is not None else 0

    vgv_disponivel = sum(
        _to_float(r[tabela_col]) for r in rows
        if tabela_col is not None and status_col is not None
        and str(r[status_col]).lower() == 'disponível'
    ) if tabela_col is not None else 0

    area_priv       = sum(_area(r, ap_col)  for r in rows)
    area_priv_acess = sum(_area(r, apa_col) for r in rows)
    area_comum      = sum(_area(r, ac_col)  for r in rows)
    total_priv      = area_priv + area_priv_acess
    area_total      = total_priv + area_comum

    STATUS_CSS = {
        'vendido':    'badge-vendido',
        'disponível': 'badge-disponivel',
        'reservado':  'badge-reservado',
        'permuta':    'badge-permuta',
    }

    def make_cells(row):
        cells = []
        for i, v in enumerate(row):
            h = headers[i].lower()
            if ('tabela' in h or 'venda' in h) and v not in ('', 0, None):
                cells.append({'val': _fmt_brl(float(v)), 'cls': 'text-end fw-semibold', 'badge': ''})
            elif h == 'status':
                badge = STATUS_CSS.get(str(v).lower(), '')
                cells.append({'val': v, 'cls': 'text-center', 'badge': badge})
            else:
                cells.append({'val': v, 'cls': '', 'badge': ''})
        return cells

    context = {
        'headers':        headers,
        'rows':           [make_cells(r) for r in rows],
        'total':          total,
        'vendidos':       vendidos,
        'disponiveis':    disponiveis,
        'reservados':     reservados,
        'permutas_n':     permutas_n,
        'vgv_vendido':    _fmt_brl(vgv_vendido),
        'vgv_disponivel': _fmt_brl(vgv_disponivel),
        'area_priv':      _fmt_m2(area_priv),
        'area_priv_acess': _fmt_m2(area_priv_acess),
        'total_priv':     _fmt_m2(total_priv),
        'area_comum':     _fmt_m2(area_comum),
        'area_total':     _fmt_m2(area_total),
    }
    return render(request, 'cota365/unidades.html', context)


def export_unidades(request):
    fmt = request.GET.get('format', 'xlsx')
    headers, rows = _load_unidades()
    price_cols = {i for i, h in enumerate(headers) if 'tabela' in h.lower() or 'venda' in h.lower()}
    if fmt == 'pdf':
        return _export_unidades_pdf(headers, rows, price_cols)
    return _export_unidades_xlsx(headers, rows, price_cols)


def vendas(request):
    sort     = request.GET.get('sort', 'cliente')
    sort_dir = request.GET.get('dir', 'asc')

    vendas_rows = _load_vendas()
    fluxo_rows  = _load_fluxo()
    fluxo_by_id = {r['id']: r for r in fluxo_rows}
    vinculos    = _load_vinculos()

    contracts = []
    total_geral = 0.0
    for v in vendas_rows:
        f    = fluxo_by_id.get(v['numero'], {})
        valor = f.get('vgv', 0) or f.get('pv', 0)
        vinc  = vinculos.get(v['unidade'], {})
        contracts.append({
            'numero':    f"#{v['numero']}",
            'cliente':   v['cliente'],
            'unidade':   v['unidade'],
            'valor':     valor,
            'valor_fmt': _fmt_brl(valor),
            'garagens':  vinc.get('garagens', ''),
            'hb':        vinc.get('hb', ''),
        })
        total_geral += valor

    _SORT_KEYS = {
        'numero':   lambda x: int(x['numero'].lstrip('#')) if x['numero'].lstrip('#').isdigit() else 0,
        'cliente':  lambda x: x['cliente'].lower(),
        'unidade':  lambda x: x['unidade'],
        'garagens': lambda x: x['garagens'],
        'hb':       lambda x: x['hb'],
    }
    contracts.sort(key=_SORT_KEYS.get(sort, _SORT_KEYS['cliente']), reverse=(sort_dir == 'desc'))
    context = {
        'contracts':   contracts,
        'total_geral': _fmt_brl(total_geral),
        'n_contratos': len(contracts),
        'sort':        sort,
        'sort_dir':    sort_dir,
        'colums': [
            ('numero',   'Nº'),
            ('cliente',  'CLIENTE'),
            ('unidade',  'UNIDADE'),
            ('garagens', 'GARAGENS'),
            ('hb',       'HB'),
        ],
    }
    return render(request, 'cota365/vendas.html', context)


def comparativo_valores(request):
    raw_rows, tot_tab_mes, tot_contrato, tot_desconto, tot_cubs, cub_atual, tot_tab_mes_all, tot_contrato_all, pct_resumo, latest_comp = _get_descontos_rows()

    def _pct_fmt(v):
        if v is None:
            return None
        sinal = '+' if v > 0 else ''
        return f'{sinal}{v:.2f}%'.replace('.', ',')

    rows = [{
        **r,
        'vtm_fmt':            _fmt_brl(r['vtm']) if r['vtm'] else None,
        'valor_contrato_fmt': _fmt_brl(r['valor_contrato']),
        'desconto_fmt':       _fmt_brl(r['desconto']),
        'pct_mes_fmt':        _pct_fmt(r['pct_mes']),
        'pct_mes_neg':        r['pct_mes'] is not None and r['pct_mes'] < 0,
        'qtd_cubs_fmt':       f'{r["qtd_cubs"]:.2f}'.replace('.', ',') if r['qtd_cubs'] is not None else None,
    } for r in raw_rows]

    sort     = request.GET.get('sort', 'cliente')
    sort_dir = request.GET.get('dir', 'asc')
    _SORTS = {
        'numero':         lambda x: int(x['numero']) if x['numero'].isdigit() else 0,
        'unidade':        lambda x: x['unidade'],
        'cliente':        lambda x: x['cliente'].lower(),
        'data_venda':     lambda x: x['data_venda'],
        'valor_contrato': lambda x: x['valor_contrato'],
        'desconto':       lambda x: (x['desconto'] or 0),
        'pct_mes':        lambda x: (x['pct_mes'] or 0),
    }
    rows.sort(key=_SORTS.get(sort, _SORTS['cliente']), reverse=(sort_dir == 'desc'))

    return render(request, 'cota365/descontos.html', {
        'rows':           rows,
        'n':              len(rows),
        'latest_comp':    latest_comp.strftime('%m/%Y') if latest_comp else '—',
        'tot_contrato':   _fmt_brl(tot_contrato),
        'tot_tab_mes':    _fmt_brl(tot_tab_mes),
        'tot_desconto':   _fmt_brl(tot_desconto),
        'tot_cubs':         f'{tot_cubs:.2f}'.replace('.', ','),
        'tot_cubs_brl':     _fmt_brl(tot_cubs * cub_atual) if cub_atual else None,
        'cub_atual_fmt':    _fmt_brl(cub_atual) if cub_atual else '—',
        'tot_pct_cubs':     _pct_fmt((tot_cubs * cub_atual) / tot_tab_mes * 100) if (cub_atual and tot_tab_mes) else None,
        'resumo_tab_all':   _fmt_brl(tot_tab_mes_all),
        'resumo_cont_all':  _fmt_brl(tot_contrato_all),
        'resumo_desconto':  _fmt_brl(tot_cubs * cub_atual) if cub_atual else _fmt_brl(tot_desconto),
        'resumo_pct':       _pct_fmt((tot_cubs * cub_atual) / tot_tab_mes_all * 100) if (cub_atual and tot_tab_mes_all) else _pct_fmt(pct_resumo),
        'resumo_pct_neg':   (tot_cubs * cub_atual) / tot_tab_mes_all * 100 < 0 if (cub_atual and tot_tab_mes_all) else (pct_resumo is not None and pct_resumo < 0),
        'sort':           sort,
        'sort_dir':       sort_dir,
        'columns': [
            ('numero',         'RESERVA'),
            ('unidade',        'UNIDADE'),
            ('cliente',        'CLIENTE'),
            ('data_venda',     'DATA VENDA'),
            ('vtm',            'TAB. MÊS VENDA'),
            ('valor_contrato', 'VALOR CONTRATO'),
            ('desconto',       'DESCONTO'),
            ('pct_mes',        'Δ%'),
            ('qtd_cubs',       'QTD CUBs'),
        ],
    })


def _get_descontos_rows():
    tabelas = defaultdict(dict)
    for t in Tabela.objects.all():
        tabelas[t.competencia][t.unidade] = t.valor_total

    latest_comp = max(tabelas.keys(), default=None)

    cub_map = {
        d.data: float(d.valor)
        for d in IndiceData.objects.filter(indice_id=1)
    }

    rows = []
    tot_contrato = tot_tab_mes = 0.0
    tot_tab_mes_all = tot_contrato_all = 0.0
    for v in Venda.objects.filter(valor_contrato__gt=0).order_by('cliente'):
        comp_mes = v.data_venda.replace(day=1) if v.data_venda else None
        vtm      = tabelas.get(comp_mes, {}).get(v.unidade) if comp_mes else None
        vc       = v.valor_contrato
        tot_contrato_all += vc
        if vtm:
            tot_tab_mes_all += vtm
        desconto = (vtm - vc) if vtm else None
        if not desconto:
            continue
        pct_mes  = desconto / vtm * 100 if (vtm and vtm > 0) else None
        cub_val  = cub_map.get(comp_mes) if comp_mes else None
        qtd_cubs = desconto / cub_val if (cub_val and cub_val > 0) else None
        rows.append({
            'numero':             v.numero,
            'unidade':            v.unidade,
            'cliente':            v.cliente,
            'data_venda':         v.data_venda.strftime('%d/%m/%Y') if v.data_venda else '—',
            'comp_mes':           comp_mes.strftime('%m/%Y') if comp_mes else '—',
            'vtm':                vtm,
            'valor_contrato':     vc,
            'desconto':           desconto,
            'pct_mes':            pct_mes,
            'qtd_cubs':           qtd_cubs,
        })
        tot_contrato += vc
        tot_tab_mes  += vtm or 0

    tot_desconto  = sum(r['desconto'] for r in rows)
    tot_cubs      = sum(r['qtd_cubs'] for r in rows if r['qtd_cubs'] is not None)
    cub_atual_obj = IndiceData.objects.filter(indice_id=1).order_by('-data').first()
    cub_atual     = float(cub_atual_obj.valor) if cub_atual_obj else None
    pct_resumo    = tot_desconto / tot_tab_mes_all * 100 if tot_tab_mes_all else None
    return rows, tot_tab_mes, tot_contrato, tot_desconto, tot_cubs, cub_atual, tot_tab_mes_all, tot_contrato_all, pct_resumo, latest_comp


def _export_descontos_pdf(rows, tot_tab_mes, tot_contrato, tot_desconto, tot_cubs, cub_atual, tot_tab_mes_all, tot_contrato_all, pct_resumo, latest_comp):
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.platypus import HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm

    NAVY  = colors.HexColor('#1a1a2e')
    GRAY1 = colors.HexColor('#f8f9fa')
    RED   = colors.HexColor('#dc3545')
    GREEN = colors.HexColor('#198754')
    GRID  = colors.HexColor('#dee2e6')
    TOTAL = colors.HexColor('#e9ecef')

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    lc_str = latest_comp.strftime('%m/%Y') if latest_comp else '—'

    story = [
        Paragraph('Cota 365 — Descontos',
                  ParagraphStyle('title', parent=styles['Heading1'], fontSize=14, spaceAfter=4)),
        Paragraph(
            f'Tab. Mês Venda: {_fmt_brl(tot_tab_mes)}   |   '
            f'Total Contratos: {_fmt_brl(tot_contrato)}   |   '
            f'Total Descontos: {_fmt_brl(tot_desconto)}   |   '
            f'{len(rows)} contratos   |   Tabela vigente: {lc_str}',
            ParagraphStyle('sub', parent=styles['Normal'], fontSize=9, spaceAfter=12)),
    ]

    header = ['RESERVA', 'UNIDADE', 'CLIENTE', 'DATA VENDA', 'TAB. MÊS', 'VALOR CONTRATO', 'DESCONTO', 'Δ%', 'QTD CUBs']
    data = [header]
    for r in rows:
        pct_str = ''
        if r['pct_mes'] is not None:
            sinal = '+' if r['pct_mes'] > 0 else ''
            pct_str = f"{sinal}{r['pct_mes']:.2f}%".replace('.', ',')
        cubs_str = f"{r['qtd_cubs']:.2f}".replace('.', ',') if r['qtd_cubs'] is not None else '—'
        data.append([
            f"#{r['numero']}",
            r['unidade'],
            r['cliente'],
            r['data_venda'],
            _fmt_brl(r['vtm']) if r['vtm'] else '—',
            _fmt_brl(r['valor_contrato']),
            _fmt_brl(r['desconto']),
            pct_str,
            cubs_str,
        ])
    tot_cubs_str = f'{tot_cubs:.2f}'.replace('.', ',')
    tot_cubs_brl = _fmt_brl(tot_cubs * cub_atual) if cub_atual else '—'
    cubs_brl_val = tot_cubs * cub_atual if cub_atual else None
    if cubs_brl_val and tot_tab_mes:
        pct_val = cubs_brl_val / tot_tab_mes * 100
        sinal = '+' if pct_val > 0 else ''
        tot_pct_cubs_str = f'{sinal}{pct_val:.2f}%'.replace('.', ',')
    else:
        tot_pct_cubs_str = ''
    data.append(['', '', '', '', _fmt_brl(tot_tab_mes), _fmt_brl(tot_contrato), _fmt_brl(tot_desconto), tot_pct_cubs_str, f'{tot_cubs_str}\n{tot_cubs_brl}'])

    col_widths = [1.6*cm, 2.2*cm, 5.8*cm, 2.4*cm, 3.0*cm, 3.2*cm, 3.2*cm, 1.6*cm, 2.0*cm]
    t = Table(data, colWidths=col_widths, repeatRows=1)

    num_rows = len(data)
    styles_ts = [
        ('BACKGROUND',    (0, 0),  (-1, 0),       NAVY),
        ('TEXTCOLOR',     (0, 0),  (-1, 0),       colors.white),
        ('FONTNAME',      (0, 0),  (-1, 0),       'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0),  (-1, -1),      7),
        ('ALIGN',         (0, 0),  (-1, -1),      'CENTER'),
        ('ALIGN',         (2, 1),  (2, num_rows-2), 'LEFT'),
        ('ALIGN',         (4, 1),  (-1, -1),      'RIGHT'),
        ('ROWBACKGROUNDS',(0, 1),  (-1, num_rows-2), [colors.white, GRAY1]),
        ('BACKGROUND',    (0, -1), (-1, -1),      TOTAL),
        ('FONTNAME',      (0, -1), (-1, -1),      'Helvetica-Bold'),
        ('GRID',          (0, 0),  (-1, -1),      0.5, GRID),
        ('TOPPADDING',    (0, 0),  (-1, -1),      3),
        ('BOTTOMPADDING', (0, 0),  (-1, -1),      3),
    ]
    # colorir descontos positivos em verde, negativos em vermelho
    for i, r in enumerate(rows, 1):
        cor = RED if (r['desconto'] or 0) < 0 else GREEN
        styles_ts.append(('TEXTCOLOR', (6, i), (6, i), cor))
        styles_ts.append(('TEXTCOLOR', (7, i), (7, i), cor))

    t.setStyle(TableStyle(styles_ts))
    story.append(t)

    from reportlab.platypus import Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    _styles = getSampleStyleSheet()
    cubs_brl_resumo = tot_cubs * cub_atual if cub_atual else None
    desc_resumo_val = cubs_brl_resumo if cubs_brl_resumo is not None else tot_desconto
    pct_resumo_val  = desc_resumo_val / tot_tab_mes_all * 100 if tot_tab_mes_all else pct_resumo
    pct_str = ''
    if pct_resumo_val is not None:
        sinal = '+' if pct_resumo_val > 0 else ''
        pct_str = f'{sinal}{pct_resumo_val:.2f}%'.replace('.', ',')
    story.append(Spacer(1, 10))
    resumo_data = [
        [Paragraph('RESUMO GERAL', ParagraphStyle('rh', parent=_styles['Normal'],
                   fontSize=7, fontName='Helvetica-Bold', textColor=colors.white, alignment=1))],
        [Table([
            [Paragraph('Total (Tabela vendas)', ParagraphStyle('rl', parent=_styles['Normal'], fontSize=7, textColor=colors.HexColor('#6c757d'))),
             Paragraph(_fmt_brl(tot_tab_mes_all), ParagraphStyle('rv', parent=_styles['Normal'], fontSize=8, fontName='Helvetica-Bold', alignment=2))],
            [Paragraph('Total (Contratos)', ParagraphStyle('rl2', parent=_styles['Normal'], fontSize=7, textColor=colors.HexColor('#6c757d'))),
             Paragraph(_fmt_brl(tot_contrato_all), ParagraphStyle('rv2', parent=_styles['Normal'], fontSize=8, fontName='Helvetica-Bold', alignment=2))],
            [Paragraph('Descontos (corrigido CUB)', ParagraphStyle('rl3', parent=_styles['Normal'], fontSize=7, textColor=colors.HexColor('#6c757d'))),
             Paragraph(_fmt_brl(desc_resumo_val), ParagraphStyle('rv3', parent=_styles['Normal'], fontSize=8, fontName='Helvetica-Bold', alignment=2))],
            [Paragraph('Percentual', ParagraphStyle('rl4', parent=_styles['Normal'], fontSize=7, textColor=colors.HexColor('#6c757d'))),
             Paragraph(pct_str, ParagraphStyle('rv4', parent=_styles['Normal'], fontSize=8, fontName='Helvetica-Bold', alignment=2,
                       textColor=(RED if (pct_resumo_val or 0) < 0 else GREEN)))],
        ], colWidths=[4*cm, 4*cm])],
    ]
    resumo_tbl = Table(resumo_data, colWidths=[8*cm])
    resumo_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0), NAVY),
        ('BACKGROUND',    (0, 1), (-1, 1), GRAY1),
        ('BOX',           (0, 0), (-1, -1), 0.5, GRID),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 6),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 6),
    ]))
    story.append(resumo_tbl)

    doc.build(story)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/pdf')
    resp['Content-Disposition'] = 'attachment; filename="descontos.pdf"'
    return resp


def export_descontos(request):
    rows, tot_tab_mes, tot_contrato, tot_desconto, tot_cubs, cub_atual, tot_tab_mes_all, tot_contrato_all, pct_resumo, latest_comp = _get_descontos_rows()
    return _export_descontos_pdf(rows, tot_tab_mes, tot_contrato, tot_desconto, tot_cubs, cub_atual, tot_tab_mes_all, tot_contrato_all, pct_resumo, latest_comp)


def fluxo_mensal(request):
    TIPOS = ['AT', 'PM', 'RA', 'PE', 'CH']
    TIPO_LABELS = {'AT': 'Ato', 'PM': 'Mensais', 'RA': 'Ref. Anuais',
                   'PE': 'Permuta', 'CH': 'Chaves'}

    _base = Parcela.objects.filter(vencimento__isnull=False)

    monthly = defaultdict(lambda: defaultdict(float))
    for r in (_base
              .annotate(ano=ExtractYear('vencimento'), mes=ExtractMonth('vencimento'))
              .values('ano', 'mes', 'tipo')
              .annotate(total=Sum('valor'))
              .order_by('ano', 'mes')):
        monthly[(r['ano'], r['mes'])][r['tipo']] = r['total'] or 0.0

    monthly_rec  = {}
    monthly_pend = {}
    for r in (Parcela.objects.filter(data_pagamento__isnull=False)
              .annotate(ano=ExtractYear('data_pagamento'), mes=ExtractMonth('data_pagamento'))
              .values('ano', 'mes')
              .annotate(rec=Sum('valor'))
              .order_by('ano', 'mes')):
        monthly_rec[(r['ano'], r['mes'])] = r['rec'] or 0.0

    for r in (_base.filter(data_pagamento__isnull=True)
              .annotate(ano=ExtractYear('vencimento'), mes=ExtractMonth('vencimento'))
              .values('ano', 'mes')
              .annotate(pend=Sum('valor'))
              .order_by('ano', 'mes')):
        monthly_pend[(r['ano'], r['mes'])] = r['pend'] or 0.0

    totals        = defaultdict(float)
    total_rec     = 0.0
    total_pend    = 0.0
    rows = []

    for key in sorted(monthly.keys()):
        yr, mo = key
        td = monthly[key]
        at = td.get('AT', 0); pm = td.get('PM', 0); ra = td.get('RA', 0)
        pe = td.get('PE', 0); ch = td.get('CH', 0); fi = td.get('FI', 0)
        poupanca  = at + pm + ra + pe + ch
        row_total = poupanca + fi
        rec   = monthly_rec.get(key, 0.0)
        pend  = monthly_pend.get(key, 0.0)
        for t, v in (('AT', at), ('PM', pm), ('RA', ra), ('PE', pe), ('CH', ch), ('FI', fi)):
            totals[t] += v
        totals['poupanca'] += poupanca
        totals['total']    += row_total
        total_rec  += rec
        total_pend += pend
        rows.append({
            'mes':       f'{mo:02d}/{yr}',
            'at':        _fmt_num(at),
            'pm':        _fmt_num(pm),
            'ra':        _fmt_num(ra),
            'pe':        _fmt_num(pe),
            'ch':        _fmt_num(ch),
            'poupanca':  _fmt_num(poupanca),
            'fi':        _fmt_num(fi),
            'total':     _fmt_num(row_total),
            'recebido':  _fmt_num(rec),
            'a_receber': _fmt_num(pend),
        })

    grand_total = totals['total'] or 1

    def pct(v):
        return f'{v / grand_total * 100:.2f}%'.replace('.', ',')

    summary = [
        {'tipo': TIPO_LABELS[t], 'total': _fmt_num(totals[t]), 'pct': pct(totals[t])}
        for t in TIPOS
    ] + [
        {'tipo': 'Poupança',      'total': _fmt_num(totals['poupanca']), 'pct': pct(totals['poupanca'])},
        {'tipo': 'Financiamento', 'total': _fmt_num(totals['FI']),       'pct': pct(totals['FI'])},
        {'tipo': 'Total',         'total': _fmt_num(grand_total),        'pct': '100,00%'},
        {'tipo': 'Recebido',      'total': _fmt_num(total_rec),          'pct': pct(total_rec)},
        {'tipo': 'A receber',     'total': _fmt_num(total_pend),         'pct': pct(total_pend)},
    ]

    context = {
        'rows':        rows,
        'summary':     summary,
        'total_geral': _fmt_brl(grand_total),
    }
    return render(request, 'cota365/fluxo.html', context)


def parcelas_view(request):
    TIPOS = ['AT', 'PM', 'RA', 'CH', 'PE', 'FI']

    tipo_filtro   = request.GET.get('tipo', '')
    status_filtro = request.GET.get('status', '')
    q             = request.GET.get('q', '').strip()
    sort          = request.GET.get('sort', 'vencimento')
    sort_dir      = request.GET.get('dir', 'asc')

    qs = Parcela.objects.all()
    if tipo_filtro:
        qs = qs.filter(tipo=tipo_filtro)
    if status_filtro == 'pago':
        qs = qs.exclude(data_pagamento=None)
    elif status_filtro == 'pendente':
        qs = qs.filter(data_pagamento=None)
    if q:
        qs = qs.filter(cliente__icontains=q) | Parcela.objects.filter(titulo__icontains=q)
        if tipo_filtro:
            qs = qs.filter(tipo=tipo_filtro)

    _agg = qs.aggregate(
        total_geral=Sum('valor'),
        total_pago=Sum('valor', filter=Q(data_pagamento__isnull=False)),
        total_pendente=Sum('valor', filter=Q(data_pagamento__isnull=True)),
        n_pagas=Count('id', filter=Q(data_pagamento__isnull=False)),
        n_pendentes=Count('id', filter=Q(data_pagamento__isnull=True)),
    )
    total_geral    = _agg['total_geral']    or 0.0
    total_pago     = _agg['total_pago']     or 0.0
    total_pendente = _agg['total_pendente'] or 0.0
    n_pagas        = _agg['n_pagas']        or 0
    n_pendentes    = _agg['n_pendentes']    or 0
    _ORM_FIELD = {
        'titulo':     'titulo',
        'tipo':       'tipo',
        'vencimento': 'vencimento',
        'pagamento':  'data_pagamento',
        'cliente':    'cliente',
    }
    _field = _ORM_FIELD.get(sort, 'vencimento')
    _order = F(_field).desc(nulls_last=True) if sort_dir == 'desc' else F(_field).asc(nulls_last=True)
    paginator = Paginator(qs.order_by(_order), 50)
    page_obj  = paginator.get_page(request.GET.get('page', 1))

    def fmt_date(d):
        return d.strftime('%d/%m/%Y') if d else ''

    rows = [
        {
            'titulo':         p.titulo,
            'parcela':        p.parcela,
            'tipo':           p.tipo,
            'vencimento':     fmt_date(p.vencimento),
            'data_pagamento': fmt_date(p.data_pagamento),
            'valor':          _fmt_brl(p.valor),
            'cliente':        p.cliente,
            'pago':           bool(p.data_pagamento),
        }
        for p in page_obj
    ]

    qd = request.GET.copy()
    qd.pop('page', None)

    context = {
        'rows':           rows,
        'page_obj':       page_obj,
        'query_string':   qd.urlencode(),
        'tipos':          TIPOS,
        'tipo_filtro':    tipo_filtro,
        'status_filtro':  status_filtro,
        'q':              q,
        'total_geral':    _fmt_brl(total_geral),
        'total_pago':     _fmt_brl(total_pago),
        'total_pendente': _fmt_brl(total_pendente),
        'n_pagas':        n_pagas,
        'n_pendentes':    n_pendentes,
        'n_total':        paginator.count,
        'sort':           sort,
        'sort_dir':       sort_dir,
    }
    return render(request, 'cota365/parcelas.html', context)


# ---------------------------------------------------------------------------
# Exports
# ---------------------------------------------------------------------------

def export_parcelas(request):
    fmt           = request.GET.get('format', 'xlsx')
    tipo_filtro   = request.GET.get('tipo', '')
    status_filtro = request.GET.get('status', '')
    q             = request.GET.get('q', '').strip()
    sort          = request.GET.get('sort', 'vencimento')
    sort_dir      = request.GET.get('dir', 'asc')

    qs = Parcela.objects.all()
    if tipo_filtro:
        qs = qs.filter(tipo=tipo_filtro)
    if status_filtro == 'pago':
        qs = qs.exclude(data_pagamento=None)
    elif status_filtro == 'pendente':
        qs = qs.filter(data_pagamento=None)
    if q:
        qs = qs.filter(cliente__icontains=q) | Parcela.objects.filter(titulo__icontains=q)
        if tipo_filtro:
            qs = qs.filter(tipo=tipo_filtro)

    _ORM_FIELD = {'titulo': 'titulo', 'tipo': 'tipo', 'vencimento': 'vencimento',
                  'pagamento': 'data_pagamento', 'cliente': 'cliente'}
    _field = _ORM_FIELD.get(sort, 'vencimento')
    _order = F(_field).desc(nulls_last=True) if sort_dir == 'desc' else F(_field).asc(nulls_last=True)
    qs = qs.order_by(_order)

    _agg = qs.aggregate(
        total_geral=Sum('valor'),
        total_pago=Sum('valor', filter=Q(data_pagamento__isnull=False)),
        total_pendente=Sum('valor', filter=Q(data_pagamento__isnull=True)),
    )

    def fmt_date(d):
        return d.strftime('%d/%m/%Y') if d else ''

    rows = [{'titulo': p.titulo, 'parcela': p.parcela, 'tipo': p.tipo,
             'vencimento': fmt_date(p.vencimento), 'data_pagamento': fmt_date(p.data_pagamento),
             'valor': p.valor or 0.0, 'cliente': p.cliente,
             'pago': bool(p.data_pagamento)} for p in qs]

    total_geral    = _agg['total_geral']    or 0.0
    total_pago     = _agg['total_pago']     or 0.0
    total_pendente = _agg['total_pendente'] or 0.0

    if fmt == 'pdf':
        return _export_parcelas_pdf(rows, total_geral, total_pago, total_pendente)
    return _export_parcelas_xlsx(rows, total_geral, total_pago, total_pendente)


def _export_parcelas_xlsx(rows, total_geral, total_pago, total_pendente):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Parcelas'

    ws.merge_cells('A1:H1')
    ws['A1'] = 'Cota 365 — Parcelas'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:H2')
    ws['A2'] = (f'Total: {_fmt_brl(total_geral)}   |   '
                f'Pagas: {_fmt_brl(total_pago)}   |   '
                f'Pendentes: {_fmt_brl(total_pendente)}')
    ws['A2'].font = Font(bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['TÍTULO', 'PARCELA', 'TIPO', 'VENCIMENTO', 'PAGAMENTO', 'VALOR', 'CLIENTE', 'STATUS']
    widths  = [10, 10, 8, 14, 14, 18, 35, 12]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w

    pago_fill    = PatternFill('solid', fgColor='F0FFF4')
    pendente_fill = PatternFill('solid', fgColor='FFFEF8')

    for i, r in enumerate(rows):
        row  = 5 + i
        fill = pago_fill if r['pago'] else pendente_fill
        vals = [f"#{r['titulo']}", r['parcela'], r['tipo'], r['vencimento'],
                r['data_pagamento'], r['valor'], r['cliente'],
                'Paga' if r['pago'] else 'Pendente']
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = _thin_border()
            cell.fill = fill
            if col == 6:
                cell.number_format = '"R$ "#,##0.00'
                cell.alignment = Alignment(horizontal='right')

    total_row = 5 + len(rows)
    ws.cell(row=total_row, column=1, value=f'Total ({len(rows)} parcelas)').font = _BOLD
    ws.cell(row=total_row, column=1).fill = _TOTAL_FILL
    for col in range(2, 6):
        ws.cell(row=total_row, column=col).fill = _TOTAL_FILL
    c = ws.cell(row=total_row, column=6, value=total_geral)
    c.number_format = '"R$ "#,##0.00'
    c.font = _BOLD
    c.fill = _TOTAL_FILL
    c.alignment = Alignment(horizontal='right')
    for col in range(7, 9):
        ws.cell(row=total_row, column=col).fill = _TOTAL_FILL

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="parcelas.xlsx"'
    return resp


def _export_parcelas_pdf(rows, total_geral, total_pago, total_pendente):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story = [
        Paragraph('Cota 365 — Parcelas',
                  ParagraphStyle('title', parent=styles['Heading1'], fontSize=14, spaceAfter=4)),
        Paragraph(f'Total: {_fmt_brl(total_geral)}   |   Pagas: {_fmt_brl(total_pago)}   |   Pendentes: {_fmt_brl(total_pendente)}',
                  ParagraphStyle('sub', parent=styles['Normal'], fontSize=9, spaceAfter=12)),
    ]

    header = ['TÍTULO', 'PARCELA', 'TIPO', 'VENCIMENTO', 'PAGAMENTO', 'VALOR', 'CLIENTE', 'STATUS']
    data   = [header]
    for r in rows:
        data.append([f"#{r['titulo']}", r['parcela'], r['tipo'], r['vencimento'],
                     r['data_pagamento'], _fmt_brl(r['valor']), r['cliente'],
                     'Paga' if r['pago'] else 'Pendente'])
    data.append(['', '', '', '', f'TOTAL ({len(rows)})', _fmt_brl(total_geral), '', ''])

    t = Table(data,
              colWidths=[1.5*cm, 1.5*cm, 1.5*cm, 2.5*cm, 2.5*cm, 3*cm, 7*cm, 2*cm],
              repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
        ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
        ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, -1), 7),
        ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN',         (6, 1), (6, -2), 'LEFT'),
        ('ALIGN',         (5, 1), (5, -1), 'RIGHT'),
        ('ROWBACKGROUNDS',(0, 1), (-1, -2), [colors.HexColor('#F0FFF4'), colors.HexColor('#FFFEF8')]),
        ('BACKGROUND',    (0, -1), (-1, -1), colors.HexColor('#e9ecef')),
        ('FONTNAME',      (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID',          (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('TOPPADDING',    (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(t)
    doc.build(story)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/pdf')
    resp['Content-Disposition'] = 'attachment; filename="parcelas.pdf"'
    return resp


def export_vendas(request):
    fmt = request.GET.get('format', 'xlsx')
    vendas_rows = _load_vendas()
    fluxo_rows  = _load_fluxo()
    fluxo_by_id = {r['id']: r for r in fluxo_rows}
    vinculos    = _load_vinculos()

    contracts = []
    total_geral = 0.0
    for v in vendas_rows:
        f = fluxo_by_id.get(v['numero'], {})
        valor = f.get('vgv', 0) or f.get('pv', 0)
        status = 'VENDIDO' if 'Vend' in v['situacao'] else v['situacao'].upper()
        vinc = vinculos.get(v['unidade'], {})
        contracts.append({
            'numero':        f"#{v['numero']}",
            'cliente':       v['cliente'],
            'empreendimento': f.get('empreendimento', 'Cota 365'),
            'unidade':       v['unidade'],
            'status':        status,
            'valor':         valor,
            'garagens':      vinc.get('garagens', ''),
            'hb':            vinc.get('hb', ''),
        })
        total_geral += valor
    contracts.sort(key=lambda x: x['cliente'].lower())

    if fmt == 'pdf':
        return _export_vendas_pdf(contracts, total_geral)
    return _export_vendas_xlsx(contracts, total_geral)


def export_fluxo(request):
    fmt = request.GET.get('format', 'xlsx')

    _base = Parcela.objects.filter(vencimento__isnull=False)

    monthly = defaultdict(lambda: defaultdict(float))
    for r in (_base
              .annotate(ano=ExtractYear('vencimento'), mes=ExtractMonth('vencimento'))
              .values('ano', 'mes', 'tipo')
              .annotate(total=Sum('valor'))
              .order_by('ano', 'mes')):
        monthly[(r['ano'], r['mes'])][r['tipo']] = r['total'] or 0.0

    monthly_rec  = {}
    monthly_pend = {}
    for r in (Parcela.objects.filter(data_pagamento__isnull=False)
              .annotate(ano=ExtractYear('data_pagamento'), mes=ExtractMonth('data_pagamento'))
              .values('ano', 'mes')
              .annotate(rec=Sum('valor'))
              .order_by('ano', 'mes')):
        monthly_rec[(r['ano'], r['mes'])] = r['rec'] or 0.0

    for r in (_base.filter(data_pagamento__isnull=True)
              .annotate(ano=ExtractYear('vencimento'), mes=ExtractMonth('vencimento'))
              .values('ano', 'mes')
              .annotate(pend=Sum('valor'))
              .order_by('ano', 'mes')):
        monthly_pend[(r['ano'], r['mes'])] = r['pend'] or 0.0

    totals     = defaultdict(float)
    total_rec  = 0.0
    total_pend = 0.0
    rows = []
    for key in sorted(monthly.keys()):
        yr, mo = key
        td = monthly[key]
        at = td.get('AT', 0); pm = td.get('PM', 0); ra = td.get('RA', 0)
        pe = td.get('PE', 0); ch = td.get('CH', 0); fi = td.get('FI', 0)
        poupanca  = at + pm + ra + pe + ch
        row_total = poupanca + fi
        rec   = monthly_rec.get(key, 0.0)
        pend  = monthly_pend.get(key, 0.0)
        for t, v in (('AT', at), ('PM', pm), ('RA', ra), ('PE', pe),
                     ('CH', ch), ('FI', fi), ('poupanca', poupanca)):
            totals[t] += v
        totals['total'] += row_total
        total_rec  += rec
        total_pend += pend
        rows.append({
            'mes': f'{mo:02d}/{yr}',
            'at': at, 'pm': pm, 'ra': ra, 'pe': pe, 'ch': ch,
            'poupanca': poupanca, 'fi': fi,
            'total': row_total, 'recebido': rec, 'a_receber': pend,
        })

    grand_total = totals['total']
    totals['recebido']  = total_rec
    totals['a_receber'] = total_pend

    if fmt == 'pdf':
        return _export_fluxo_pdf(rows, totals, grand_total)
    return _export_fluxo_xlsx(rows, totals, grand_total)


# -- Excel helpers -----------------------------------------------------------

_HEADER_FILL = PatternFill('solid', fgColor='1a1a2e')
_HEADER_FONT = Font(color='FFFFFF', bold=True)
_STATUS_FILL = PatternFill('solid', fgColor='d4edda')
_STATUS_FONT = Font(color='155724', bold=True)
_ALT_FILL    = PatternFill('solid', fgColor='f8f9fa')
_TOTAL_FILL  = PatternFill('solid', fgColor='e9ecef')
_BOLD        = Font(bold=True)


def _thin_border():
    s = Side(style='thin', color='dee2e6')
    return Border(left=s, right=s, top=s, bottom=s)


def _export_unidades_xlsx(headers, rows, price_cols):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Unidades'

    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    ws['A1'] = 'Cota 365 — Cadastro de Unidades'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col)].width = max(14, len(str(h)) + 4)

    status_col = next((i for i, h in enumerate(headers) if h.lower() == 'status'), None)
    STATUS_COLORS = {
        'vendido':    ('d4edda', '155724'),
        'disponível': ('d1ecf1', '0c5460'),
        'reservado':  ('fff3cd', '856404'),
        'permuta':    ('f3e8ff', '5a1e96'),
    }

    for i, row in enumerate(rows):
        r = 4 + i
        fill = _ALT_FILL if i % 2 == 0 else PatternFill()
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=col, value=_xlsx_val(val))
            cell.border = _thin_border()
            if col - 1 in price_cols and val not in ('', 0, None):
                cell.number_format = '"R$ "#,##0.00'
                cell.alignment = Alignment(horizontal='right')
                cell.value = float(val) if val else 0
            elif col - 1 == status_col and val:
                cp = STATUS_COLORS.get(str(val).lower(), ('', ''))
                if cp[0]:
                    cell.fill = PatternFill('solid', fgColor=cp[0])
                    cell.font = Font(color=cp[1], bold=True)
                    cell.alignment = Alignment(horizontal='center')
            else:
                cell.fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="unidades_cota365.xlsx"'
    return resp


def _export_unidades_pdf(headers, rows, price_cols):
    buf = io.BytesIO()
    page = landscape(A4) if len(headers) > 6 else A4
    doc = SimpleDocTemplate(buf, pagesize=page,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    title_s = ParagraphStyle('t', parent=styles['Heading1'], fontSize=13, spaceAfter=6)
    sub_s   = ParagraphStyle('s', parent=styles['Normal'],  fontSize=9,  spaceAfter=14,
                              textColor=colors.HexColor('#6c757d'))

    story = [
        Paragraph('Cota 365 — Cadastro de Unidades', title_s),
        Paragraph(f'Total de unidades: {len(rows)}  |  Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}', sub_s),
    ]

    base_w    = doc.width / len(headers)
    first_w   = base_w * 0.6
    remaining = (doc.width - first_w) / (len(headers) - 1) if len(headers) > 1 else base_w
    tabela_idx = next((i for i, h in enumerate(headers) if 'tabela' in h.lower()), None)
    venda_idx  = next((i for i, h in enumerate(headers) if 'venda'  in h.lower()), None)
    wide_cols  = {i for i in (tabela_idx, venda_idx) if i is not None}
    total_extra = remaining * 0.15 * len(wide_cols)
    n_other = len(headers) - 1 - len(wide_cols)
    shrink = total_extra / n_other if n_other > 0 else 0
    col_widths = []
    for i in range(len(headers)):
        if i == 0:          col_widths.append(first_w)
        elif i in wide_cols: col_widths.append(remaining * 1.15)
        else:               col_widths.append(remaining - shrink)

    HEADER_LABELS = {'Complemento do Tipo': 'Compl. do Tipo'}
    header_row = [
        Paragraph(f'<b><font color="white">{HEADER_LABELS.get(h, h)}</font></b>',
                  ParagraphStyle('th', parent=styles['Normal'], fontSize=7, alignment=1))
        for h in headers
    ]
    data = [header_row]
    for row in rows:
        pdf_row = []
        for i, val in enumerate(row):
            if i in price_cols and val not in ('', None, 0):
                txt = _fmt_brl(float(val))
                pdf_row.append(Paragraph(txt, ParagraphStyle('cr', parent=styles['Normal'], fontSize=7, alignment=2)))
            else:
                pdf_row.append(Paragraph(str(val) if val is not None else '',
                                         ParagraphStyle('c', parent=styles['Normal'], fontSize=7)))
        data.append(pdf_row)

    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#dee2e6')),
        ('TOPPADDING',    (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    story.append(t)
    doc.build(story)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/pdf')
    resp['Content-Disposition'] = 'attachment; filename="unidades_cota365.pdf"'
    return resp


def _export_vendas_xlsx(contracts, total_geral):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Unidades Vendidas'

    ws.merge_cells('A1:H1')
    ws['A1'] = 'Cota 365 — Unidades Vendidas'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:H2')
    ws['A2'] = f'Total de Contratos: {len(contracts)}    |    Total Geral: {_fmt_brl(total_geral)}'
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(bold=True)

    headers = ['Nº', 'CLIENTE', 'EMPREENDIMENTO', 'UNIDADE', 'GARAGENS', 'HB', 'STATUS', 'TOTAL CONTRATO']
    widths  = [8, 38, 16, 12, 14, 10, 12, 22]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w

    for i, c in enumerate(contracts):
        row = 5 + i
        fill = _ALT_FILL if i % 2 == 0 else PatternFill()
        data = [c['numero'], c['cliente'], c['empreendimento'], c['unidade'],
                c['garagens'], c['hb'], c['status'], c['valor']]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=_xlsx_val(val))
            cell.border = _thin_border()
            if col not in (7, 8): cell.fill = fill
            if col == 8:
                cell.number_format = '"R$ "#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            if col == 7:
                cell.fill = _STATUS_FILL
                cell.font = _STATUS_FONT
                cell.alignment = Alignment(horizontal='center')

    total_row = 5 + len(contracts)
    ws.cell(row=total_row, column=7, value='TOTAL GERAL').font = _BOLD
    ws.cell(row=total_row, column=7).fill = _TOTAL_FILL
    ws.cell(row=total_row, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=total_row, column=8, value=total_geral).number_format = '"R$ "#,##0.00'
    ws.cell(row=total_row, column=8).font = _BOLD
    ws.cell(row=total_row, column=8).fill = _TOTAL_FILL
    ws.cell(row=total_row, column=8).alignment = Alignment(horizontal='right')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="unidades_vendidas.xlsx"'
    return resp


def _export_fluxo_xlsx(rows, totals, total_geral):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Fluxo Mensal'

    n_cols = 11
    ws.merge_cells(f'A1:{get_column_letter(n_cols)}1')
    ws['A1'] = 'Cota 365 — Fluxo Mensal de Receitas'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'A2:{get_column_letter(n_cols)}2')
    ws['A2'] = f'Total Geral: {_fmt_brl(total_geral)}'
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(bold=True)

    headers = ['MÊS', 'ATO', 'MENSAIS', 'REF. ANUAIS', 'PERMUTA',
               'CHAVES', 'POUPANÇA', 'FINANCIAMENTO', 'TOTAL', 'RECEBIDO', 'A RECEBER']
    widths  = [12, 16, 16, 16, 16, 16, 16, 18, 18, 18, 18]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w

    fields = ['mes', 'at', 'pm', 'ra', 'pe', 'ch', 'poupanca', 'fi', 'total', 'recebido', 'a_receber']
    for i, r in enumerate(rows):
        row = 5 + i
        fill = _ALT_FILL if i % 2 == 0 else PatternFill()
        for col, key in enumerate(fields, 1):
            val = r[key]
            cell = ws.cell(row=row, column=col, value=_xlsx_val(val))
            cell.border = _thin_border()
            cell.fill = fill
            if col > 1:
                cell.number_format = '"R$ "#,##0.00'
                cell.alignment = Alignment(horizontal='right')

    total_row = 5 + len(rows)
    total_vals = [totals.get('AT', 0), totals.get('PM', 0), totals.get('RA', 0),
                  totals.get('PE', 0), totals.get('CH', 0), totals.get('poupanca', 0),
                  totals.get('FI', 0), totals.get('total', 0),
                  totals.get('recebido', 0), totals.get('a_receber', 0)]
    ws.cell(row=total_row, column=1, value='TOTAL').font = _BOLD
    ws.cell(row=total_row, column=1).fill = _TOTAL_FILL
    for col, val in enumerate(total_vals, 2):
        cell = ws.cell(row=total_row, column=col, value=val)
        cell.number_format = '"R$ "#,##0.00'
        cell.font = _BOLD
        cell.fill = _TOTAL_FILL
        cell.alignment = Alignment(horizontal='right')
        cell.border = _thin_border()

    # Resumo por tipo
    grand = totals.get('total', 0) or 1

    def _pct(v):
        return round(v / grand * 100, 2)

    summary_items = [
        ('Ato',           totals.get('AT', 0)),
        ('Mensais',       totals.get('PM', 0)),
        ('Ref. Anuais',   totals.get('RA', 0)),
        ('Permuta',       totals.get('PE', 0)),
        ('Chaves',        totals.get('CH', 0)),
        ('Poupança',      totals.get('poupanca', 0)),
        ('Financiamento', totals.get('FI', 0)),
        ('Total',         totals.get('total', 0)),
        ('Recebido',      totals.get('recebido', 0)),
        ('A receber',     totals.get('a_receber', 0)),
    ]

    summary_start = total_row + 2
    for col, h in enumerate(['TIPO', 'TOTAL', '%'], 1):
        cell = ws.cell(row=summary_start, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = _thin_border()
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10

    for i, (tipo, val) in enumerate(summary_items):
        row = summary_start + 1 + i
        fill = _TOTAL_FILL if tipo in ('Poupança', 'Financiamento', 'Total') else (
               _ALT_FILL if i % 2 == 0 else PatternFill())
        bold = tipo in ('Poupança', 'Financiamento', 'Total')
        c1 = ws.cell(row=row, column=1, value=tipo)
        c1.border = _thin_border(); c1.fill = fill
        if bold: c1.font = _BOLD
        c2 = ws.cell(row=row, column=2, value=val)
        c2.number_format = '"R$ "#,##0.00'
        c2.alignment = Alignment(horizontal='right')
        c2.border = _thin_border(); c2.fill = fill
        if bold: c2.font = _BOLD
        c3 = ws.cell(row=row, column=3, value=_pct(val))
        c3.number_format = '0.00"%"'
        c3.alignment = Alignment(horizontal='right')
        c3.border = _thin_border(); c3.fill = fill
        if bold: c3.font = _BOLD

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="fluxo_mensal.xlsx"'
    return resp


def _export_vendas_pdf(contracts, total_geral):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story = [
        Paragraph('Cota 365 — Unidades Vendidas',
                  ParagraphStyle('title', parent=styles['Heading1'], fontSize=14, spaceAfter=4)),
        Paragraph(f'Nº de Contratos: {len(contracts)}    |    Total Geral: {_fmt_brl(total_geral)}',
                  ParagraphStyle('sub', parent=styles['Normal'], fontSize=10, spaceAfter=12)),
    ]

    header = ['Nº', 'CLIENTE', 'EMPREENDIMENTO', 'UNIDADE', 'GARAGENS', 'HB', 'STATUS', 'TOTAL CONTRATO']
    data = [header]
    for c in contracts:
        data.append([c['numero'], c['cliente'], c['empreendimento'],
                     c['unidade'], c['garagens'], c['hb'], c['status'], _fmt_brl(c['valor'])])
    data.append(['', '', '', '', '', '', 'TOTAL GERAL', _fmt_brl(total_geral)])

    t = Table(data, colWidths=[1.2*cm, 7*cm, 3.5*cm, 2.5*cm, 2.5*cm, 1.8*cm, 2.2*cm, 3.8*cm], repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
        ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
        ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0), 8),
        ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN',         (1, 1), (1, -2),  'LEFT'),
        ('ALIGN',         (7, 1), (7, -1),  'RIGHT'),
        ('FONTSIZE',      (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS',(0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')]),
        ('BACKGROUND',    (0, -1), (-1, -1), colors.HexColor('#e9ecef')),
        ('FONTNAME',      (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID',          (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(t)
    doc.build(story)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/pdf')
    resp['Content-Disposition'] = 'attachment; filename="unidades_vendidas.pdf"'
    return resp


def _export_fluxo_pdf(rows, totals, grand_total):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.2*cm, bottomMargin=1.2*cm)
    styles = getSampleStyleSheet()

    NAVY  = colors.HexColor('#1a1a2e')
    GRAY1 = colors.HexColor('#f8f9fa')
    GRID  = colors.HexColor('#dee2e6')
    POUPA = colors.HexColor('#fffde7')
    FIN   = colors.HexColor('#fff3e0')

    def n(v):
        return _fmt_num(v)

    def pct(v):
        if not grand_total:
            return '0,00%'
        return f'{v / grand_total * 100:.2f}%'.replace('.', ',')

    story = [
        Paragraph('Cota 365 — Fluxo Mensal de Receitas',
                  ParagraphStyle('title', parent=styles['Normal'],
                                 fontName='Helvetica-Bold', fontSize=13, spaceAfter=8)),
    ]

    # ── Tabela principal ──────────────────────────────────────────────────────
    HEADERS = ['Mês', 'Ato', 'Mensais', 'Ref. Anuais', 'Permuta', 'Chaves',
               'Poupança', 'Financiamento', 'Total', 'Recebido', 'A receber']

    # página landscape A4 = 29,7cm - 3cm margens = 26,7cm
    W_MES = 1.7 * cm
    W_COL = (26.7 * cm - W_MES) / 10
    col_widths = [W_MES] + [W_COL] * 10

    def hdr_cell(txt):
        return Paragraph(f'<b><font color="white">{txt}</font></b>',
                         ParagraphStyle('h', fontName='Helvetica-Bold',
                                        fontSize=7, alignment=1))

    data = [[hdr_cell(h) for h in HEADERS]]

    for r in rows:
        data.append([
            Paragraph(f'<b>{r["mes"]}</b>', ParagraphStyle('m', fontSize=6.5, fontName='Helvetica-Bold')),
            n(r['at']), n(r['pm']), n(r['ra']), n(r['pe']), n(r['ch']),
            n(r['poupanca']), n(r['fi']),
            Paragraph(f'<b>{n(r["total"])}</b>', ParagraphStyle('t', fontSize=6.5, fontName='Helvetica-Bold')),
            n(r['recebido']), n(r['a_receber']),
        ])

    # Linha Total
    data.append([
        Paragraph('<b><font color="white">Total</font></b>',
                  ParagraphStyle('tot', fontName='Helvetica-Bold', fontSize=7, alignment=1)),
        *[Paragraph(f'<b><font color="white">{n(totals.get(k, 0))}</font></b>',
                    ParagraphStyle('tv', fontName='Helvetica-Bold', fontSize=6.5, alignment=2))
          for k in ('AT', 'PM', 'RA', 'PE', 'CH', 'poupanca', 'FI', 'total', 'recebido', 'a_receber')],
    ])

    # Linha %
    def pct_cell(v):
        return Paragraph(f'<i>{pct(v)}</i>',
                         ParagraphStyle('pc', fontName='Helvetica-Oblique', fontSize=6, alignment=2))

    data.append([
        Paragraph('<i>%</i>', ParagraphStyle('pm', fontName='Helvetica-Oblique', fontSize=6)),
        *[pct_cell(totals.get(k, 0))
          for k in ('AT', 'PM', 'RA', 'PE', 'CH', 'poupanca', 'FI', 'total', 'recebido', 'a_receber')],
    ])

    n_data   = len(data)
    total_r  = n_data - 2   # linha Total (0-based)
    pct_r    = n_data - 1   # linha %

    main_style = TableStyle([
        # Header
        ('BACKGROUND',    (0, 0),  (-1, 0),       NAVY),
        ('FONTNAME',      (0, 0),  (-1, 0),       'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0),  (-1, 0),       7),
        ('ALIGN',         (0, 0),  (-1, 0),       'CENTER'),
        ('VALIGN',        (0, 0),  (-1, -1),      'MIDDLE'),
        # Data rows — alternating
        ('ROWBACKGROUNDS',(0, 1),  (-1, total_r - 1), [colors.white, GRAY1]),
        ('FONTSIZE',      (0, 1),  (-1, -1),      6.5),
        ('ALIGN',         (0, 1),  (0, -1),       'CENTER'),
        ('ALIGN',         (1, 1),  (-1, -1),      'RIGHT'),
        # Total row
        ('BACKGROUND',    (0, total_r), (-1, total_r), NAVY),
        ('FONTNAME',      (0, total_r), (-1, total_r), 'Helvetica-Bold'),
        # % row
        ('BACKGROUND',    (0, pct_r),  (-1, pct_r),  GRAY1),
        ('FONTSIZE',      (0, pct_r),  (-1, pct_r),  6),
        # Grid
        ('GRID',          (0, 0),  (-1, -1),      0.4, GRID),
        ('TOPPADDING',    (0, 0),  (-1, -1),      3),
        ('BOTTOMPADDING', (0, 0),  (-1, -1),      3),
        ('LEFTPADDING',   (0, 0),  (-1, -1),      3),
        ('RIGHTPADDING',  (0, 0),  (-1, -1),      3),
    ])

    main_table = Table(data, colWidths=col_widths, repeatRows=1)
    main_table.setStyle(main_style)
    story.append(main_table)
    story.append(PageBreak())

    # ── Tabela resumo por tipo ────────────────────────────────────────────────
    summary_rows = [
        ('Ato',           totals.get('AT', 0)),
        ('Mensais',       totals.get('PM', 0)),
        ('Ref. Anuais',   totals.get('RA', 0)),
        ('Permuta',       totals.get('PE', 0)),
        ('Chaves',        totals.get('CH', 0)),
        ('Poupança',      totals.get('poupanca', 0)),
        ('Financiamento', totals.get('FI', 0)),
        ('Total',         grand_total),
        ('Recebido',      totals.get('recebido', 0)),
        ('A receber',     totals.get('a_receber', 0)),
    ]

    def shdr(txt):
        return Paragraph(f'<b><font color="white">{txt}</font></b>',
                         ParagraphStyle('sh', fontName='Helvetica-Bold', fontSize=8, alignment=1))

    sum_data = [[shdr('Tipo'), shdr('Total'), shdr('%')]]
    for label, val in summary_rows:
        bold = label in ('Total',)
        fn   = 'Helvetica-Bold' if bold else 'Helvetica'
        sum_data.append([
            Paragraph(f'<b>{label}</b>' if bold else label,
                      ParagraphStyle('sl', fontName=fn, fontSize=8)),
            Paragraph(f'<b>{n(val)}</b>' if bold else n(val),
                      ParagraphStyle('sv', fontName=fn, fontSize=8, alignment=2)),
            Paragraph(f'<b>{pct(val)}</b>' if bold else pct(val),
                      ParagraphStyle('sp', fontName=fn, fontSize=8, alignment=2)),
        ])

    sum_table = Table(sum_data, colWidths=[4.5*cm, 4.5*cm, 2.5*cm])
    sum_style = TableStyle([
        ('BACKGROUND',    (0, 0),  (-1, 0),  NAVY),
        ('FONTNAME',      (0, 0),  (-1, 0),  'Helvetica-Bold'),
        ('ALIGN',         (0, 0),  (-1, 0),  'CENTER'),
        ('VALIGN',        (0, 0),  (-1, -1), 'MIDDLE'),
        ('ALIGN',         (1, 1),  (-1, -1), 'RIGHT'),
        ('FONTSIZE',      (0, 1),  (-1, -1), 8),
        # Poupança row (index 6 = 7th data row after header)
        ('BACKGROUND',    (0, 6),  (-1, 6),  POUPA),
        # Financiamento row (index 7)
        ('BACKGROUND',    (0, 7),  (-1, 7),  FIN),
        # Total row (index 8)
        ('FONTNAME',      (0, 8),  (-1, 8),  'Helvetica-Bold'),
        ('LINEABOVE',     (0, 8),  (-1, 8),  1, NAVY),
        # Grid
        ('GRID',          (0, 0),  (-1, -1), 0.4, GRID),
        ('TOPPADDING',    (0, 0),  (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0),  (-1, -1), 4),
        ('LEFTPADDING',   (0, 0),  (-1, -1), 4),
        ('RIGHTPADDING',  (0, 0),  (-1, -1), 4),
    ])
    sum_table.setStyle(sum_style)
    story.append(sum_table)

    doc.build(story)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/pdf')
    resp['Content-Disposition'] = 'attachment; filename="fluxo_mensal.pdf"'
    return resp


# ---------------------------------------------------------------------------
# Comparativo de Áreas — Tabela × Unidades
# ---------------------------------------------------------------------------

def export_areas_comparativo(request):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HDR_FILL  = PatternFill('solid', fgColor='1A1A2E')
    HDR_FONT  = Font(bold=True, color='FFFFFF', size=10)
    DIFF_FILL = PatternFill('solid', fgColor='FFE0E0')
    OK_FILL   = PatternFill('solid', fgColor='E0F4E0')
    ONLY_FILL = PatternFill('solid', fgColor='FFF3CD')
    TOTAL_FONT = Font(bold=True, size=10)
    BORDER = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    def _hdr(ws, row, cols):
        for c, txt in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=txt)
            cell.font  = HDR_FONT
            cell.fill  = HDR_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = BORDER

    def _auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)

    wb = openpyxl.Workbook()

    # ── Aba 1: Tabela ─────────────────────────────────────────────────────────
    ws_tab = wb.active
    ws_tab.title = 'Tabela'
    _hdr(ws_tab, 1, ['UNIDADE', 'TIPOLOGIA', 'SITUAÇÃO', 'ÁREA PRIV. (m²)', 'VALOR TOTAL'])
    ws_tab.row_dimensions[1].height = 28
    _tab_qs = list(_tabela_qs().order_by('unidade'))
    for i, t in enumerate(_tab_qs, 2):
        row = [t.unidade, t.tipologia, t.situacao, t.area_privativa, t.valor_total]
        for c, val in enumerate(row, 1):
            cell = ws_tab.cell(row=i, column=c, value=_xlsx_val(val))
            cell.border = BORDER
            if c in (4, 5):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
    # total
    n = len(_tab_qs)
    total_row = n + 2
    ws_tab.cell(row=total_row, column=1, value='TOTAL').font = TOTAL_FONT
    ws_tab.cell(row=total_row, column=4,
                value=sum(t.area_privativa for t in _tab_qs)).number_format = '#,##0.00'
    ws_tab.cell(row=total_row, column=4).font = TOTAL_FONT
    ws_tab.cell(row=total_row, column=5,
                value=sum(t.valor_total for t in _tab_qs)).number_format = '#,##0.00'
    ws_tab.cell(row=total_row, column=5).font = TOTAL_FONT
    _auto_width(ws_tab)

    # ── Aba 2: Unidades ───────────────────────────────────────────────────────
    ws_uni = wb.create_sheet('Unidades')
    _hdr(ws_uni, 1, ['UNIDADE', 'TIPO', 'COMPLEMENTO', 'ÁREA PRIV. (m²)',
                     'ÁREA PRIV. ACESS. (m²)', 'ÁREA COMUM (m²)', 'TOTAL PRIV. (m²)', 'FRAÇÃO IDEAL'])
    ws_uni.row_dimensions[1].height = 28
    for i, u in enumerate(Unidade.objects.order_by('unidade'), 2):
        total_priv = u.area_privativa + u.area_priv_acessoria
        row = [u.unidade, u.tipo, u.complemento_tipo, u.area_privativa,
               u.area_priv_acessoria, u.area_comum, total_priv, u.fracao_ideal]
        for c, val in enumerate(row, 1):
            cell = ws_uni.cell(row=i, column=c, value=_xlsx_val(val))
            cell.border = BORDER
            if c in (4, 5, 6, 7):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
    n2 = Unidade.objects.count()
    total_row2 = n2 + 2
    ws_uni.cell(row=total_row2, column=1, value='TOTAL').font = TOTAL_FONT
    for col, attr in [(4, 'area_privativa'), (5, 'area_priv_acessoria'), (6, 'area_comum')]:
        val = sum(getattr(u, attr) for u in Unidade.objects.all())
        c = ws_uni.cell(row=total_row2, column=col, value=val)
        c.number_format = '#,##0.00'
        c.font = TOTAL_FONT
    tp = sum(u.area_privativa + u.area_priv_acessoria for u in Unidade.objects.all())
    c7 = ws_uni.cell(row=total_row2, column=7, value=tp)
    c7.number_format = '#,##0.00'
    c7.font = TOTAL_FONT
    _auto_width(ws_uni)

    # ── Aba 3: Comparativo ────────────────────────────────────────────────────
    ws_cmp = wb.create_sheet('Comparativo')
    _hdr(ws_cmp, 1, [
        'UNIDADE',
        'SITUAÇÃO (Tab.)', 'ÁREA PRIV. Tabela (m²)',
        'ÁREA PRIV. Unidade (m²)', 'ÁREA PRIV. ACESS. (m²)', 'TOTAL PRIV. Unid. (m²)',
        'DIFERENÇA (Tab − Unid)',
        'STATUS',
    ])
    ws_cmp.row_dimensions[1].height = 28

    tab_map  = {t.unidade: t for t in _tabela_qs()}
    uni_map  = {u.unidade: u for u in Unidade.objects.all()}
    all_keys = sorted(set(tab_map) | set(uni_map))

    diff_count = 0
    for i, key in enumerate(all_keys, 2):
        t = tab_map.get(key)
        u = uni_map.get(key)

        ap_tab  = t.area_privativa if t else None
        sit_tab = t.situacao       if t else None
        ap_uni  = u.area_privativa      if u else None
        apa_uni = u.area_priv_acessoria if u else None
        tp_uni  = (ap_uni + apa_uni)    if u else None

        if t and u:
            diff  = round(ap_tab - ap_uni, 4)
            status = 'OK' if diff == 0 else f'DIVERGE {diff:+.4f}'
            fill   = OK_FILL if diff == 0 else DIFF_FILL
            if diff != 0:
                diff_count += 1
        elif t:
            diff   = None
            status = 'Só na Tabela'
            fill   = ONLY_FILL
        else:
            diff   = None
            status = 'Só em Unidades'
            fill   = ONLY_FILL

        row_vals = [key, sit_tab, ap_tab, ap_uni, apa_uni, tp_uni, diff, status]
        for c, val in enumerate(row_vals, 1):
            cell = ws_cmp.cell(row=i, column=c, value=_xlsx_val(val))
            cell.border = BORDER
            cell.fill   = fill
            if c in (3, 4, 5, 6, 7) and val is not None:
                cell.number_format = '#,##0.0000'
                cell.alignment = Alignment(horizontal='right')

    # totais comparativo
    tr = len(all_keys) + 2
    ws_cmp.cell(row=tr, column=1, value='TOTAL').font = TOTAL_FONT
    for col, vals in [
        (3, [t.area_privativa for t in tab_map.values()]),
        (4, [u.area_privativa for u in uni_map.values()]),
        (5, [u.area_priv_acessoria for u in uni_map.values()]),
        (6, [u.area_privativa + u.area_priv_acessoria for u in uni_map.values()]),
    ]:
        c = ws_cmp.cell(row=tr, column=col, value=sum(vals))
        c.number_format = '#,##0.0000'
        c.font = TOTAL_FONT

    ws_cmp.cell(row=tr + 1, column=1,
                value=f'Unidades com divergência: {diff_count}').font = Font(bold=True, color='CC0000')

    _auto_width(ws_cmp)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="comparativo_areas_cota365.xlsx"'
    return resp


# ---------------------------------------------------------------------------
# Comissões — exportações
# ---------------------------------------------------------------------------

def export_comissoes_excel(request):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HDR_FILL  = PatternFill('solid', fgColor='1A1A2E')
    HDR_FONT  = Font(bold=True, color='FFFFFF', size=10)
    BORDER    = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC'),
    )
    TOTAL_FONT = Font(bold=True, size=10)

    def _hdr(ws, cols):
        for c, txt in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=txt)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = BORDER
        ws.row_dimensions[1].height = 28

    def _auto_width(ws):
        for col in ws.columns:
            w = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 40)

    wb = openpyxl.Workbook()

    # ── Aba Lista ─────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Comissões'
    headers = ['Nº', 'Reserva', 'Unidade', 'Cliente', 'Corretor', 'Imobiliária',
               'Beneficiário', 'Tipo Comissão', 'Valor Contrato', '% Comissão',
               'Valor Comissão', 'A Pagar']
    _hdr(ws, headers)

    qs = list(Comissao.objects.all())
    for i, c in enumerate(sorted(qs, key=lambda x: (x.unidade, x.beneficiario.lower())), 2):
        row = [
            c.numero, c.reserva, c.unidade, c.cliente, c.corretor, c.imobiliaria,
            c.beneficiario, c.tipo_comissao,
            c.valor_contrato, c.pct_comissao, c.valor_comissao, c.valor_comissao_pagar,
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=_xlsx_val(val))
            cell.border = BORDER
            if col in (9, 11, 12):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            elif col == 10:
                cell.number_format = '0.00"%"'
                cell.alignment = Alignment(horizontal='right')

    tr = len(qs) + 2
    ws.cell(row=tr, column=1, value='TOTAL').font = TOTAL_FONT
    for col, attr in [(9, 'valor_contrato'), (11, 'valor_comissao'), (12, 'valor_comissao_pagar')]:
        c = ws.cell(row=tr, column=col, value=sum(getattr(x, attr) for x in qs))
        c.number_format = '#,##0.00'
        c.font = TOTAL_FONT
    _auto_width(ws)

    # ── Aba Por Imobiliária ───────────────────────────────────────────────────
    ws2 = wb.create_sheet('Por Imobiliária')
    _hdr(ws2, ['Imobiliária', 'Vendas', 'Comissão', 'A Pagar'])
    imob_map      = defaultdict(lambda: {'comissao': 0.0, 'pagar': 0.0})
    imob_reservas = defaultdict(set)
    for c in qs:
        k = c.imobiliaria or '(sem imobiliária)'
        imob_map[k]['comissao'] += c.valor_comissao
        imob_map[k]['pagar']    += c.valor_comissao_pagar
        imob_reservas[k].add(c.reserva)
    for i, (k, v) in enumerate(sorted(imob_map.items(), key=lambda x: x[0].lower()), 2):
        for col, val in enumerate([k, len(imob_reservas[k]), v['comissao'], v['pagar']], 1):
            cell = ws2.cell(row=i, column=col, value=_xlsx_val(val))
            cell.border = BORDER
            if col in (3, 4):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
    _auto_width(ws2)

    # ── Aba Por Beneficiário ──────────────────────────────────────────────────
    ws3 = wb.create_sheet('Por Beneficiário')
    _hdr(ws3, ['Beneficiário', 'Qtde', 'Comissão', 'A Pagar'])
    benef_map = defaultdict(lambda: {'n': 0, 'comissao': 0.0, 'pagar': 0.0})
    for c in qs:
        k = c.beneficiario or '(sem beneficiário)'
        benef_map[k]['n']        += 1
        benef_map[k]['comissao'] += c.valor_comissao
        benef_map[k]['pagar']    += c.valor_comissao_pagar
    for i, (k, v) in enumerate(sorted(benef_map.items(), key=lambda x: x[0].lower()), 2):
        for col, val in enumerate([k, v['n'], v['comissao'], v['pagar']], 1):
            cell = ws3.cell(row=i, column=col, value=_xlsx_val(val))
            cell.border = BORDER
            if col in (3, 4):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
    _auto_width(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="comissoes_cota365.xlsx"'
    return resp


def export_comissoes_pdf(request):
    from reportlab.platypus import HRFlowable

    qs = list(Comissao.objects.all())
    if not qs:
        return HttpResponse('Sem dados.', status=404)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    W = doc.width
    styles = getSampleStyleSheet()

    NAVY   = colors.HexColor('#1a1a2e')
    BORDER = colors.HexColor('#dee2e6')

    def ps(name, **kw):
        return ParagraphStyle(name, parent=styles['Normal'], **kw)

    title_s = ps('T', fontSize=14, fontName='Helvetica-Bold', textColor=NAVY, spaceAfter=2)
    sub_s   = ps('S', fontSize=8,  textColor=colors.HexColor('#6c757d'), spaceAfter=8)
    sec_s   = ps('H', fontSize=10, fontName='Helvetica-Bold', textColor=NAVY,
                 spaceBefore=10, spaceAfter=4)

    def th(txt):
        return Paragraph(f'<b><font color="white">{txt}</font></b>',
                         ps('TH', fontSize=7, alignment=1))
    def td(txt):
        return Paragraph(str(txt), ps('TD', fontSize=7))
    def tdr(txt):
        return Paragraph(str(txt), ps('TR', fontSize=7, alignment=2))
    def tdb(txt):
        return Paragraph(f'<b>{txt}</b>', ps('TB', fontSize=7, alignment=2))

    def tbl(data, col_widths, total_last=False):
        t = Table(data, colWidths=col_widths, repeatRows=1)
        cmds = [
            ('BACKGROUND',    (0, 0), (-1, 0), NAVY),
            ('GRID',          (0, 0), (-1, -1), 0.3, BORDER),
            ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('TOPPADDING',    (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING',   (0, 0), (-1, -1), 4),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
        ]
        if total_last:
            cmds += [
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f0fe')),
                ('FONTNAME',   (0, -1), (-1, -1), 'Helvetica-Bold'),
            ]
        t.setStyle(TableStyle(cmds))
        return t

    # KPIs
    reservas_vt = {}
    for c in qs:
        if c.reserva not in reservas_vt:
            reservas_vt[c.reserva] = c.valor_contrato
    total_vendas   = len(reservas_vt)
    total_contrato = sum(reservas_vt.values())
    total_comissao = sum(c.valor_comissao       for c in qs)
    total_pagar    = sum(c.valor_comissao_pagar for c in qs)

    story = []
    story.append(Paragraph('Cota 365 — Comissões', title_s))
    story.append(Paragraph(
        f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}  |  {total_vendas} vendas  |  {len(qs)} linhas',
        sub_s))
    story.append(HRFlowable(width='100%', thickness=1, color=BORDER, spaceAfter=8))

    # Resumo KPI
    story.append(Paragraph('Resumo Geral', sec_s))
    story.append(tbl([
        [th('VENDAS'), th('LINHAS'), th('VALOR CONTRATOS'), th('TOTAL COMISSÕES'), th('A PAGAR')],
        [tdr(str(total_vendas)), tdr(str(len(qs))), tdr(_fmt_brl(total_contrato)),
         tdr(_fmt_brl(total_comissao)), tdr(_fmt_brl(total_pagar))],
    ], [W/5]*5))
    story.append(Spacer(1, 8))

    # Por Imobiliária
    imob_map      = defaultdict(lambda: {'comissao': 0.0, 'pagar': 0.0})
    imob_reservas = defaultdict(set)
    for c in qs:
        k = c.imobiliaria or '(sem imobiliária)'
        imob_map[k]['comissao'] += c.valor_comissao
        imob_map[k]['pagar']    += c.valor_comissao_pagar
        imob_reservas[k].add(c.reserva)

    story.append(Paragraph('Por Imobiliária', sec_s))
    imob_rows = [[th('IMOBILIÁRIA'), th('VENDAS'), th('COMISSÃO'), th('A PAGAR')]]
    for k, v in sorted(imob_map.items(), key=lambda x: x[0].lower()):
        imob_rows.append([td(k), tdr(str(len(imob_reservas[k]))),
                          tdr(_fmt_brl(v['comissao'])), tdr(_fmt_brl(v['pagar']))])
    imob_rows.append([tdb('TOTAL'), tdb(str(total_vendas)),
                      tdb(_fmt_brl(total_comissao)), tdb(_fmt_brl(total_pagar))])
    story.append(tbl(imob_rows, [W*0.48, W*0.12, W*0.20, W*0.20], total_last=True))
    story.append(Spacer(1, 8))

    # Por Beneficiário
    benef_map = defaultdict(lambda: {'n': 0, 'comissao': 0.0, 'pagar': 0.0})
    for c in qs:
        k = c.beneficiario or '(sem beneficiário)'
        benef_map[k]['n']        += 1
        benef_map[k]['comissao'] += c.valor_comissao
        benef_map[k]['pagar']    += c.valor_comissao_pagar

    story.append(Paragraph('Por Beneficiário', sec_s))
    benef_rows = [[th('BENEFICIÁRIO'), th('QTDE'), th('COMISSÃO'), th('A PAGAR')]]
    for k, v in sorted(benef_map.items(), key=lambda x: x[0].lower()):
        benef_rows.append([td(k), tdr(str(v['n'])),
                           tdr(_fmt_brl(v['comissao'])), tdr(_fmt_brl(v['pagar']))])
    benef_rows.append([tdb('TOTAL'), tdb(str(len(qs))),
                       tdb(_fmt_brl(total_comissao)), tdb(_fmt_brl(total_pagar))])
    story.append(tbl(benef_rows, [W*0.48, W*0.12, W*0.20, W*0.20], total_last=True))

    # Lista completa
    story.append(PageBreak())
    story.append(Paragraph('Lista de Comissões', sec_s))
    list_rows = [[th('Nº'), th('RESERVA'), th('UNID.'), th('CLIENTE'),
                  th('IMOBILIÁRIA'), th('BENEFICIÁRIO'), th('TIPO'), th('%'), th('COMISSÃO'), th('A PAGAR')]]
    for c in sorted(qs, key=lambda x: (x.unidade, x.beneficiario.lower())):
        list_rows.append([
            td(c.numero), td(c.reserva), td(c.unidade),
            td(c.cliente[:22] if c.cliente else ''),
            td(c.imobiliaria[:20] if c.imobiliaria else ''),
            td(c.beneficiario[:20] if c.beneficiario else ''),
            td(c.tipo_comissao[:10]),
            tdr(f"{c.pct_comissao:.1f}%".replace('.', ',')),
            tdr(_fmt_brl(c.valor_comissao)),
            tdr(_fmt_brl(c.valor_comissao_pagar)),
        ])
    list_rows.append([tdb('TOTAL'), tdb(''), tdb(''), tdb(''), tdb(''), tdb(''), tdb(''), tdb(''),
                      tdb(_fmt_brl(total_comissao)), tdb(_fmt_brl(total_pagar))])
    story.append(tbl(list_rows,
                     [W*0.04, W*0.06, W*0.06, W*0.14, W*0.14, W*0.14, W*0.08, W*0.06, W*0.14, W*0.14],
                     total_last=True))

    doc.build(story)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/pdf')
    resp['Content-Disposition'] = 'attachment; filename="comissoes_cota365.pdf"'
    return resp
