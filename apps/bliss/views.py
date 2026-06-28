import csv
import io
import json
import os
import time
import unicodedata
import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.core.mail import EmailMessage
from django.core.validators import validate_email
from django.db import transaction
from django.db.models import F, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import get_template
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from openpyxl import load_workbook
from xhtml2pdf import pisa

from .forms import BlissForm
from .models import Bliss


WEBHOOK_PASSWORD = os.getenv("BLISS_RESUMO_WEBHOOK_PASSWORD", "12345")

@login_required
def bliss_unidades(request):
    registros = Bliss.objects.all()
    return render(request, 'bliss/bliss_unidades.html', {'registros': registros})

# Criar
@login_required
def bliss_create(request):
    if request.method == 'POST':
        form = BlissForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('bliss_unidades')
    else:
        form = BlissForm()
    return render(request, 'bliss/bliss_form.html', {'form': form})

# Editar
@login_required
def bliss_update(request, pk):
    registro = get_object_or_404(Bliss, pk=pk)
    form = BlissForm(request.POST or None, instance=registro)
    if form.is_valid():
        form.save()
        return redirect('bliss_unidades')
    return render(request, 'bliss/bliss_form.html', {'form': form})

# Excluir
@login_required
def bliss_delete(request, pk):
    registro = get_object_or_404(Bliss, pk=pk)
    if request.method == 'POST':
        registro.delete()
        return redirect('bliss_unidades')
    return render(request, 'bliss/bliss_delete_confirm.html', {'registro': registro})

# Relatório HTML
@login_required
def bliss_unidades_full(request):
    sort = request.GET.get('sort', 'bloco')
    direction = request.GET.get('dir', 'asc')  # 'asc' ou 'desc'

    allowed = {
        'bloco', 'unidade', 'area_privativa', 'garagem', 'deposito',
        'area_total', 'tipologia', 'situacao', 'valor_tabela',
        'data_venda', 'valor_venda', 'cliente', 'email', 'perc_permuta'
    }
    if sort not in allowed:
        sort = 'bloco'
    if direction not in ('asc', 'desc'):
        direction = 'asc'

    def with_dir(field: str) -> str:
        return field if direction == 'asc' else f'-{field}'

    # monta a ordenaÃ§Ã£o
    order_by = []
    if sort == 'bloco':
        # 1Âº por bloco, 2Âº por unidade (mesma direÃ§Ã£o escolhida)
        order_by.append(with_dir('bloco'))
        order_by.append(with_dir('unidade'))
    else:
        order_by.append(with_dir(sort))

    registros = Bliss.objects.all().order_by(*order_by)

    context = {
        'registros': registros,
        'sort': sort,
        'dir': direction,
        'invert_dir': 'desc' if direction == 'asc' else 'asc',
    }
    return render(request, 'bliss/bliss_unidades_full.html', context)

# Relatório PDF
@login_required
def bliss_unidades_full_pdf(request):
    registros = Bliss.objects.all()
    template = get_template('bliss/bliss_unidades_full_pdf.html')
    html = template.render({'registros': registros})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    return response

# Resumo de Situações
@login_required
def bliss_summary(request):
    ctx = _build_bliss_resumo_context()
    sit = ctx['situacao']
    r   = ctx['resumo']

    resumo = [
        {
            'situacao':           label,
            'total_valor_tabela': data['valor'],
            'total_unidades':     data['qtde'],
            'pct_valor_tabela':   data['valor_perc'],
            'pct_unidades':       data['qtde_perc'],
        }
        for label, data in sit.items()
        if label != 'Total'
    ]

    totais = {
        'valor_tabela': sit['Total']['valor'],
        'unidades':     sit['Total']['qtde'],
    }

    resumo_lojas = [
        {
            'preco_medio_tipo': None,
            'quantidade':       r['qtde_lojas'],
            'tipo':             'Loja',
            'm2':               r['priv_loja'],
            'valor_venda':      r['valor_loja'],
            'valor_m2':         r['m2_loja'],
        },
        {
            'preco_medio_tipo': r['preco_medio_tipo'],
            'quantidade':       r['qtde_tipos'],
            'tipo':             'Tipos',
            'm2':               r['priv_tipos'],
            'valor_venda':      r['valor_tipos'],
            'valor_m2':         r['m2_tipos'],
        },
        {
            'preco_medio_tipo': None,
            'quantidade':       '',
            'tipo':             'Total',
            'm2':               r['priv_total'],
            'valor_venda':      r['valor_total'],
            'valor_m2':         r['m2_total'],
        },
    ]

    return render(request, 'bliss/summary.html', {
        'resumo':       resumo,
        'totais':       totais,
        'resumo_lojas': resumo_lojas,
    })

def _aplicar_situacoes_fixas() -> int:
    """Sobrescreve situações específicas após qualquer importação de tabela."""
    _FIXAS = {
        'QA': [
            ('1-SUN', '201-SUN'),
            ('1-SUN', '206-SUN'),
            ('2-SHINE', '501-SHINE'),
        ],
        'Bloqueada': [
            ('1-SUN', '306-SUN'),
            ('1-SUN', '406-SUN'),
            ('2-SHINE', '305-SHINE'),
            ('2-SHINE', '405-SHINE'),
        ],
        'Permuta': [
            ('1-SUN', '101-SUN'),
            ('1-SUN', '303-SUN'),
            ('1-SUN', '505-SUN'),
            ('1-SUN', '701-SUN'),
            ('1-SUN', '802-SUN'),
            ('2-SHINE', '102-SHINE'),
            ('2-SHINE', '302-SHINE'),
            ('2-SHINE', '401-SHINE'),
            ('2-SHINE', '703-SHINE'),
        ],
        'Permuta/Venda': [
            ('1-SUNxxx', 'Loja'),
        ],
    }
    total = 0
    for situacao, pares in _FIXAS.items():
        for bloco, unidade in pares:
            total += Bliss.objects.filter(bloco=bloco, unidade=unidade).update(situacao=situacao)
    return total


@login_required
@require_POST
def atualizar_situacoes(request):
    total = _aplicar_situacoes_fixas()
    messages.success(request, f'{total} registros atualizados com sucesso.')
    return redirect('bliss_unidades')

# Importar planilha Excel
@login_required
@transaction.atomic
def bliss_import(request):
    excel_file = request.FILES.get('excel_file')
    if request.method != 'POST' or not excel_file:
        return render(request, 'bliss/bliss_import.html')

    try:
        wb = load_workbook(excel_file)
    except Exception:
        messages.error(request, 'Não foi possível ler o arquivo. Verifique se é um .xlsx válido.')
        return render(request, 'bliss/bliss_import.html')

    ws = wb.active
    objs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        objs.append(Bliss(
            bloco=row[0],
            unidade=row[1],
            area_privativa=row[2] or 0,
            garagem=row[3],
            deposito=row[4],
            tipologia=row[5],
            situacao=row[6] or '',
            valor_tabela=row[7] or 0,
            data_venda=row[8] if isinstance(row[8], datetime.date) else None,
            valor_venda=row[9] or 0,
            cliente=row[10] or '',
            email=row[11] or '',
        ))

    if not objs:
        messages.error(request, 'Nenhuma linha válida encontrada na planilha.')
        return render(request, 'bliss/bliss_import.html')

    Bliss.objects.all().delete()
    Bliss.objects.bulk_create(objs)
    messages.success(request, f'{len(objs)} unidades importadas com sucesso.')
    return redirect('bliss_unidades')


EXCECOES: set[tuple[str, str]] = set()  # populate to skip specific (bloco, unidade) pairs during atualizacao_mensal

def _parse_money_br(valor_str):
    if valor_str is None:
        return Decimal('0')
    s = str(valor_str).strip()
    if not s:
        return Decimal('0')
    s = s.replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal('0')

def _get_ci(row, key_lower):
    # getter case-insensitive para DictReader
    for k, v in row.items():
        if k and k.lower().strip() == key_lower:
            return v
    return None

@login_required
@transaction.atomic
def atualizacao_mensal(request):
    """
    Atualiza valor_tabela e situacao casando por (bloco, unidade).
    Aceita delimitadores ; , \t | e cabeÃ§alhos em qualquer ordem:
    bloco, unidade, valor_tabela, situacao
    """
    if request.method == 'POST' and request.FILES.get('csv_file'):
        f = request.FILES['csv_file']

        # decodifica respeitando BOM
        wrapper = io.TextIOWrapper(f.file, encoding='utf-8-sig', newline='')
        sample = wrapper.read(2048)
        wrapper.seek(0)

        # tenta detectar o delimitador; se falhar, forÃ§a ';'
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=';,|\t')
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ';'

        reader = csv.DictReader(wrapper, delimiter=delimiter)
        if not reader.fieldnames:
            messages.error(request, 'NÃ£o foi possÃ­vel ler os cabeÃ§alhos do CSV.')
            return redirect('bliss_unidades')

        obrigatorios = {'bloco', 'unidade', 'situação', 'valor total'}
        faltando = obrigatorios - {h.lower().strip() for h in reader.fieldnames if h}
        if faltando:
            messages.error(request, f'Cabeçalhos ausentes no CSV: {", ".join(sorted(faltando))}.')
            return redirect('bliss_unidades')

        total = puladas_excecao = nao_encontradas = sem_mudanca = 0
        objetos_para_update = []

        for row in reader:
            # pula linhas totalmente vazias
            if not any((row or {}).values()):
                continue

            total += 1
            bloco = (_get_ci(row, 'bloco') or '').strip()
            unidade = (_get_ci(row, 'unidade') or '').strip()
            if not bloco or not unidade:
                continue

            if (bloco, unidade) in EXCECOES:
                puladas_excecao += 1
                continue

            valor_tabela_csv = _parse_money_br(_get_ci(row, 'valor total'))
            situacao_csv = (_get_ci(row, 'situação') or '').strip()

            obj = Bliss.objects.filter(bloco=bloco, unidade=unidade).first()
            if not obj:
                nao_encontradas += 1
                continue

            mudou = False
            if obj.valor_tabela != valor_tabela_csv:
                obj.valor_tabela = valor_tabela_csv
                mudou = True
            if (obj.situacao or '') != situacao_csv:
                obj.situacao = situacao_csv
                mudou = True

            if mudou:
                objetos_para_update.append(obj)
            else:
                sem_mudanca += 1

        if objetos_para_update:
            Bliss.objects.bulk_update(objetos_para_update, ['valor_tabela', 'situacao'])

        fixas = _aplicar_situacoes_fixas()

        messages.success(
            request,
            (
                f'Processadas: {total}. '
                f'Atualizadas: {len(objetos_para_update)}. '
                f'Sem mudança: {sem_mudanca}. '
                f'Não encontradas: {nao_encontradas}. '
                f'Situações fixas aplicadas: {fixas}.'
            )
        )
        return redirect('bliss_unidades')

    return render(request, 'bliss/bliss_atualizacao_mensal.html')

import re as _re
import re

def _strip_hyperlink(value: str) -> str:
    """Extrai o texto de exibição de fórmulas =HIPERLINK(...) / =HYPERLINK(...)."""
    s = (value or '').strip()
    m = _re.match(r'=(?:HIPER|HYPER)LINK\([^;,]+[;,]"([^"]+)"\)', s, _re.IGNORECASE)
    return m.group(1).strip() if m else s


@login_required
@transaction.atomic
def bliss_import_clientes(request):
    """Importa CSV com UNIDADE, CLIENTE e E-MAIL e atualiza os registros existentes."""
    if request.method != 'POST' or not request.FILES.get('csv_file'):
        return render(request, 'bliss/bliss_import_clientes.html')

    f = request.FILES['csv_file']
    wrapper = io.TextIOWrapper(f.file, encoding='utf-8-sig', newline='')
    sample = wrapper.read(2048)
    wrapper.seek(0)

    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=';,|\t')
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ';'

    reader = csv.DictReader(wrapper, delimiter=delimiter)
    if not reader.fieldnames:
        messages.error(request, 'Não foi possível ler os cabeçalhos do CSV.')
        return render(request, 'bliss/bliss_import_clientes.html')

    cabecalhos = {h.lower().strip() for h in reader.fieldnames if h}
    if 'unidade' not in cabecalhos:
        messages.error(request, 'Coluna UNIDADE não encontrada no CSV.')
        return render(request, 'bliss/bliss_import_clientes.html')

    tem_bloco   = 'bloco'  in cabecalhos
    tem_cliente = 'cliente' in cabecalhos
    tem_email   = 'e-mail' in cabecalhos or 'email' in cabecalhos

    if not tem_cliente and not tem_email:
        messages.error(request, 'O CSV precisa ter ao menos a coluna CLIENTE ou E-MAIL.')
        return render(request, 'bliss/bliss_import_clientes.html')

    atualizados = nao_encontrados = sem_mudanca = 0
    objs = []

    for row in reader:
        if not any((row or {}).values()):
            continue

        unidade = (_get_ci(row, 'unidade') or '').strip()
        if not unidade:
            continue

        qs = Bliss.objects.filter(unidade=unidade)
        if tem_bloco:
            bloco = (_get_ci(row, 'bloco') or '').strip()
            if bloco:
                qs = qs.filter(bloco=bloco)

        obj = qs.first()
        if not obj:
            nao_encontrados += 1
            continue

        mudou = False
        if tem_cliente:
            novo = _strip_hyperlink(_get_ci(row, 'cliente') or '')
            if obj.cliente != novo:
                obj.cliente = novo
                mudou = True
        if tem_email:
            chave_email = 'e-mail' if 'e-mail' in cabecalhos else 'email'
            novo = _strip_hyperlink(_get_ci(row, chave_email) or '')
            if obj.email != novo:
                obj.email = novo
                mudou = True

        if mudou:
            objs.append(obj)
        else:
            sem_mudanca += 1

    campos = [f for f, ok in [('cliente', tem_cliente), ('email', tem_email)] if ok]
    if objs:
        Bliss.objects.bulk_update(objs, campos)
        atualizados = len(objs)

    messages.success(
        request,
        f'Atualizados: {atualizados}. Sem mudança: {sem_mudanca}. Não encontrados: {nao_encontrados}.'
    )
    return redirect('bliss_unidades_full')


def _build_bliss_resumo_context():
    registros = list(Bliss.objects.all())

    situacao = {
        'Disponível': {'qtde': 0, 'valor': Decimal('0'), 'area': Decimal('0')},
        'Reservada': {'qtde': 0, 'valor': Decimal('0'), 'area': Decimal('0')},
        'Bloqueada': {'qtde': 0, 'valor': Decimal('0'), 'area': Decimal('0')}, 
        'Vendida': {'qtde': 0, 'valor': Decimal('0'), 'area': Decimal('0')},
        'Permuta': {'qtde': 0, 'valor': Decimal('0'), 'area': Decimal('0')},
        'QA': {'qtde': 0, 'valor': Decimal('0'), 'area': Decimal('0')},
        'Total': {'qtde': 0, 'valor': Decimal('0'), 'area': Decimal('0')},
    }

    status_labels = [label for label in situacao.keys() if label != 'Total']

    def normalizar_status(valor: str) -> str:
        if not valor:
            return ''
        base = unicodedata.normalize('NFKD', valor.strip())
        sem_acentos = ''.join(ch for ch in base if not unicodedata.combining(ch))
        return sem_acentos.casefold()

    status_lookup: dict[str, str] = {}

    def registrar_lookup(label: str) -> None:
        chave_norm = normalizar_status(label)
        if not chave_norm:
            return
        if chave_norm not in status_lookup:
            status_lookup[chave_norm] = label
        chave_lower = label.strip().casefold()
        if chave_lower not in status_lookup:
            status_lookup[chave_lower] = label

    for label in status_labels:
        registrar_lookup(label)

    status_disponivel_norm = normalizar_status('Disponivel')

    agrupadas = {label: [] for label in status_labels}
    fator_permuta = Decimal('0.12826')
    fator_restante = Decimal('0.87174')

    def adicionar_unidade(label, valor_calculado, area_calculada, registro):
        label = (label or '').strip()
        if not label:
            return
        chave_norm = normalizar_status(label)
        chave_mapeada = status_lookup.get(chave_norm) or status_lookup.get(label.casefold()) or label
        registrar_lookup(chave_mapeada)
        agrupadas.setdefault(chave_mapeada, []).append({
            'id': registro.id,
            'bloco': registro.bloco,
            'unidade': registro.unidade,
            'situacao': chave_mapeada,
            'valor_tabela': valor_calculado,
            'area_privativa': area_calculada,
            'garagem': registro.garagem,
            'deposito': registro.deposito,
            'tipologia': registro.tipologia,
        })

    for registro in registros:
        valor = registro.valor_tabela or Decimal('0')
        area = registro.area_privativa or Decimal('0')
        unidade_nome = (registro.unidade or '').strip().lower()
        situacao_registro = (registro.situacao or '').strip()

        if unidade_nome == 'loja':
            adicionar_unidade('Permuta', valor * fator_permuta, area * fator_permuta, registro)
            adicionar_unidade(situacao_registro, valor * fator_restante, area * fator_restante, registro)
        else:
            adicionar_unidade(situacao_registro, valor, area, registro)

    def ordenar_unidades(itens):
        return sorted(
            itens,
            key=lambda item: (
                (item.get('bloco') or ''),
                (item.get('unidade') or ''),
            ),
        )

    unidades = [
        {'situacao': label, 'unidades': ordenar_unidades(agrupadas.get(label, []))}
        for label in status_labels
    ]

    extras = [label for label in agrupadas.keys() if label not in status_labels]
    for label in extras:
        registrar_lookup(label)
        unidades.append({'situacao': label, 'unidades': ordenar_unidades(agrupadas[label])})

    default_totais = {
        'qtde': 0,
        'valor': Decimal('0'),
        'area': Decimal('0'),
        'qtde_perc': Decimal('0'),
        'valor_perc': Decimal('0'),
    }
    unidades = [
        {**item, 'totais': situacao.get(item['situacao'], default_totais)}
        for item in unidades
    ]
    unidades_por_status = {item['situacao']: item['unidades'] for item in unidades}

    def to_float(value):
        if value is None:
            return 0.0
        try:
            return float(value)
        except (TypeError, ValueError):
            try:
                return float(str(value).replace(',', '.'))
            except (TypeError, ValueError):
                return 0.0

    resumo = {
        'preco_medio_tipo': 0,
        'qtde_lojas': 0,
        'qtde_tipos': 0,
        'qtde_total': 0,

        'valor_loja': 0,
        'valor_tipos': 0,
        'valor_total': 0,

        'priv_loja': 0,
        'priv_tipos': 0,
        'priv_avenda': 0,

        'm2_loja': 0,
        'm2_tipos': 0,
        'm2_avenda': 0,
    }

    for registro in registros:
        chave = (registro.situacao or '').strip()
        valor = registro.valor_tabela or Decimal('0')
        area = registro.area_privativa or Decimal('0')

        if registro.unidade == 'Loja':
            valor_permuta = valor * fator_permuta
            valor_restante = valor * fator_restante
            area_permuta = area * fator_permuta
            area_restante = area * fator_restante

            situacao["Permuta"]['qtde'] += 1
            situacao["Permuta"]['valor'] += valor_permuta
            situacao["Permuta"]['area'] += area_permuta

            destino = (registro.situacao or '').strip()
            destino_normalizado = status_lookup.get(destino.casefold(), destino) if destino else ''
            if destino_normalizado:
                if destino_normalizado not in situacao:
                    situacao[destino_normalizado] = {
                        'qtde': 0,
                        'valor': Decimal('0'),
                        'area': Decimal('0'),
                    }
                    registrar_lookup(destino_normalizado)
                situacao[destino_normalizado]['qtde'] += 1
                situacao[destino_normalizado]['valor'] += valor_restante
                situacao[destino_normalizado]['area'] += area_restante

            if normalizar_status(destino_normalizado) == status_disponivel_norm:
                resumo["qtde_lojas"] = 1
                resumo["valor_loja"] = valor_restante
                resumo["priv_loja"] = area_restante

        else:
            if chave and chave not in situacao:
                situacao[chave] = {
                    'qtde': 0,
                    'valor': Decimal('0'),
                    'area': Decimal('0'),
                }
                registrar_lookup(chave)
            if chave in situacao:
                situacao[chave]['qtde'] += 1
                situacao[chave]['valor'] += valor
                situacao[chave]['area'] += area
                if normalizar_status(registro.situacao) == status_disponivel_norm:
                    resumo["qtde_tipos"] += 1
                    resumo["priv_tipos"] += area
                    resumo["valor_tipos"] += registro.valor_tabela
        situacao['Total']['qtde'] += 1
        situacao['Total']['valor'] += valor
        situacao['Total']['area'] += area

    total_qtde = situacao['Total']['qtde']
    total_valor = situacao['Total']['valor']

    if resumo["qtde_tipos"]:
        resumo["preco_medio_tipo"] = resumo["valor_tipos"] / resumo["qtde_tipos"]
    else:
        resumo["preco_medio_tipo"] = 0    

    if resumo["priv_loja"]:
        resumo["m2_loja"] = resumo["valor_loja"] / resumo["priv_loja"] 
    else:
        resumo["m2_loja"] = 0   

    if resumo["priv_tipos"]:
        resumo["m2_tipos"] = resumo["valor_tipos"] / resumo["priv_tipos"] 
    else:
        resumo["m2_tipos"] = 0 

    resumo["qtde_total"] = resumo["qtde_lojas"] + resumo["qtde_tipos"]  
    resumo["valor_total"] = resumo["valor_loja"] + resumo["valor_tipos"] 
    resumo["priv_total"] = resumo["priv_loja"] + resumo["priv_tipos"] 

    if resumo["priv_total"]:
        resumo["m2_total"] = resumo["valor_total"] / resumo["priv_total"] 
    else:
        resumo["m2_total"] = 0 

    for dados in situacao.values():
        dados['qtde_perc'] = (Decimal(dados['qtde']) / Decimal(total_qtde) * Decimal('100')) if total_qtde else Decimal('0')
        dados['valor_perc'] = (dados['valor'] / total_valor * Decimal('100')) if total_valor else Decimal('0')
        dados['area_perc'] = (dados['area'] / situacao['Total']['area'] * Decimal('100')) if situacao['Total']['area'] else Decimal('0')

    chart_data = {
        'situacao_labels': [item['situacao'] for item in unidades],
        'situacao_valores': [to_float((situacao.get(item['situacao']) or {}).get('valor')) for item in unidades],
        'situacao_areas': [to_float((situacao.get(item['situacao']) or {}).get('area')) for item in unidades],
        'situacao_qtdes': [int((situacao.get(item['situacao']) or {}).get('qtde') or 0) for item in unidades],
        'tipo_labels': ['Tipos', 'Lojas'],
        'tipo_valores': [to_float(resumo.get('valor_tipos')), to_float(resumo.get('valor_loja'))],
        'tipo_areas': [to_float(resumo.get('priv_tipos')), to_float(resumo.get('priv_loja'))],
    }


    return {
        'situacao': situacao,
        'registros': registros,
        'unidades': unidades,
        'chart_data': chart_data,
        'unidades_por_status': unidades_por_status,
        'resumo': resumo,
    }

def _render_bliss_resumo_pdf(context):
    template = get_template('bliss/bliss_resumo_pdf.html')
    html = template.render(context)
    buffer = BytesIO()
    pisa_status = pisa.CreatePDF(html, dest=buffer)
    if pisa_status.err:
        raise ValueError('Erro ao gerar PDF do resumo Bliss')
    buffer.seek(0)
    return buffer


def send_bliss_resumo_pdf_email(recipient_email, *, subject=None, body=None):
    context = _build_bliss_resumo_context()
    pdf_buffer = _render_bliss_resumo_pdf(context)
    email = EmailMessage(
        subject or 'Resumo Bliss',
        body or 'Segue em anexo o resumo Bliss.',
        to=[recipient_email],
    )
    email.attach('bliss_resumo.pdf', pdf_buffer.getvalue(), 'application/pdf')
    sent_count = email.send(fail_silently=False)
    if sent_count == 0:
        raise ValueError('Nenhum e-mail foi enviado.')


def _attach_email_flag(request, context):
    sent_info = request.session.get('bliss_resumo_email_sent')
    email_flag = False
    if isinstance(sent_info, dict):
        ts = sent_info.get('ts')
        if ts and (time.time() - ts) <= 5:
            email_flag = sent_info.get('email') or True
        else:
            request.session.pop('bliss_resumo_email_sent', None)
    elif sent_info:
        email_flag = True
        request.session.pop('bliss_resumo_email_sent', None)
    context['email_enviado'] = email_flag
    return context


@login_required
def bliss_resumo(request):
    context = _attach_email_flag(request, _build_bliss_resumo_context())
    return render(request, 'bliss/bliss_resumo.html', context)



@login_required
def bliss_dashboard(request):
    context = _attach_email_flag(request, _build_bliss_resumo_context())
    return render(request, 'bliss/bliss_dashboard.html', context)

@login_required
def bliss_resumo_pdf(request):
    context = _build_bliss_resumo_context()
    try:
        pdf_buffer = _render_bliss_resumo_pdf(context)
    except ValueError as exc:
        return HttpResponse(str(exc), status=500)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="bliss_resumo.pdf"'
    response.write(pdf_buffer.getvalue())
    return response


@csrf_exempt
def bliss_resumo_email_webhook(request):
    if request.method != 'POST':
        return JsonResponse({'detail': 'Metodo nao permitido'}, status=405)

    try:
        raw_body = request.body.decode('utf-8') if request.body else ''
        payload = json.loads(raw_body) if raw_body else {}
    except json.JSONDecodeError:
        payload = request.POST

    email = (payload.get('email') or payload.get('to') or '').strip()
    password = (payload.get('senha') or payload.get('password') or '').strip()

    if password != WEBHOOK_PASSWORD:
        return JsonResponse({'detail': 'Credenciais invalidas'}, status=403)

    if not email:
        return JsonResponse({'detail': 'E-mail nao informado'}, status=400)

    try:
        validate_email(email)
    except ValidationError:
        return JsonResponse({'detail': 'E-mail invalido'}, status=400)

    try:
        send_bliss_resumo_pdf_email(email)
    except ValueError as exc:
        return JsonResponse({'detail': str(exc)}, status=500)
    except Exception:
        return JsonResponse({'detail': 'Erro inesperado ao enviar o e-mail'}, status=500)

    return JsonResponse({'status': 'ok'})


@login_required
def bliss_resumo_send_email(request):
    if request.method != 'POST':
        return redirect('bliss_resumo')

    email = (request.user.email or '').strip()
    if not email:
        messages.error(request, 'Seu usuario nao possui e-mail cadastrado.')
        return redirect('bliss_resumo')

    try:
        send_bliss_resumo_pdf_email(email)
    except ValueError as exc:
        messages.error(request, str(exc))
    except Exception:
        messages.error(request, 'Nao foi possivel enviar o resumo por e-mail.')
    else:
        messages.success(request, f'Resumo enviado para {email}.')
        request.session['bliss_resumo_email_sent'] = {'email': email, 'ts': time.time()}

    return redirect('bliss_resumo')


@csrf_exempt
def bliss_resumo_test_email(request):
    if request.method not in ('POST', 'GET'):
        return JsonResponse({'detail': 'Metodo nao permitido'}, status=405)

    if request.method == 'GET':
        payload = request.GET
    else:
        try:
            raw_body = request.body.decode('utf-8') if request.body else ''
            payload = json.loads(raw_body) if raw_body else {}
        except json.JSONDecodeError:
            payload = request.POST

    email = (payload.get('email') or '').strip()
    password = (payload.get('senha') or payload.get('password') or '').strip()

    if password != WEBHOOK_PASSWORD:
        return JsonResponse({'detail': 'Credenciais invalidas'}, status=403)

    if not email:
        return JsonResponse({'detail': 'E-mail nao informado'}, status=400)

    try:
        validate_email(email)
    except ValidationError:
        return JsonResponse({'detail': 'E-mail invalido'}, status=400)

    email_message = EmailMessage(
        'Teste de envio Bliss',
        'Este e um e-mail de teste do sistema Bliss.',
        to=[email],
    )

    try:
        sent = email_message.send(fail_silently=False)
    except Exception:
        return JsonResponse({'detail': 'Erro ao enviar e-mail de teste'}, status=500)

    if sent == 0:
        return JsonResponse({'detail': 'Nenhum e-mail foi enviado'}, status=500)

    return JsonResponse({'status': 'ok', 'message': 'E-mail de teste enviado'})


# ---------------------------------------------------------------------------
# Cartório Bliss Living
# ---------------------------------------------------------------------------

from django.conf import settings as _settings

_CARTORIO_BLISS_XLSX = _settings.BASE_DIR / 'cartorio - bliss.xlsx'


def _bliss_fmt_area(v):
    if v is None:
        return '—'
    return f'{v:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')


def _bliss_fmt_fi(v):
    if v is None:
        return '—'
    return f'{v:.3f}%'.replace('.', ',')


def _bliss_cartorio_parse_keys(col_val, skip_words=('sem vaga', 'sem hobby')):
    if not col_val:
        return []
    s = str(col_val).strip()
    if any(s.lower().startswith(w) for w in skip_words):
        return []
    s = re.sub(r'\s+c/', ';', s, flags=re.IGNORECASE)
    result = []
    for item in s.split(';'):
        item = item.strip().upper()
        item = re.sub(r'^GPCD\s+', 'G', item)
        item = item.replace(' ', '')
        if item:
            result.append(item)
    return result


def _bliss_cartorio_load_grouped():
    import openpyxl as _openpyxl
    from pathlib import Path
    if not Path(_CARTORIO_BLISS_XLSX).exists():
        raise FileNotFoundError(_CARTORIO_BLISS_XLSX)
    wb = _openpyxl.load_workbook(_CARTORIO_BLISS_XLSX)
    ws = wb['cartorio']
    rows = list(ws.iter_rows(values_only=True))

    # Normaliza header para lowercase para evitar problemas de case do xlsx
    header = [str(h).strip().lower() if h is not None else None for h in rows[0]]
    data = [dict(zip(header, r)) for r in rows[1:] if r[0] is not None]
    data = [r for r in data if r.get('tipo')]

    def _get(r, *keys):
        """Busca coluna ignorando acentuação e case (já normalizados no header)."""
        for k in keys:
            v = r.get(k.lower())
            if v is not None:
                return v
        return None

    def make_unit(r, stub=False):
        return {
            'unidade':    _get(r, 'unidade'),
            'localizacao': _get(r, 'localização', 'localizacao', 'localização'),
            'tipologia':  _get(r, 'tipologia'),
            'ap':         _get(r, 'área privativa', 'area privativa'),
            'apa':        _get(r, 'área privativa acessória', 'area privativa acessoria'),
            'apt':        _get(r, 'área privativa total', 'area privativa total'),
            'ac':         _get(r, 'área de uso comum', 'area de uso comum'),
            'art':        _get(r, 'área real total', 'area real total'),
            'fi':         _get(r, 'fração ideal', 'fracao ideal'),
            'matricula':  _get(r, 'matricula', 'matrícula'),
            'vinculo':    _get(r, 'vinculo-matricula', 'vínculo-matricula'),
            'stub':       stub,
        }

    comp_dict = {}
    for r in data:
        if (r.get('tipo') or '').lower() in ('garagem', 'hobby box', 'moto'):
            key = str(_get(r, 'unidade') or '').strip().upper().replace(' ', '')
            comp_dict[key] = make_unit(r)

    def resolve(key, garagens_do_grupo):
        if key in comp_dict:
            return comp_dict[key]
        for g in garagens_do_grupo:
            vinculo_parts = [x.strip() for x in (g.get('vinculo') or '').split(';')]
            if key in vinculo_parts:
                return {
                    'unidade': key, 'localizacao': g.get('localizacao'),
                    'tipologia': None, 'ap': g.get('apa'), 'apa': None,
                    'apt': g.get('apa'), 'ac': None, 'art': None,
                    'fi': None, 'matricula': g.get('matricula'), 'vinculo': None, 'stub': True,
                }
        return {
            'unidade': key, 'localizacao': None, 'tipologia': None,
            'ap': None, 'apa': None, 'apt': None, 'ac': None, 'art': None,
            'fi': None, 'matricula': None, 'vinculo': None, 'stub': True,
        }

    groups_apt, groups_loja = [], []
    for r in data:
        tipo = (r.get('tipo') or '').lower()
        if tipo not in ('apartamento', 'loja'):
            continue
        gar_keys = _bliss_cartorio_parse_keys(_get(r, 'garagens'), ('sem vaga',))
        hb_keys  = _bliss_cartorio_parse_keys(_get(r, 'hbs', 'hb'), ('sem hobby',))
        garagens = [comp_dict.get(k, {'unidade': k, 'stub': True, **{f: None for f in
                    ('localizacao','tipologia','ap','apa','apt','ac','art','fi','matricula','vinculo')}})
                    for k in gar_keys]
        hbs = [resolve(k, garagens) for k in hb_keys]
        group = {'principal': make_unit(r), 'garagens': garagens, 'hobby_boxes': hbs}
        (groups_apt if tipo == 'apartamento' else groups_loja).append(group)

    return {'apartamentos': groups_apt, 'lojas': groups_loja}


def bliss_cartorio_view(request):
    try:
        grupos = _bliss_cartorio_load_grouped()
    except FileNotFoundError:
        return HttpResponse('cartorio - bliss.xlsx não encontrado na raiz do projeto.', status=404)
    totais = {k: len(v) for k, v in grupos.items()}
    return render(request, 'bliss/cartorio.html', {'grupos': grupos, 'totais': totais})


def bliss_cartorio_pdf(request):
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, PageBreak
    from django.http import FileResponse
    from datetime import datetime

    try:
        grupos = _bliss_cartorio_load_grouped()
    except FileNotFoundError:
        return HttpResponse('cartorio - bliss.xlsx não encontrado.', status=404)

    import io as _io
    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.2*cm, rightMargin=1.2*cm,
                            topMargin=1.4*cm, bottomMargin=1.4*cm)
    styles = getSampleStyleSheet()

    title_s = ParagraphStyle('bt', parent=styles['Heading1'], fontSize=13, spaceAfter=3)
    sub_s   = ParagraphStyle('bs', parent=styles['Normal'], fontSize=8, spaceAfter=10,
                              textColor=colors.HexColor('#6c757d'))
    sec_s   = ParagraphStyle('be', parent=styles['Heading2'], fontSize=10, spaceAfter=4,
                              spaceBefore=8, textColor=colors.HexColor('#1a1a2e'))
    cell_s  = ParagraphStyle('bc', parent=styles['Normal'], fontSize=6.5, leading=8)
    nr_s    = ParagraphStyle('bn', parent=styles['Normal'], fontSize=6.5, leading=8, alignment=2)
    hdr_s   = ParagraphStyle('bh', parent=styles['Normal'], fontSize=6, leading=7.5, alignment=1,
                              textColor=colors.white)
    mat_s   = ParagraphStyle('bm', parent=styles['Normal'], fontSize=5.5, leading=7,
                              textColor=colors.HexColor('#444'))
    unit_p  = ParagraphStyle('bu', parent=styles['Normal'], fontSize=10, leading=12,
                              fontName='Helvetica-Bold', textColor=colors.HexColor('#1a1a2e'))

    HEADERS = [
        'Unidade', 'Localização', 'Tipologia',
        'Área Priv.\n(m²)', 'Área Priv.\nAces. (m²)', 'Área Priv.\nTotal (m²)',
        'Área\nComum (m²)', 'Área Real\nTotal (m²)', 'Fração\nIdeal (%)',
        'Matrícula',
    ]
    W = doc.width
    fixed_w = [1.8, 3.0, 1.8, 1.8, 1.8, 1.8, 1.8, 1.9, 1.6]
    mat_w = W - sum(x * cm for x in fixed_w)
    COL_W = [x * cm for x in fixed_w] + [mat_w]

    C_PRINCIPAL = colors.HexColor('#dbeafe')
    C_GARAGEM   = colors.HexColor('#fff3e0')
    C_HB        = colors.HexColor('#f3e5f5')
    C_SEP       = colors.HexColor('#f8f9fa')
    NAVY        = colors.HexColor('#1a1a2e')

    def _fmt_unidade(u):
        """Remove sufixo após '-' (ex: 101-SUN → 101, G46 → G46)."""
        nome = str(u['unidade'] or '')
        return nome.split('-')[0] if '-' in nome else nome

    def make_row(u, principal=False):
        return [
            Paragraph(_fmt_unidade(u), unit_p if principal else cell_s),
            Paragraph(str(u['localizacao'] or ''), cell_s),
            Paragraph(str(u['tipologia'] or ''), cell_s),
            Paragraph(_bliss_fmt_area(u['ap']),  nr_s),
            Paragraph(_bliss_fmt_area(u['apa']), nr_s),
            Paragraph(_bliss_fmt_area(u['apt']), nr_s),
            Paragraph(_bliss_fmt_area(u['ac']),  nr_s),
            Paragraph(_bliss_fmt_area(u['art']), nr_s),
            Paragraph(_bliss_fmt_fi(u['fi']),    nr_s),
            Paragraph(str(u['matricula'] or ''), mat_s),
        ]

    def make_blank_row():
        return [Paragraph('', cell_s)] * len(HEADERS)

    def build_table(group_list):
        hrow = [Paragraph(f'<b>{h}</b>', hdr_s) for h in HEADERS]
        table_data, row_colors = [hrow], []
        for group in group_list:
            p_idx = len(table_data)
            table_data.append(make_row(group['principal'], principal=True))
            row_colors.append((p_idx, C_PRINCIPAL))
            for g in group['garagens']:
                idx = len(table_data); table_data.append(make_row(g))
                row_colors.append((idx, C_GARAGEM))
            for h in group['hobby_boxes']:
                idx = len(table_data); table_data.append(make_row(h))
                row_colors.append((idx, C_HB))
            sep = len(table_data); table_data.append(make_blank_row())
            row_colors.append((sep, C_SEP))
        t = Table(table_data, colWidths=COL_W, repeatRows=1)
        cmds = [
            ('BACKGROUND',    (0, 0), (-1, 0), NAVY),
            ('GRID',          (0, 0), (-1, -1), 0.3, colors.HexColor('#dee2e6')),
            ('TOPPADDING',    (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('VALIGN',        (0, 0), (-1, -1), 'TOP'),
        ]
        for ri, c in row_colors:
            cmds.append(('BACKGROUND', (0, ri), (-1, ri), c))
        t.setStyle(TableStyle(cmds))
        return t

    story = [
        Paragraph('Bliss Living — Matrícula Cartório', title_s),
        Paragraph(f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}', sub_s),
    ]
    first = True
    for key, label in [('apartamentos', 'APARTAMENTOS'), ('lojas', 'LOJAS')]:
        group_list = grupos.get(key, [])
        if not group_list:
            continue
        if not first:
            story.append(PageBreak())
        first = False
        story.append(Paragraph(f'{label} — {len(group_list)} unidades', sec_s))
        story.append(build_table(group_list))

    doc.build(story)
    buf.seek(0)
    resp = FileResponse(buf, content_type='application/pdf')
    resp['Content-Disposition'] = 'inline; filename="cartorio_bliss.pdf"'
    return resp



