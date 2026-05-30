from datetime import date

from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect

from .models import Empresa, Empreendimento, Bloco, Unidade, TabelaVendas, SeriePagamento, ItemTabelaVendas, ValorSerie, ImportLog
from .forms import EmpresaForm, EmpreendimentoForm, BlocoForm, UnidadeForm, TabelaVendasForm, SeriePagamentoForm
from .utils import render_to_pdf


# ── Empresa ──────────────────────────────────────────────────────────────────

@login_required
def empresa_list(request):
    empresas = Empresa.objects.all()
    return render(request, 'incorporadora/empresa_list.html', {'empresas': empresas})


@login_required
def empresa_create(request):
    form = EmpresaForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Empresa cadastrada com sucesso.')
        return redirect('incorporadora:empresa_list')
    return render(request, 'incorporadora/empresa_form.html', {'form': form, 'titulo': 'Nova Empresa'})


@login_required
def empresa_edit(request, pk):
    empresa = get_object_or_404(Empresa, pk=pk)
    form = EmpresaForm(request.POST or None, instance=empresa)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Empresa atualizada com sucesso.')
        return redirect('incorporadora:empresa_list')
    return render(request, 'incorporadora/empresa_form.html', {'form': form, 'titulo': 'Editar Empresa', 'empresa': empresa})


@login_required
def empresa_delete(request, pk):
    empresa = get_object_or_404(Empresa, pk=pk)
    if request.method == 'POST':
        empresa.delete()
        messages.success(request, 'Empresa excluída com sucesso.')
        return redirect('incorporadora:empresa_list')
    return render(request, 'incorporadora/empresa_confirm_delete.html', {'empresa': empresa})


# ── Empreendimento ───────────────────────────────────────────────────────────

@login_required
def empreendimento_list(request):
    empreendimentos = Empreendimento.objects.select_related('empresa').all()
    return render(request, 'incorporadora/empreendimento_list.html', {'empreendimentos': empreendimentos})


@login_required
def empreendimento_create(request):
    form = EmpreendimentoForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Empreendimento cadastrado com sucesso.')
        return redirect('incorporadora:empreendimento_list')
    return render(request, 'incorporadora/empreendimento_form.html', {'form': form, 'titulo': 'Novo Empreendimento'})


@login_required
def empreendimento_edit(request, pk):
    obj = get_object_or_404(Empreendimento, pk=pk)
    form = EmpreendimentoForm(request.POST or None, instance=obj)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Empreendimento atualizado com sucesso.')
        return redirect('incorporadora:empreendimento_list')
    return render(request, 'incorporadora/empreendimento_form.html', {'form': form, 'titulo': 'Editar Empreendimento', 'obj': obj})


@login_required
def empreendimento_delete(request, pk):
    obj = get_object_or_404(Empreendimento, pk=pk)
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Empreendimento excluído com sucesso.')
        return redirect('incorporadora:empreendimento_list')
    return render(request, 'incorporadora/empreendimento_confirm_delete.html', {'obj': obj})


# ── Bloco ─────────────────────────────────────────────────────────────────────

def _build_vinculos_rows(bloco):
    def _split_aliases(u):
        if not u.numeros_adicionais:
            return [], []
        parts = [a.strip() for a in u.numeros_adicionais.split(',') if a.strip()]
        hb     = [a for a in parts if a.upper().startswith('HB')]
        outros = [a for a in parts if not a.upper().startswith('HB')]
        return outros, hb

    def _display(numero, aliases):
        return f'{numero} ({" · ".join(aliases)})' if aliases else numero

    principais = (
        bloco.unidades
        .filter(tipo__in=['apartamento', 'sala', 'loja'])
        .prefetch_related('vinculadas')
        .order_by('ordem', 'numero')
    )
    rows = []
    for u in principais:
        vinculadas = list(u.vinculadas.all())
        garagens, hbs, outros = [], [], []
        for v in vinculadas:
            outros_aliases, hb_aliases = _split_aliases(v)
            if v.tipo == 'garagem':
                garagens.append(_display(v.numero, outros_aliases) + ('*' if hb_aliases else ''))
                hbs.extend(hb_aliases)
            elif v.tipo == 'hobby_box':
                hbs.append(_display(v.numero, outros_aliases))
            else:
                outros.append(_display(v.numero, outros_aliases + hb_aliases))
        if vinculadas:
            rows.append({'unidade': u, 'garagens': garagens, 'hbs': hbs, 'outros': outros})
    return rows


@login_required
def bloco_list(request, empreendimento_pk):
    empreendimento = get_object_or_404(Empreendimento, pk=empreendimento_pk)
    blocos = empreendimento.blocos.prefetch_related('unidades__vinculadas').all()
    blocos_data = [{'bloco': b, 'vinculos': _build_vinculos_rows(b)} for b in blocos]
    return render(request, 'incorporadora/bloco_list.html', {
        'empreendimento': empreendimento,
        'blocos': blocos,
        'blocos_data': blocos_data,
    })


@login_required
def bloco_create(request, empreendimento_pk):
    empreendimento = get_object_or_404(Empreendimento, pk=empreendimento_pk)
    form = BlocoForm(request.POST or None, initial={'empreendimento': empreendimento})
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Bloco cadastrado com sucesso.')
        return redirect('incorporadora:bloco_list', empreendimento_pk=empreendimento.pk)
    return render(request, 'incorporadora/bloco_form.html', {
        'form': form, 'titulo': 'Novo Bloco', 'empreendimento': empreendimento,
    })


@login_required
def bloco_edit(request, pk):
    bloco = get_object_or_404(Bloco, pk=pk)
    form = BlocoForm(request.POST or None, instance=bloco)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Bloco atualizado com sucesso.')
        return redirect('incorporadora:bloco_list', empreendimento_pk=bloco.empreendimento.pk)
    return render(request, 'incorporadora/bloco_form.html', {
        'form': form, 'titulo': 'Editar Bloco', 'empreendimento': bloco.empreendimento, 'obj': bloco,
    })


@login_required
def bloco_delete(request, pk):
    bloco = get_object_or_404(Bloco, pk=pk)
    empreendimento_pk = bloco.empreendimento.pk
    if request.method == 'POST':
        bloco.delete()
        messages.success(request, 'Bloco excluído com sucesso.')
        return redirect('incorporadora:bloco_list', empreendimento_pk=empreendimento_pk)
    return render(request, 'incorporadora/bloco_confirm_delete.html', {'obj': bloco})


# ── Unidade ───────────────────────────────────────────────────────────────────

@login_required
def unidade_list(request, bloco_pk):
    bloco    = get_object_or_404(Bloco.objects.select_related('empreendimento'), pk=bloco_pk)
    unidades = bloco.unidades.order_by('ordem')

    f_numero    = request.GET.get('numero', '').strip()
    f_tipo      = request.GET.get('tipo', '').strip()
    f_tipologia = request.GET.get('tipologia', '').strip()
    f_local     = request.GET.get('localizacao', '').strip()
    f_status    = request.GET.get('status', '').strip()

    if f_numero:
        unidades = unidades.filter(numero__icontains=f_numero)
    if f_tipo:
        unidades = unidades.filter(tipo=f_tipo)
    if f_tipologia:
        unidades = unidades.filter(tipologia__icontains=f_tipologia)
    if f_local:
        unidades = unidades.filter(localizacao__icontains=f_local)
    if f_status:
        unidades = unidades.filter(status=f_status)

    filtros_ativos = any([f_numero, f_tipo, f_tipologia, f_local, f_status])

    return render(request, 'incorporadora/unidade_list.html', {
        'bloco':          bloco,
        'unidades':       unidades,
        'tipo_choices':   Unidade.TIPO_CHOICES,
        'status_choices': Unidade.STATUS_CHOICES,
        'filtros':        {'numero': f_numero, 'tipo': f_tipo, 'tipologia': f_tipologia,
                           'localizacao': f_local, 'status': f_status},
        'filtros_ativos': filtros_ativos,
    })


@login_required
def unidade_create(request, bloco_pk):
    bloco = get_object_or_404(Bloco.objects.select_related('empreendimento'), pk=bloco_pk)
    form = UnidadeForm(request.POST or None, initial={'bloco': bloco}, bloco=bloco)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Unidade cadastrada com sucesso.')
        return redirect('incorporadora:unidade_list', bloco_pk=bloco.pk)
    return render(request, 'incorporadora/unidade_form.html', {'form': form, 'titulo': 'Nova Unidade', 'bloco': bloco})


@login_required
def unidade_edit(request, pk):
    unidade = get_object_or_404(Unidade.objects.select_related('bloco__empreendimento'), pk=pk)
    form = UnidadeForm(request.POST or None, instance=unidade, bloco=unidade.bloco)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Unidade atualizada com sucesso.')
        return redirect('incorporadora:unidade_list', bloco_pk=unidade.bloco.pk)
    return render(request, 'incorporadora/unidade_form.html', {
        'form': form, 'titulo': 'Editar Unidade', 'bloco': unidade.bloco, 'obj': unidade,
    })


@login_required
def unidade_delete(request, pk):
    unidade = get_object_or_404(Unidade.objects.select_related('bloco__empreendimento'), pk=pk)
    bloco_pk = unidade.bloco.pk
    if request.method == 'POST':
        unidade.delete()
        messages.success(request, 'Unidade excluída com sucesso.')
        return redirect('incorporadora:unidade_list', bloco_pk=bloco_pk)
    return render(request, 'incorporadora/unidade_confirm_delete.html', {'obj': unidade})


# ── PDF exports ───────────────────────────────────────────────────────────────

@login_required
def empreendimento_export_excel(request):
    return _empreendimento_excel(request, empreendimento=None)


@login_required
def empreendimento_export_excel_pk(request, pk):
    empreendimento = get_object_or_404(Empreendimento, pk=pk)
    return _empreendimento_excel(request, empreendimento=empreendimento)


def _empreendimento_excel(request, empreendimento):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Unidades'

    headers = [
        'Empreendimento', 'Bloco', 'Ordem', 'Número', 'Números Adicionais', 'Tipo', 'Tipologia', 'Localização',
        'Área Privativa (m²)', 'Área Priv. Acess. (m²)', 'Área Comum (m²)',
        'Fração Ideal', 'Valor Tabela (R$)', 'Status', 'Cliente',
        'Descrição 1', 'Descrição 2', 'Descrição 3',
    ]
    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill(start_color='A7A3AB', end_color='A7A3AB', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')

    qs = Unidade.objects.select_related('bloco__empreendimento')
    if empreendimento:
        qs = qs.filter(bloco__empreendimento=empreendimento)
    qs = qs.order_by('bloco__empreendimento__nome', 'bloco__nome', 'ordem', 'numero')

    for row, u in enumerate(qs, 2):
        ws.cell(row=row, column=1,  value=u.bloco.empreendimento.nome)
        ws.cell(row=row, column=2,  value=u.bloco.nome)
        ws.cell(row=row, column=3,  value=u.ordem)
        ws.cell(row=row, column=4,  value=u.numero)
        ws.cell(row=row, column=5,  value=u.numeros_adicionais)
        ws.cell(row=row, column=6,  value=u.get_tipo_display())
        ws.cell(row=row, column=7,  value=u.tipologia)
        ws.cell(row=row, column=8,  value=u.localizacao)
        ws.cell(row=row, column=9,  value=float(u.area_privativa))
        ws.cell(row=row, column=10, value=float(u.area_privativa_acessoria))
        ws.cell(row=row, column=11, value=float(u.area_comum))
        ws.cell(row=row, column=12, value=float(u.fracao_ideal))
        ws.cell(row=row, column=13, value=float(u.valor_tabela))
        ws.cell(row=row, column=14, value=u.get_status_display())
        ws.cell(row=row, column=15, value=u.cliente_nome)
        ws.cell(row=row, column=16, value=u.descricao1)
        ws.cell(row=row, column=17, value=u.descricao2)
        ws.cell(row=row, column=18, value=u.descricao3)

    for col in ws.columns:
        width = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(width + 3, 40)

    nome_emp = empreendimento.nome.lower().replace(' ', '_') if empreendimento else 'todos'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="unidades_{nome_emp}_{date.today().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
def empreendimento_csv_template(request):
    import csv as csv_mod
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="template_unidades.csv"'
    response.write('﻿')  # BOM para Excel reconhecer UTF-8
    writer = csv_mod.writer(response, delimiter=';')
    writer.writerow([
        'ordem', 'bloco', 'numero', 'numeros_adicionais', 'tipo', 'tipologia', 'localizacao',
        'area_privativa', 'area_privativa_acessoria', 'area_comum',
        'fracao_ideal', 'valor_tabela', 'status',
        'descricao1', 'descricao2', 'descricao3',
    ])
    writer.writerow([
        '1', 'Torre A', '101', '', 'apartamento', '3D', '10º andar',
        '89.50', '0', '12.30', '0.001200', '650000.00', 'disponivel',
        '', '', '',
    ])
    writer.writerow([
        '2', 'Subsolo', 'G01', 'M03,HB60', 'garagem', 'coberta', 'Subsolo',
        '12.00', '0', '0', '0.000120', '45000.00', 'disponivel',
        '', '', '',
    ])
    return response


@login_required
def empreendimento_import_csv(request):
    import csv as csv_mod, io
    from decimal import Decimal, InvalidOperation

    if request.method != 'POST':
        return redirect('incorporadora:empreendimento_list')

    empreendimento_pk = request.POST.get('empreendimento')
    csv_file = request.FILES.get('arquivo')

    if not empreendimento_pk or not csv_file:
        messages.error(request, 'Selecione o empreendimento e o arquivo CSV.')
        return redirect('incorporadora:empreendimento_list')

    empreendimento = get_object_or_404(Empreendimento, pk=empreendimento_pk)

    tipo_map = {v.lower(): k for k, v in Unidade.TIPO_CHOICES}
    tipo_map.update({k: k for k, v in Unidade.TIPO_CHOICES})
    status_map = {v.lower(): k for k, v in Unidade.STATUS_CHOICES}
    status_map.update({k: k for k, v in Unidade.STATUS_CHOICES})

    def to_decimal(val, field, linha, erros):
        val = str(val).strip().replace(',', '.')
        try:
            return Decimal(val) if val else Decimal('0')
        except InvalidOperation:
            erros.append(f'Linha {linha}: campo "{field}" inválido ({val!r}).')
            return None

    try:
        decoded = csv_file.read().decode('utf-8-sig')
    except UnicodeDecodeError:
        csv_file.seek(0)
        decoded = csv_file.read().decode('latin-1')

    reader = csv_mod.DictReader(io.StringIO(decoded), delimiter=';')
    rows = [{k.strip().lower(): (v or '').strip() for k, v in row.items()} for row in reader]

    erros = []
    criados = 0
    blocos_cache = {}

    for i, row in enumerate(rows, 2):
        if not any(row.values()):
            continue

        bloco_nome = row.get('bloco', '').strip()
        numero = row.get('numero', '').strip()

        if not bloco_nome:
            erros.append(f'Linha {i}: coluna "bloco" obrigatória.')
            continue
        if not numero:
            erros.append(f'Linha {i}: coluna "numero" obrigatória.')
            continue

        if bloco_nome not in blocos_cache:
            bloco, _ = Bloco.objects.get_or_create(empreendimento=empreendimento, nome=bloco_nome)
            blocos_cache[bloco_nome] = bloco

        tipo_raw = row.get('tipo', 'apartamento').lower()
        tipo = tipo_map.get(tipo_raw)
        if not tipo:
            erros.append(f'Linha {i}: tipo "{tipo_raw}" inválido. Use: {", ".join(k for k,_ in Unidade.TIPO_CHOICES)}.')
            continue

        status_raw = row.get('status', 'disponivel').lower()
        status = status_map.get(status_raw, 'disponivel')

        try:
            ordem = int(row.get('ordem', '0') or '0')
        except ValueError:
            ordem = 0

        area_priv  = to_decimal(row.get('area_privativa', '0'),           'area_privativa',           i, erros)
        area_acess = to_decimal(row.get('area_privativa_acessoria', '0'), 'area_privativa_acessoria', i, erros)
        area_comum = to_decimal(row.get('area_comum', '0'),               'area_comum',               i, erros)
        fracao     = to_decimal(row.get('fracao_ideal', '0'),             'fracao_ideal',             i, erros)
        valor      = to_decimal(row.get('valor_tabela', '0'),             'valor_tabela',             i, erros)

        if None in (area_priv, area_acess, area_comum, fracao, valor):
            continue

        Unidade.objects.update_or_create(
            bloco=blocos_cache[bloco_nome],
            numero=numero,
            defaults={
                'ordem':                    ordem,
                'numeros_adicionais':       row.get('numeros_adicionais', ''),
                'tipo':                     tipo,
                'tipologia':                row.get('tipologia', ''),
                'localizacao':              row.get('localizacao', ''),
                'area_privativa':           area_priv,
                'area_privativa_acessoria': area_acess,
                'area_comum':               area_comum,
                'fracao_ideal':             fracao,
                'valor_tabela':             valor,
                'status':                   status,
                'descricao1':               row.get('descricao1', ''),
                'descricao2':               row.get('descricao2', ''),
                'descricao3':               row.get('descricao3', ''),
            }
        )
        criados += 1

    if erros:
        for e in erros[:10]:
            messages.warning(request, e)
        if len(erros) > 10:
            messages.warning(request, f'... e mais {len(erros) - 10} erro(s) não exibidos.')

    if criados:
        messages.success(request, f'{criados} unidade(s) importada(s)/atualizadas em "{empreendimento}".')
    elif not erros:
        messages.warning(request, 'Nenhuma unidade encontrada no arquivo.')

    return redirect('incorporadora:empreendimento_list')


@login_required
def empreendimento_relatorio_pdf(request, pk):
    from decimal import Decimal
    from io import BytesIO
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_LEFT

    empreendimento = get_object_or_404(
        Empreendimento.objects.select_related('empresa'), pk=pk
    )
    blocos = empreendimento.blocos.order_by('nome')

    zero = Decimal('0')
    campos = ['area_privativa', 'area_privativa_acessoria', 'area_comum', 'fracao_ideal', 'valor_tabela']

    def calcular_totais(uns):
        return {c: sum((getattr(u, c) for u in uns), zero) for c in campos}

    blocos_data = []
    for bloco in blocos:
        uns = list(Unidade.objects.filter(bloco=bloco).order_by('ordem'))
        blocos_data.append({'bloco': bloco, 'unidades': uns, 'totais': calcular_totais(uns)})
    todas = [u for b in blocos_data for u in b['unidades']]

    # ── cores ──
    C_HDR   = colors.HexColor('#A7A3AB')
    C_ALT   = colors.HexColor('#f7f7f7')
    C_SUB   = colors.HexColor('#f0eff2')
    C_BORDA = colors.HexColor('#e0e0e0')
    C_WHITE = colors.white

    # ── estilos de parágrafo ──
    def ps(name, font='Helvetica', size=8.5, color=colors.black, align=TA_LEFT, **kw):
        return ParagraphStyle(name, fontName=font, fontSize=size, textColor=color, alignment=align, **kw)

    sN  = ps('n')
    sR  = ps('r',  align=TA_RIGHT)
    sB  = ps('b',  font='Helvetica-Bold')
    sBR = ps('br', font='Helvetica-Bold', align=TA_RIGHT)
    sH  = ps('h',  font='Helvetica-Bold', size=8, color=C_WHITE)
    sHR = ps('hr', font='Helvetica-Bold', size=8, color=C_WHITE, align=TA_RIGHT)
    sSL = ps('sl', font='Helvetica-Bold', size=8)
    sSR = ps('sr', font='Helvetica-Bold', size=8, align=TA_RIGHT)
    sTL = ps('tl', font='Helvetica-Bold', size=9,  color=C_WHITE)
    sTR = ps('tr', font='Helvetica-Bold', size=9,  color=C_WHITE, align=TA_RIGHT)

    # ── larguras das colunas (total = 267mm) ──
    # Nº | Tipo | Tipologia | Localização | ÁPriv | ÁAcess | ÁComum | Fração | Valor | Status
    CW = [17*mm, 20*mm, 35*mm, 38*mm, 23*mm, 29*mm, 23*mm, 24*mm, 30*mm, 28*mm]
    # 11+26+35+38+23+29+23+24+30+28 = 267mm

    def fmt_dec(v, places=2):
        return f'{float(v):.{places}f}'

    def hdr_row():
        return [
            Paragraph('Nº', sH),
            Paragraph('Tipo', sH),
            Paragraph('Tipologia', sH),
            Paragraph('Localização', sH),
            Paragraph('Área Priv. (m²)', sHR),
            Paragraph('Área Priv. Acess. (m²)', sHR),
            Paragraph('Área Comum (m²)', sHR),
            Paragraph('Fração Ideal', sHR),
            Paragraph('Valor Tabela (R$)', sHR),
            Paragraph('Status', sH),
        ]

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=20*mm, bottomMargin=20*mm)
    els = []

    # cabeçalho do documento
    els.append(Paragraph(empreendimento.empresa.razao_social.upper(),
                         ps('emp', font='Helvetica-Bold', size=10, color=C_HDR)))
    els.append(Paragraph(f'Relatório de Unidades — {empreendimento.nome}',
                         ps('tit', font='Helvetica-Bold', size=14, color=colors.HexColor('#1a1a2e'),
                            spaceBefore=4)))
    els.append(Paragraph(
        f'{empreendimento.empresa} — {empreendimento.get_status_display()} — Gerado em {date.today().strftime("%d/%m/%Y")}',
        ps('sub', size=8, color=colors.HexColor('#666666'), spaceBefore=5)))
    els.append(HRFlowable(width='100%', thickness=2, color=C_HDR, spaceAfter=6*mm, spaceBefore=2*mm))

    for item in blocos_data:
        bloco = item['bloco']
        uns   = item['unidades']
        tots  = item['totais']

        els.append(Paragraph(f'Bloco: {bloco.nome}',
                             ps('bn', font='Helvetica-Bold', size=9,
                                color=colors.HexColor('#1a1a2e'), spaceBefore=6*mm, spaceAfter=2*mm)))
        if not uns:
            continue

        data = [hdr_row()]
        for u in uns:
            data.append([
                Paragraph(u.numero_display, sB),
                Paragraph(u.get_tipo_display(), sN),
                Paragraph(u.tipologia or '', sN),
                Paragraph(u.localizacao or '', sN),
                Paragraph(fmt_dec(u.area_privativa), sR),
                Paragraph(fmt_dec(u.area_privativa_acessoria), sR),
                Paragraph(fmt_dec(u.area_comum), sR),
                Paragraph(fmt_dec(u.fracao_ideal, 6), sR),
                Paragraph(fmt_dec(u.valor_tabela), sR),
                Paragraph(u.get_status_display(), sN),
            ])
        # subtotal
        data.append([
            Paragraph(f'Subtotal — {bloco.nome}', sSL), '', '', '',
            Paragraph(fmt_dec(tots['area_privativa']), sSR),
            Paragraph(fmt_dec(tots['area_privativa_acessoria']), sSR),
            Paragraph(fmt_dec(tots['area_comum']), sSR),
            Paragraph(fmt_dec(tots['fracao_ideal'], 6), sSR),
            Paragraph(fmt_dec(tots['valor_tabela']), sSR),
            '',
        ])

        nr = len(data)
        ts = TableStyle([
            ('BACKGROUND',  (0, 0),  (-1, 0),   C_HDR),
            ('ROWBACKGROUNDS', (0, 1), (-1, nr-2), [C_WHITE, C_ALT]),
            ('BACKGROUND',  (0, -1), (-1, -1),  C_SUB),
            ('LINEBELOW',   (0, 1),  (-1, nr-2), 0.5, C_BORDA),
            ('LINEABOVE',   (0, -1), (-1, -1),   1,   C_HDR),
            ('SPAN',        (0, -1), (3, -1)),
            ('VALIGN',      (0, 0),  (-1, -1),  'MIDDLE'),
            ('LEFTPADDING',  (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING',   (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING',(0, 0), (-1, -1), 3),
        ])
        t = Table(data, colWidths=CW, repeatRows=1)
        t.setStyle(ts)
        els.append(t)
        els.append(Spacer(1, 4*mm))

    # total geral
    if todas:
        tg = calcular_totais(todas)
        tdata = [[
            Paragraph(f'TOTAL GERAL — {len(todas)} unidade(s)', sTL), '', '', '',
            Paragraph(fmt_dec(tg['area_privativa']), sTR),
            Paragraph(fmt_dec(tg['area_privativa_acessoria']), sTR),
            Paragraph(fmt_dec(tg['area_comum']), sTR),
            Paragraph(fmt_dec(tg['fracao_ideal'], 6), sTR),
            Paragraph(fmt_dec(tg['valor_tabela']), sTR),
            '',
        ]]
        tt = Table(tdata, colWidths=CW)
        tt.setStyle(TableStyle([
            ('BACKGROUND',   (0, 0), (-1, 0), C_HDR),
            ('SPAN',         (0, 0), (3, 0)),
            ('LINEABOVE',    (0, 0), (-1, 0), 2, C_HDR),
            ('VALIGN',       (0, 0), (-1, 0), 'MIDDLE'),
            ('LEFTPADDING',  (0, 0), (-1, 0), 4),
            ('RIGHTPADDING', (0, 0), (-1, 0), 4),
            ('TOPPADDING',   (0, 0), (-1, 0), 5),
            ('BOTTOMPADDING',(0, 0), (-1, 0), 5),
        ]))
        els.append(tt)

    doc.build(els)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="relatorio_{pk}.pdf"'
    return response


@login_required
def empreendimento_vinculos_pdf(request, pk):
    from io import BytesIO
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT

    empreendimento = get_object_or_404(
        Empreendimento.objects.select_related('empresa'), pk=pk
    )
    blocos = empreendimento.blocos.prefetch_related('unidades__vinculadas').order_by('nome')

    C_HDR   = colors.HexColor('#A7A3AB')
    C_ALT   = colors.HexColor('#f7f7f7')
    C_BORDA = colors.HexColor('#e0e0e0')
    C_WHITE = colors.white

    def ps(name, font='Helvetica', size=8.5, color=colors.black, align=TA_LEFT, **kw):
        return ParagraphStyle(name, fontName=font, fontSize=size, textColor=color, alignment=align, **kw)

    sN  = ps('n')
    sB  = ps('b',  font='Helvetica-Bold')
    sH  = ps('h',  font='Helvetica-Bold', size=8, color=C_WHITE)

    CW = [30*mm, 55*mm, 55*mm, 40*mm]  # Unidade | Garagens | HBs | Outros = 180mm

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=20*mm, bottomMargin=20*mm)
    els = []

    els.append(Paragraph(empreendimento.empresa.razao_social.upper(),
                         ps('emp', font='Helvetica-Bold', size=10, color=C_HDR)))
    els.append(Paragraph(f'Vínculos — {empreendimento.nome}',
                         ps('tit', font='Helvetica-Bold', size=14, color=colors.HexColor('#1a1a2e'),
                            spaceBefore=4)))
    els.append(Paragraph(
        f'{empreendimento.empresa} — {empreendimento.get_status_display()} — Gerado em {date.today().strftime("%d/%m/%Y")}',
        ps('sub', size=8, color=colors.HexColor('#666666'), spaceBefore=5)))
    els.append(HRFlowable(width='100%', thickness=2, color=C_HDR, spaceAfter=6*mm, spaceBefore=2*mm))

    def _split_aliases(u):
        if not u.numeros_adicionais:
            return [], []
        parts = [a.strip() for a in u.numeros_adicionais.split(',') if a.strip()]
        hb     = [a for a in parts if a.upper().startswith('HB')]
        outros = [a for a in parts if not a.upper().startswith('HB')]
        return outros, hb

    def _display(numero, aliases):
        return f'{numero} ({" · ".join(aliases)})' if aliases else numero

    for bloco in blocos:
        principais = (
            bloco.unidades
            .filter(tipo__in=['apartamento', 'sala', 'loja'])
            .prefetch_related('vinculadas')
            .order_by('ordem', 'numero')
        )
        rows = []
        for u in principais:
            vinculadas = list(u.vinculadas.all())
            garagens, hbs, outros = [], [], []
            for v in vinculadas:
                outros_aliases, hb_aliases = _split_aliases(v)
                if v.tipo == 'garagem':
                    garagens.append(_display(v.numero, outros_aliases) + ('*' if hb_aliases else ''))
                    hbs.extend(hb_aliases)
                elif v.tipo == 'hobby_box':
                    hbs.append(_display(v.numero, outros_aliases))
                else:
                    outros.append(_display(v.numero, outros_aliases + hb_aliases))
            if vinculadas:
                rows.append((u, garagens, hbs, outros))

        if not rows:
            continue

        els.append(Paragraph(f'Bloco: {bloco.nome}',
                             ps('bn', font='Helvetica-Bold', size=9,
                                color=colors.HexColor('#1a1a2e'), spaceBefore=6*mm, spaceAfter=2*mm)))

        data = [[
            Paragraph('Unidade', sH),
            Paragraph('Garagens', sH),
            Paragraph('Hobby Boxes', sH),
            Paragraph('Outros', sH),
        ]]
        for i, (u, garagens, hbs, outros) in enumerate(rows):
            row_style = ps(f'r{i}', size=8.5)
            row_alt   = ps(f'ra{i}', size=8.5)
            data.append([
                Paragraph(u.numero_display, sB),
                Paragraph('  '.join(garagens) if garagens else '—', sN),
                Paragraph('  '.join(hbs)      if hbs      else '—', sN),
                Paragraph('  '.join(outros)   if outros   else '', sN),
            ])

        nr = len(data)
        ts = TableStyle([
            ('BACKGROUND',     (0, 0),  (-1, 0),   C_HDR),
            ('ROWBACKGROUNDS', (0, 1),  (-1, -1),  [C_WHITE, C_ALT]),
            ('LINEBELOW',      (0, 1),  (-1, -1),  0.5, C_BORDA),
            ('VALIGN',         (0, 0),  (-1, -1),  'MIDDLE'),
            ('LEFTPADDING',    (0, 0),  (-1, -1),  4),
            ('RIGHTPADDING',   (0, 0),  (-1, -1),  4),
            ('TOPPADDING',     (0, 0),  (-1, -1),  3),
            ('BOTTOMPADDING',  (0, 0),  (-1, -1),  3),
        ])
        t = Table(data, colWidths=CW, repeatRows=1)
        t.setStyle(ts)
        els.append(t)
        els.append(Spacer(1, 4*mm))

    doc.build(els)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="vinculos_{pk}.pdf"'
    return response


# ── Resumo por Empreendimento ─────────────────────────────────────────────────

def _build_empreendimento_resumo_context(empreendimento):
    from decimal import Decimal

    STATUS_LABELS = {
        'disponivel': 'Disponível',
        'reservado':  'Reservado',
        'vendido':    'Vendido',
        'permuta':    'Permuta',
        'bloqueado':  'Bloqueado',
        'qa':         'QA',
    }
    TIPO_LABELS = dict(Unidade.TIPO_CHOICES)

    unidades_principais = list(
        Unidade.objects
        .filter(bloco__empreendimento=empreendimento, unidade_principal__isnull=True)
        .select_related('bloco')
        .order_by('bloco__nome', 'ordem', 'numero')
    )

    # ── por status ────────────────────────────────────────────────────────────
    por_status = {v: {'qtde': 0, 'valor': Decimal('0'), 'area': Decimal('0'), 'unidades': []}
                  for v in STATUS_LABELS.values()}
    total_s = {'qtde': 0, 'valor': Decimal('0'), 'area': Decimal('0')}

    for u in unidades_principais:
        label = STATUS_LABELS.get(u.status, u.status.upper())
        if label not in por_status:
            por_status[label] = {'qtde': 0, 'valor': Decimal('0'), 'area': Decimal('0'), 'unidades': []}
        v = u.valor_tabela or Decimal('0')
        a = u.area_privativa or Decimal('0')
        perc = u.perc_permuta or Decimal('0')

        if perc > 0:
            # Permuta parcial: fraciona valor e área entre Permuta e o status real
            v_perm = v * perc
            a_perm = a * perc
            v_rest = v * (1 - perc)
            a_rest = a * (1 - perc)

            label_perm = 'Permuta'
            if label_perm not in por_status:
                por_status[label_perm] = {'qtde': 0, 'valor': Decimal('0'), 'area': Decimal('0'), 'unidades': []}
            por_status[label_perm]['qtde']  += 1
            por_status[label_perm]['valor'] += v_perm
            por_status[label_perm]['area']  += a_perm

            por_status[label]['qtde']   += 1
            por_status[label]['valor']  += v_rest
            por_status[label]['area']   += a_rest
            por_status[label]['unidades'].append(u)

            total_s['qtde']  += 1
            total_s['valor'] += v
            total_s['area']  += a
        else:
            por_status[label]['qtde']   += 1
            por_status[label]['valor']  += v
            por_status[label]['area']   += a
            por_status[label]['unidades'].append(u)
            total_s['qtde']  += 1
            total_s['valor'] += v
            total_s['area']  += a

    for dados in por_status.values():
        dados['qtde_perc']  = (Decimal(dados['qtde'])  / Decimal(total_s['qtde'])  * 100) if total_s['qtde']  else Decimal('0')
        dados['valor_perc'] = (dados['valor'] / total_s['valor'] * 100)                   if total_s['valor'] else Decimal('0')
        dados['area_perc']  = (dados['area']  / total_s['area']  * 100)                   if total_s['area']  else Decimal('0')

    total_s['qtde_perc'] = total_s['valor_perc'] = total_s['area_perc'] = Decimal('100') if total_s['qtde'] else Decimal('0')

    grupos_status = [
        {'status': label, 'dados': dados}
        for label, dados in por_status.items()
        if dados['qtde'] > 0
    ]

    # ── por tipo (só disponíveis, só apartamento e loja) ─────────────────────
    todas = list(
        Unidade.objects
        .filter(
            bloco__empreendimento=empreendimento,
            status='disponivel',
            tipo__in=['apartamento', 'loja'],
            unidade_principal__isnull=True,
        )
        .order_by('tipo')
    )
    por_tipo = {}
    for u in todas:
        label = TIPO_LABELS.get(u.tipo, u.tipo)
        if label not in por_tipo:
            por_tipo[label] = {'qtde': 0, 'valor': Decimal('0'), 'area': Decimal('0')}
        v = u.valor_tabela or Decimal('0')
        a = u.area_privativa or Decimal('0')
        perc = u.perc_permuta or Decimal('0')
        # Para loja com permuta parcial, usa apenas a fração disponível
        if perc > 0:
            v = v * (1 - perc)
            a = a * (1 - perc)
        por_tipo[label]['qtde']  += 1
        por_tipo[label]['valor'] += v
        por_tipo[label]['area']  += a

    for dados in por_tipo.values():
        dados['valor_m2'] = (dados['valor'] / dados['area']) if dados['area'] else Decimal('0')

    grupos_tipo = [{'tipo': t, 'dados': d} for t, d in por_tipo.items()]

    total_tipo = {
        'qtde':  sum(d['dados']['qtde']  for d in grupos_tipo),
        'valor': sum(d['dados']['valor'] for d in grupos_tipo),
        'area':  sum(d['dados']['area']  for d in grupos_tipo),
    }
    total_tipo['valor_m2'] = (total_tipo['valor'] / total_tipo['area']) if total_tipo['area'] else Decimal('0')

    return {
        'empreendimento': empreendimento,
        'grupos_status':  grupos_status,
        'total':          total_s,
        'grupos_tipo':    grupos_tipo,
        'total_tipo':     total_tipo,
        'data_geracao':   date.today(),
    }


@login_required
def empreendimento_resumo(request, pk):
    empreendimento = get_object_or_404(Empreendimento, pk=pk)
    ctx = _build_empreendimento_resumo_context(empreendimento)
    email_enviado = request.session.pop('resumo_email_enviado', False)
    ctx['email_enviado'] = email_enviado
    return render(request, 'incorporadora/empreendimento_resumo.html', ctx)


@login_required
def empreendimento_resumo_pdf(request, pk):
    from xhtml2pdf import pisa
    from io import BytesIO
    from django.template.loader import get_template

    empreendimento = get_object_or_404(Empreendimento, pk=pk)
    ctx = _build_empreendimento_resumo_context(empreendimento)
    tpl = get_template('incorporadora/pdf/empreendimento_resumo_pdf.html')
    html = tpl.render(ctx)
    buf = BytesIO()
    pisa.CreatePDF(html, dest=buf)
    buf.seek(0)
    nome = f'resumo_{empreendimento.nome.lower().replace(" ", "_")}.pdf'
    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nome}"'
    return response


@login_required
def empreendimento_resumo_email(request, pk):
    from django.core.mail import EmailMessage
    from xhtml2pdf import pisa
    from io import BytesIO
    from django.template.loader import get_template

    if request.method != 'POST':
        return redirect('incorporadora:empreendimento_resumo', pk=pk)

    empreendimento = get_object_or_404(Empreendimento, pk=pk)
    destinatario = request.user.email
    if not destinatario:
        messages.error(request, 'Seu usuário não tem e-mail cadastrado.')
        return redirect('incorporadora:empreendimento_resumo', pk=pk)

    ctx = _build_empreendimento_resumo_context(empreendimento)
    tpl = get_template('incorporadora/pdf/empreendimento_resumo_pdf.html')
    html = tpl.render(ctx)
    buf = BytesIO()
    pisa.CreatePDF(html, dest=buf)

    nome = f'resumo_{empreendimento.nome.lower().replace(" ", "_")}.pdf'
    email = EmailMessage(
        subject=f'Resumo — {empreendimento.nome}',
        body=f'Segue em anexo o resumo de situações de {empreendimento.nome}.',
        to=[destinatario],
    )
    email.attach(nome, buf.getvalue(), 'application/pdf')
    email.send(fail_silently=False)

    request.session['resumo_email_enviado'] = True
    messages.success(request, f'Resumo enviado para {destinatario}.')
    return redirect('incorporadora:empreendimento_resumo', pk=pk)


@login_required
def vinculo_list(request, bloco_pk):
    bloco = get_object_or_404(Bloco.objects.select_related('empreendimento'), pk=bloco_pk)

    principais = (
        bloco.unidades
        .filter(tipo__in=['apartamento', 'sala', 'loja'])
        .prefetch_related('vinculadas')
        .order_by('ordem', 'numero')
    )

    def _split_aliases(u):
        """Separa aliases em HB (hobby_box) e outros (garagem/moto/etc)."""
        if not u.numeros_adicionais:
            return [], []
        parts = [a.strip() for a in u.numeros_adicionais.split(',') if a.strip()]
        hb    = [a for a in parts if a.upper().startswith('HB')]
        outros = [a for a in parts if not a.upper().startswith('HB')]
        return outros, hb

    def _display(numero, aliases):
        if not aliases:
            return numero
        return f'{numero} ({" · ".join(aliases)})'

    rows = []
    for u in principais:
        vinculadas = list(u.vinculadas.all())
        garagens = []
        hbs      = []
        outros   = []

        for v in vinculadas:
            outros_aliases, hb_aliases = _split_aliases(v)
            if v.tipo == 'garagem':
                sufixo = '*' if hb_aliases else ''
                garagens.append(_display(v.numero, outros_aliases) + sufixo)
                hbs.extend(hb_aliases)
            elif v.tipo == 'hobby_box':
                hbs.append(_display(v.numero, outros_aliases))
            else:
                outros.append(_display(v.numero, outros_aliases + hb_aliases))

        if vinculadas:
            rows.append({'unidade': u, 'garagens': garagens, 'hbs': hbs, 'outros': outros})

    sem_vinculo = bloco.unidades.filter(
        tipo__in=['garagem', 'hobby_box'],
        unidade_principal__isnull=True,
    ).order_by('tipo', 'ordem', 'numero')

    return render(request, 'incorporadora/vinculo_list.html', {
        'bloco':       bloco,
        'rows':        rows,
        'sem_vinculo': sem_vinculo,
    })


@login_required
def vinculo_csv_template(request, bloco_pk):
    import csv as csv_mod
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="template_vinculos.csv"'
    response.write('﻿')
    writer = csv_mod.writer(response, delimiter=';')
    writer.writerow(['unidade', 'vinculada', 'numeros_adicionais'])
    writer.writerow(['201', 'G57', 'M03,HB60'])
    writer.writerow(['202', 'G18', ''])
    writer.writerow(['202', 'HB13', ''])
    return response


@login_required
def vinculo_import_csv(request, bloco_pk):
    import csv as csv_mod, io

    bloco = get_object_or_404(Bloco.objects.select_related('empreendimento'), pk=bloco_pk)

    if request.method != 'POST':
        return redirect('incorporadora:vinculo_list', bloco_pk=bloco_pk)

    csv_file = request.FILES.get('arquivo')
    if not csv_file:
        messages.error(request, 'Selecione um arquivo CSV.')
        return redirect('incorporadora:vinculo_list', bloco_pk=bloco_pk)

    try:
        decoded = csv_file.read().decode('utf-8-sig')
    except UnicodeDecodeError:
        csv_file.seek(0)
        decoded = csv_file.read().decode('latin-1')

    reader = csv_mod.DictReader(io.StringIO(decoded), delimiter=';')
    rows = [{k.strip().lower(): (v or '').strip() for k, v in row.items()} for row in reader]

    todas_unidades = list(bloco.unidades.all())
    unidades_bloco = {u.numero: u for u in todas_unidades}

    def buscar_unidade(numero):
        if numero in unidades_bloco:
            return unidades_bloco[numero]
        for u in todas_unidades:
            if u.numeros_adicionais:
                aliases = [a.strip() for a in u.numeros_adicionais.split(',') if a.strip()]
                if numero in aliases:
                    return u
        return None

    # limpa todos os vínculos e aliases do bloco antes de reimportar
    bloco.unidades.filter(
        tipo__in=['garagem', 'hobby_box']
    ).update(unidade_principal=None, numeros_adicionais='')

    erros = []
    vinculados = 0

    for i, row in enumerate(rows, 2):
        if not any(row.values()):
            continue

        num_principal = row.get('unidade', '').strip()
        num_vinculada = row.get('vinculada', '').strip()
        numeros_adicionais = row.get('numeros_adicionais', '').strip()

        if not num_principal:
            erros.append(f'Linha {i}: coluna "unidade" é obrigatória.')
            continue

        # sem vinculada mas com numeros_adicionais: apenas atualiza o campo na unidade
        if not num_vinculada:
            if numeros_adicionais:
                unidade = buscar_unidade(num_principal)
                if not unidade:
                    erros.append(f'Linha {i}: unidade "{num_principal}" não encontrada no bloco.')
                    continue
                unidade.numeros_adicionais = numeros_adicionais
                unidade.save(update_fields=['numeros_adicionais'])
                vinculados += 1
            continue

        principal = buscar_unidade(num_principal)
        if not principal:
            erros.append(f'Linha {i}: unidade "{num_principal}" não encontrada no bloco.')
            continue

        vinculada = buscar_unidade(num_vinculada)
        if not vinculada:
            erros.append(f'Linha {i}: unidade "{num_vinculada}" não encontrada no bloco.')
            continue

        if vinculada.pk == principal.pk:
            erros.append(f'Linha {i}: uma unidade não pode ser vinculada a si mesma.')
            continue

        update_fields = ['unidade_principal']
        vinculada.unidade_principal = principal
        if numeros_adicionais:
            vinculada.numeros_adicionais = numeros_adicionais
            update_fields.append('numeros_adicionais')
        vinculada.save(update_fields=update_fields)
        vinculados += 1

    if erros:
        for e in erros[:10]:
            messages.warning(request, e)
        if len(erros) > 10:
            messages.warning(request, f'... e mais {len(erros) - 10} erro(s) não exibidos.')

    if vinculados:
        messages.success(request, f'{vinculados} vínculo(s) importado(s) em "{bloco.nome}".')
    elif not erros:
        messages.warning(request, 'Nenhum vínculo encontrado no arquivo.')

    return redirect('incorporadora:vinculo_list', bloco_pk=bloco_pk)


@login_required
def unidade_export_excel(request, bloco_pk):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    bloco = get_object_or_404(Bloco.objects.select_related('empreendimento'), pk=bloco_pk)
    unidades = bloco.unidades.order_by('ordem')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Unidades'

    headers = [
        'Ordem', 'Número', 'Números Adicionais', 'Tipo', 'Tipologia', 'Localização',
        'Área Privativa (m²)', 'Área Priv. Acess. (m²)', 'Área Comum (m²)',
        'Fração Ideal', 'Valor Tabela (R$)', 'Status',
        'Descrição 1', 'Descrição 2', 'Descrição 3',
    ]
    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill(start_color='A7A3AB', end_color='A7A3AB', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')

    for row, u in enumerate(unidades, 2):
        ws.cell(row=row, column=1,  value=u.ordem)
        ws.cell(row=row, column=2,  value=u.numero)
        ws.cell(row=row, column=3,  value=u.numeros_adicionais)
        ws.cell(row=row, column=4,  value=u.get_tipo_display())
        ws.cell(row=row, column=5,  value=u.tipologia)
        ws.cell(row=row, column=6,  value=u.localizacao)
        ws.cell(row=row, column=7,  value=float(u.area_privativa))
        ws.cell(row=row, column=8,  value=float(u.area_privativa_acessoria))
        ws.cell(row=row, column=9,  value=float(u.area_comum))
        ws.cell(row=row, column=10, value=float(u.fracao_ideal))
        ws.cell(row=row, column=11, value=float(u.valor_tabela))
        ws.cell(row=row, column=12, value=u.get_status_display())
        ws.cell(row=row, column=13, value=u.descricao1)
        ws.cell(row=row, column=14, value=u.descricao2)
        ws.cell(row=row, column=15, value=u.descricao3)

    for col in ws.columns:
        width = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(width + 3, 40)

    nome_bloco = bloco.nome.replace(' ', '_')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="unidades_{nome_bloco}_{date.today().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
def unidade_csv_template(request, bloco_pk):
    import csv as csv_mod
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="template_unidades.csv"'
    response.write('﻿')
    writer = csv_mod.writer(response, delimiter=';')
    writer.writerow([
        'ordem', 'numero', 'numeros_adicionais', 'tipo', 'tipologia', 'localizacao',
        'area_privativa', 'area_privativa_acessoria', 'area_comum',
        'fracao_ideal', 'valor_tabela', 'status',
        'descricao1', 'descricao2', 'descricao3',
    ])
    writer.writerow([
        '1', '101', '', 'apartamento', '3D', '10º andar',
        '89.50', '0', '12.30', '0.001200', '650000.00', 'disponivel',
        '', '', '',
    ])
    writer.writerow([
        '2', 'G01', 'M03,HB60', 'garagem', 'coberta', 'Subsolo',
        '12.00', '0', '0', '0.000120', '45000.00', 'disponivel',
        '', '', '',
    ])
    return response


@login_required
def unidade_csv_template_empreendimento(request, empreendimento_pk):
    import csv as csv_mod
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="template_unidades_empreendimento.csv"'
    response.write('﻿')
    writer = csv_mod.writer(response, delimiter=';')
    writer.writerow([
        'ordem', 'bloco', 'numero', 'numeros_adicionais', 'tipo', 'tipologia', 'localizacao',
        'area_privativa', 'area_privativa_acessoria', 'area_comum',
        'fracao_ideal', 'valor_tabela', 'status',
        'descricao1', 'descricao2', 'descricao3',
    ])
    writer.writerow([
        '1', 'Único', '101', '', 'apartamento', '3D', '10º andar',
        '89.50', '0', '12.30', '0.001200', '650000.00', 'disponivel',
        '', '', '',
    ])
    writer.writerow([
        '2', 'Único', 'G01', 'M03,HB60', 'garagem', 'coberta', 'Subsolo',
        '12.00', '0', '0', '0.000120', '45000.00', 'disponivel',
        '', '', '',
    ])
    return response


@login_required
def unidade_import_empreendimento_csv(request, empreendimento_pk):
    import csv as csv_mod, io
    from decimal import Decimal, InvalidOperation

    empreendimento = get_object_or_404(Empreendimento.objects.prefetch_related('blocos'), pk=empreendimento_pk)

    if request.method != 'POST':
        return redirect('incorporadora:bloco_list', empreendimento_pk=empreendimento_pk)

    csv_file = request.FILES.get('arquivo')
    if not csv_file:
        messages.error(request, 'Selecione um arquivo CSV.')
        return redirect('incorporadora:bloco_list', empreendimento_pk=empreendimento_pk)

    tipo_map = {v.lower(): k for k, v in Unidade.TIPO_CHOICES}
    tipo_map.update({k: k for k, v in Unidade.TIPO_CHOICES})
    status_map = {v.lower(): k for k, v in Unidade.STATUS_CHOICES}
    status_map.update({k: k for k, v in Unidade.STATUS_CHOICES})

    def to_decimal(val, field, linha, erros):
        val = str(val).strip().replace(',', '.')
        try:
            return Decimal(val) if val else Decimal('0')
        except InvalidOperation:
            erros.append(f'Linha {linha}: campo "{field}" inválido ({val!r}).')
            return None

    try:
        decoded = csv_file.read().decode('utf-8-sig')
    except UnicodeDecodeError:
        csv_file.seek(0)
        decoded = csv_file.read().decode('latin-1')

    reader = csv_mod.DictReader(io.StringIO(decoded), delimiter=';')
    rows = [{k.strip().lower(): (v or '').strip() for k, v in row.items()} for row in reader]

    blocos_cache = {b.nome: b for b in empreendimento.blocos.all()}

    erros = []
    criados = 0

    for i, row in enumerate(rows, 2):
        if not any(row.values()):
            continue

        bloco_nome = row.get('bloco', '').strip()
        if not bloco_nome:
            erros.append(f'Linha {i}: coluna "bloco" obrigatória.')
            continue
        bloco = blocos_cache.get(bloco_nome)
        if not bloco:
            erros.append(f'Linha {i}: bloco "{bloco_nome}" não encontrado neste empreendimento.')
            continue

        numero = row.get('numero', '').strip()
        if not numero:
            erros.append(f'Linha {i}: coluna "numero" obrigatória.')
            continue

        tipo_raw = row.get('tipo', 'apartamento').lower()
        tipo = tipo_map.get(tipo_raw)
        if not tipo:
            erros.append(f'Linha {i}: tipo "{tipo_raw}" inválido. Use: {", ".join(k for k,_ in Unidade.TIPO_CHOICES)}.')
            continue

        status_raw = row.get('status', 'disponivel').lower()
        status = status_map.get(status_raw, 'disponivel')

        try:
            ordem = int(row.get('ordem', '0') or '0')
        except ValueError:
            ordem = 0

        area_priv  = to_decimal(row.get('area_privativa', '0'),           'area_privativa',           i, erros)
        area_acess = to_decimal(row.get('area_privativa_acessoria', '0'), 'area_privativa_acessoria', i, erros)
        area_com   = to_decimal(row.get('area_comum', '0'),               'area_comum',               i, erros)
        fracao     = to_decimal(row.get('fracao_ideal', '0'),             'fracao_ideal',             i, erros)
        valor      = to_decimal(row.get('valor_tabela', '0'),             'valor_tabela',             i, erros)

        if None in (area_priv, area_acess, area_com, fracao, valor):
            continue

        Unidade.objects.update_or_create(
            bloco=bloco,
            numero=numero,
            defaults={
                'ordem':                    ordem,
                'numeros_adicionais':       row.get('numeros_adicionais', ''),
                'tipo':                     tipo,
                'tipologia':                row.get('tipologia', ''),
                'localizacao':              row.get('localizacao', ''),
                'area_privativa':           area_priv,
                'area_privativa_acessoria': area_acess,
                'area_comum':               area_com,
                'fracao_ideal':             fracao,
                'valor_tabela':             valor,
                'status':                   status,
                'descricao1':               row.get('descricao1', ''),
                'descricao2':               row.get('descricao2', ''),
                'descricao3':               row.get('descricao3', ''),
            }
        )
        criados += 1

    if erros:
        for e in erros[:10]:
            messages.warning(request, e)
        if len(erros) > 10:
            messages.warning(request, f'... e mais {len(erros) - 10} erro(s) não exibidos.')

    if criados:
        messages.success(request, f'{criados} unidade(s) importada(s)/atualizadas em "{empreendimento.nome}".')
    elif not erros:
        messages.warning(request, 'Nenhuma unidade encontrada no arquivo.')

    return redirect('incorporadora:bloco_list', empreendimento_pk=empreendimento_pk)


@login_required
def vinculos_import_empreendimento_csv(request, empreendimento_pk):
    import csv as csv_mod, io

    empreendimento = get_object_or_404(Empreendimento, pk=empreendimento_pk)

    if request.method != 'POST':
        return redirect('incorporadora:bloco_list', empreendimento_pk=empreendimento_pk)

    csv_file = request.FILES.get('arquivo')
    if not csv_file:
        messages.error(request, 'Selecione um arquivo CSV.')
        return redirect('incorporadora:bloco_list', empreendimento_pk=empreendimento_pk)

    try:
        decoded = csv_file.read().decode('utf-8-sig')
    except UnicodeDecodeError:
        csv_file.seek(0)
        decoded = csv_file.read().decode('latin-1')

    reader = csv_mod.DictReader(io.StringIO(decoded), delimiter=';')
    rows = [{k.strip().lower(): (v or '').strip() for k, v in row.items()} for row in reader]

    todas_unidades = list(
        Unidade.objects.filter(bloco__empreendimento=empreendimento)
        .select_related('bloco')
    )
    por_numero = {}
    for u in todas_unidades:
        por_numero.setdefault(u.numero, []).append(u)
        if u.numeros_adicionais:
            for alias in [a.strip() for a in u.numeros_adicionais.split(',') if a.strip()]:
                por_numero.setdefault(alias, []).append(u)

    def buscar(numero):
        matches = por_numero.get(numero, [])
        if len(matches) == 1:
            return matches[0], None
        if len(matches) > 1:
            blocos = ', '.join(u.bloco.nome for u in matches)
            return None, f'número "{numero}" ambíguo (blocos: {blocos})'
        return None, f'unidade "{numero}" não encontrada'

    # Limpa vínculos anteriores de todo o empreendimento
    Unidade.objects.filter(
        bloco__empreendimento=empreendimento,
        tipo__in=['garagem', 'hobby_box'],
    ).update(unidade_principal=None)

    erros = []
    vinculados = 0

    for i, row in enumerate(rows, 2):
        if not any(row.values()):
            continue

        num_principal  = row.get('unidade', '').strip()
        num_vinculada  = row.get('vinculada', '').strip()
        nums_adicionais = row.get('numeros_adicionais', '').strip()

        if not num_principal:
            erros.append(f'Linha {i}: coluna "unidade" é obrigatória.')
            continue

        if not num_vinculada:
            if nums_adicionais:
                principal, err = buscar(num_principal)
                if err:
                    erros.append(f'Linha {i}: {err}.')
                    continue
                principal.numeros_adicionais = nums_adicionais
                principal.save(update_fields=['numeros_adicionais'])
                vinculados += 1
            continue

        principal, err = buscar(num_principal)
        if err:
            erros.append(f'Linha {i}: {err}.')
            continue

        vinculada, err = buscar(num_vinculada)
        if err:
            erros.append(f'Linha {i}: {err}.')
            continue

        if vinculada.pk == principal.pk:
            erros.append(f'Linha {i}: uma unidade não pode ser vinculada a si mesma.')
            continue

        update_fields = ['unidade_principal']
        vinculada.unidade_principal = principal
        if nums_adicionais:
            vinculada.numeros_adicionais = nums_adicionais
            update_fields.append('numeros_adicionais')
        vinculada.save(update_fields=update_fields)
        vinculados += 1

    if erros:
        for e in erros[:10]:
            messages.warning(request, e)
        if len(erros) > 10:
            messages.warning(request, f'... e mais {len(erros) - 10} erro(s) não exibidos.')

    if vinculados:
        messages.success(request, f'{vinculados} vínculo(s) importado(s) em "{empreendimento}".')
    elif not erros:
        messages.warning(request, 'Nenhum vínculo encontrado no arquivo.')

    return redirect('incorporadora:bloco_list', empreendimento_pk=empreendimento_pk)


@login_required
def unidade_import_tabela_cv(request, empreendimento_pk):
    import csv as csv_mod, io, re
    from decimal import Decimal, InvalidOperation

    empreendimento = get_object_or_404(Empreendimento, pk=empreendimento_pk)

    if request.method != 'POST':
        return redirect('incorporadora:bloco_list', empreendimento_pk=empreendimento_pk)

    csv_file = request.FILES.get('arquivo')
    if not csv_file:
        messages.error(request, 'Selecione um arquivo CSV.')
        return redirect('incorporadora:bloco_list', empreendimento_pk=empreendimento_pk)

    status_map = {
        'disponível': 'disponivel',
        'disponivel': 'disponivel',
        'reservada':  'reservado',
        'reservado':  'reservado',
        'vendida':    'vendido',
        'vendido':    'vendido',
    }

    def parse_valor(raw):
        raw = re.sub(r'[R$\s]', '', raw)
        raw = raw.replace('.', '').replace(',', '.')
        try:
            return Decimal(raw)
        except InvalidOperation:
            return None

    try:
        decoded = csv_file.read().decode('utf-8-sig')
    except UnicodeDecodeError:
        csv_file.seek(0)
        decoded = csv_file.read().decode('latin-1')

    reader = csv_mod.DictReader(io.StringIO(decoded), delimiter=';')
    rows = [{k.strip(): (v or '').strip() for k, v in row.items()} for row in reader]

    unidades_map = {
        u.numero: u
        for u in Unidade.objects.filter(bloco__empreendimento=empreendimento)
    }

    atualizadas = 0
    erros = []

    for i, row in enumerate(rows, 2):
        numero = (row.get('UNIDADE') or row.get('unidade') or '').strip()
        if not numero:
            continue

        unidade = unidades_map.get(str(numero))
        if not unidade:
            erros.append(f'Linha {i}: unidade "{numero}" não encontrada.')
            continue

        situacao_raw = (row.get('SITUAÇÃO') or row.get('SITUACAO') or row.get('situação') or row.get('situacao') or '').strip()
        valor_raw    = (row.get('VALOR TOTAL') or row.get('valor total') or '').strip()

        campos = {}
        if situacao_raw:
            status = status_map.get(situacao_raw.lower())
            if status:
                campos['status'] = status

        if valor_raw:
            valor = parse_valor(valor_raw)
            if valor is not None:
                campos['valor_tabela'] = valor
            else:
                erros.append(f'Linha {i}: valor "{valor_raw}" inválido.')

        if campos:
            for attr, val in campos.items():
                setattr(unidade, attr, val)
            unidade.save(update_fields=list(campos.keys()))
            atualizadas += 1

    if erros:
        for e in erros[:10]:
            messages.warning(request, e)
        if len(erros) > 10:
            messages.warning(request, f'... e mais {len(erros) - 10} aviso(s) não exibidos.')

    if atualizadas:
        messages.success(request, f'{atualizadas} unidade(s) atualizadas com dados da tabela CV.')
    elif not erros:
        messages.warning(request, 'Nenhuma unidade encontrada no arquivo.')

    return redirect('incorporadora:bloco_list', empreendimento_pk=empreendimento_pk)


@login_required
def unidade_import_csv(request, bloco_pk):
    import csv as csv_mod, io
    from decimal import Decimal, InvalidOperation

    bloco = get_object_or_404(Bloco.objects.select_related('empreendimento'), pk=bloco_pk)

    if request.method != 'POST':
        return redirect('incorporadora:unidade_list', bloco_pk=bloco_pk)

    csv_file = request.FILES.get('arquivo')
    if not csv_file:
        messages.error(request, 'Selecione um arquivo CSV.')
        return redirect('incorporadora:unidade_list', bloco_pk=bloco_pk)

    tipo_map = {v.lower(): k for k, v in Unidade.TIPO_CHOICES}
    tipo_map.update({k: k for k, v in Unidade.TIPO_CHOICES})
    status_map = {v.lower(): k for k, v in Unidade.STATUS_CHOICES}
    status_map.update({k: k for k, v in Unidade.STATUS_CHOICES})

    def to_decimal(val, field, linha, erros):
        val = str(val).strip().replace(',', '.')
        try:
            return Decimal(val) if val else Decimal('0')
        except InvalidOperation:
            erros.append(f'Linha {linha}: campo "{field}" inválido ({val!r}).')
            return None

    try:
        decoded = csv_file.read().decode('utf-8-sig')
    except UnicodeDecodeError:
        csv_file.seek(0)
        decoded = csv_file.read().decode('latin-1')

    reader = csv_mod.DictReader(io.StringIO(decoded), delimiter=';')
    rows = [{k.strip().lower(): (v or '').strip() for k, v in row.items()} for row in reader]

    erros = []
    criados = 0

    for i, row in enumerate(rows, 2):
        if not any(row.values()):
            continue

        numero = row.get('numero', '').strip()
        if not numero:
            erros.append(f'Linha {i}: coluna "numero" obrigatória.')
            continue

        tipo_raw = row.get('tipo', 'apartamento').lower()
        tipo = tipo_map.get(tipo_raw)
        if not tipo:
            erros.append(f'Linha {i}: tipo "{tipo_raw}" inválido. Use: {", ".join(k for k,_ in Unidade.TIPO_CHOICES)}.')
            continue

        status_raw = row.get('status', 'disponivel').lower()
        status = status_map.get(status_raw, 'disponivel')

        try:
            ordem = int(row.get('ordem', '0') or '0')
        except ValueError:
            ordem = 0

        area_priv  = to_decimal(row.get('area_privativa', '0'),           'area_privativa',           i, erros)
        area_acess = to_decimal(row.get('area_privativa_acessoria', '0'), 'area_privativa_acessoria', i, erros)
        area_com   = to_decimal(row.get('area_comum', '0'),               'area_comum',               i, erros)
        fracao     = to_decimal(row.get('fracao_ideal', '0'),             'fracao_ideal',             i, erros)
        valor      = to_decimal(row.get('valor_tabela', '0'),             'valor_tabela',             i, erros)

        if None in (area_priv, area_acess, area_com, fracao, valor):
            continue

        Unidade.objects.update_or_create(
            bloco=bloco,
            numero=numero,
            defaults={
                'ordem':                    ordem,
                'numeros_adicionais':       row.get('numeros_adicionais', ''),
                'tipo':                     tipo,
                'tipologia':                row.get('tipologia', ''),
                'localizacao':              row.get('localizacao', ''),
                'area_privativa':           area_priv,
                'area_privativa_acessoria': area_acess,
                'area_comum':               area_com,
                'fracao_ideal':             fracao,
                'valor_tabela':             valor,
                'status':                   status,
                'descricao1':               row.get('descricao1', ''),
                'descricao2':               row.get('descricao2', ''),
                'descricao3':               row.get('descricao3', ''),
            }
        )
        criados += 1

    if erros:
        for e in erros[:10]:
            messages.warning(request, e)
        if len(erros) > 10:
            messages.warning(request, f'... e mais {len(erros) - 10} erro(s) não exibidos.')

    if criados:
        messages.success(request, f'{criados} unidade(s) importada(s)/atualizadas em "{bloco.nome}".')
    elif not erros:
        messages.warning(request, 'Nenhuma unidade encontrada no arquivo.')

    return redirect('incorporadora:unidade_list', bloco_pk=bloco_pk)


@login_required
def empresa_list_pdf(request):
    empresas = Empresa.objects.all()
    primeira = empresas.filter(ativo=True).first()
    return render_to_pdf('incorporadora/pdf/empresa_list.html', {
        'empresas': empresas,
        'empresa_nome': primeira.razao_social if primeira else '',
        'data': date.today().strftime('%d/%m/%Y'),
    }, filename='empresas.pdf')


@login_required
def empreendimento_list_pdf(request):
    empreendimentos = Empreendimento.objects.select_related('empresa').all()
    primeira = Empresa.objects.filter(ativo=True).first()
    return render_to_pdf('incorporadora/pdf/empreendimento_list.html', {
        'empreendimentos': empreendimentos,
        'empresa_nome': primeira.razao_social if primeira else '',
        'data': date.today().strftime('%d/%m/%Y'),
    }, filename='empreendimentos.pdf')


@login_required
def bloco_list_pdf(request, empreendimento_pk):
    empreendimento = get_object_or_404(Empreendimento.objects.select_related('empresa'), pk=empreendimento_pk)
    blocos = list(empreendimento.blocos.all())
    return render_to_pdf('incorporadora/pdf/bloco_list.html', {
        'empreendimento': empreendimento,
        'blocos': blocos,
        'empresa_nome': empreendimento.empresa.razao_social,
        'data': date.today().strftime('%d/%m/%Y'),
    }, filename=f'blocos_{empreendimento.pk}.pdf')


@login_required
def unidade_list_pdf(request, bloco_pk):
    from io import BytesIO
    from decimal import Decimal
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_LEFT

    bloco = get_object_or_404(Bloco.objects.select_related('empreendimento__empresa'), pk=bloco_pk)
    unidades = list(bloco.unidades.order_by('ordem'))

    zero = Decimal('0')
    campos = ['area_privativa', 'area_privativa_acessoria', 'area_comum', 'fracao_ideal', 'valor_tabela']
    tots = {c: sum((getattr(u, c) for u in unidades), zero) for c in campos}

    C_HDR   = colors.HexColor('#A7A3AB')
    C_ALT   = colors.HexColor('#f7f7f7')
    C_SUB   = colors.HexColor('#f0eff2')
    C_BORDA = colors.HexColor('#e0e0e0')
    C_WHITE = colors.white

    def ps(name, font='Helvetica', size=8.5, color=colors.black, align=TA_LEFT, **kw):
        return ParagraphStyle(name, fontName=font, fontSize=size, textColor=color, alignment=align, **kw)

    sN  = ps('n')
    sR  = ps('r',  align=TA_RIGHT)
    sB  = ps('b',  font='Helvetica-Bold')
    sBR = ps('br', font='Helvetica-Bold', align=TA_RIGHT)
    sH  = ps('h',  font='Helvetica-Bold', size=8, color=C_WHITE)
    sHR = ps('hr', font='Helvetica-Bold', size=8, color=C_WHITE, align=TA_RIGHT)
    sTL = ps('tl', font='Helvetica-Bold', size=9, color=C_WHITE)
    sTR = ps('tr', font='Helvetica-Bold', size=9, color=C_WHITE, align=TA_RIGHT)

    CW = [17*mm, 20*mm, 35*mm, 38*mm, 23*mm, 29*mm, 23*mm, 24*mm, 30*mm, 28*mm]

    def fmt_dec(v, places=2):
        return f'{float(v):.{places}f}'

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=20*mm, bottomMargin=20*mm)
    els = []

    emp = bloco.empreendimento
    els.append(Paragraph(emp.empresa.razao_social.upper(),
                         ps('emp', font='Helvetica-Bold', size=10, color=C_HDR)))
    els.append(Paragraph(f'Relatório de Unidades — {emp.nome}',
                         ps('tit', font='Helvetica-Bold', size=14, color=colors.HexColor('#1a1a2e'),
                            spaceBefore=4)))
    els.append(Paragraph(
        f'Bloco: {bloco.nome} — {emp.get_status_display()} — Gerado em {date.today().strftime("%d/%m/%Y")}',
        ps('sub', size=8, color=colors.HexColor('#666666'), spaceBefore=5)))
    els.append(HRFlowable(width='100%', thickness=2, color=C_HDR, spaceAfter=6*mm, spaceBefore=2*mm))

    data = [[
        Paragraph('Nº', sH),
        Paragraph('Tipo', sH),
        Paragraph('Tipologia', sH),
        Paragraph('Localização', sH),
        Paragraph('Área Priv. (m²)', sHR),
        Paragraph('Área Priv. Acess. (m²)', sHR),
        Paragraph('Área Comum (m²)', sHR),
        Paragraph('Fração Ideal', sHR),
        Paragraph('Valor Tabela (R$)', sHR),
        Paragraph('Status', sH),
    ]]
    for u in unidades:
        data.append([
            Paragraph(u.numero_display, sB),
            Paragraph(u.get_tipo_display(), sN),
            Paragraph(u.tipologia or '', sN),
            Paragraph(u.localizacao or '', sN),
            Paragraph(fmt_dec(u.area_privativa), sR),
            Paragraph(fmt_dec(u.area_privativa_acessoria), sR),
            Paragraph(fmt_dec(u.area_comum), sR),
            Paragraph(fmt_dec(u.fracao_ideal, 6), sR),
            Paragraph(fmt_dec(u.valor_tabela), sR),
            Paragraph(u.get_status_display(), sN),
        ])

    if unidades:
        sSL = ps('sl', font='Helvetica-Bold', size=8)
        sSR = ps('sr', font='Helvetica-Bold', size=8, align=TA_RIGHT)
        data.append([
            Paragraph(f'Total — {len(unidades)} unidade(s)', sSL), '', '', '',
            Paragraph(fmt_dec(tots['area_privativa']), sSR),
            Paragraph(fmt_dec(tots['area_privativa_acessoria']), sSR),
            Paragraph(fmt_dec(tots['area_comum']), sSR),
            Paragraph(fmt_dec(tots['fracao_ideal'], 6), sSR),
            Paragraph(fmt_dec(tots['valor_tabela']), sSR),
            '',
        ])

    nr = len(data)
    ts = TableStyle([
        ('BACKGROUND',   (0, 0),  (-1, 0),   C_HDR),
        ('ROWBACKGROUNDS', (0, 1), (-1, nr-2 if unidades else -1), [C_WHITE, C_ALT]),
        ('BACKGROUND',   (0, -1), (-1, -1),  C_SUB),
        ('LINEBELOW',    (0, 1),  (-1, nr-2), 0.5, C_BORDA),
        ('LINEABOVE',    (0, -1), (-1, -1),   1,   C_HDR),
        ('SPAN',         (0, -1), (3, -1)),
        ('VALIGN',       (0, 0),  (-1, -1),  'MIDDLE'),
        ('LEFTPADDING',  (0, 0),  (-1, -1),  4),
        ('RIGHTPADDING', (0, 0),  (-1, -1),  4),
        ('TOPPADDING',   (0, 0),  (-1, -1),  3),
        ('BOTTOMPADDING',(0, 0),  (-1, -1),  3),
    ])
    t = Table(data, colWidths=CW, repeatRows=1)
    t.setStyle(ts)
    els.append(t)

    doc.build(els)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="unidades_{bloco_pk}.pdf"'
    return response


# ── Tabela de Vendas ──────────────────────────────────────────────────────────

@login_required
def tabela_list(request, empreendimento_pk):
    from django.http import JsonResponse as _JsonResponse
    empreendimento = get_object_or_404(Empreendimento, pk=empreendimento_pk)
    tabelas = TabelaVendas.objects.filter(empreendimento=empreendimento)
    if request.GET.get('json'):
        return _JsonResponse([{'pk': t.pk, 'nome': t.nome} for t in tabelas], safe=False)
    return render(request, 'incorporadora/tabela_list.html', {
        'empreendimento': empreendimento,
        'tabelas': tabelas,
    })


@login_required
def tabela_create(request, empreendimento_pk):
    empreendimento = get_object_or_404(Empreendimento, pk=empreendimento_pk)
    form = TabelaVendasForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        tabela = form.save(commit=False)
        tabela.empreendimento = empreendimento
        tabela.save()
        messages.success(request, 'Tabela criada com sucesso.')
        return redirect('incorporadora:tabela_detail', pk=tabela.pk)
    return render(request, 'incorporadora/tabela_form.html', {
        'form': form,
        'empreendimento': empreendimento,
        'titulo': 'Nova Tabela de Vendas',
    })


@login_required
def tabela_edit(request, pk):
    tabela = get_object_or_404(TabelaVendas.objects.select_related('empreendimento'), pk=pk)
    form = TabelaVendasForm(request.POST or None, instance=tabela)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Tabela atualizada com sucesso.')
        return redirect('incorporadora:tabela_detail', pk=tabela.pk)
    return render(request, 'incorporadora/tabela_form.html', {
        'form': form,
        'empreendimento': tabela.empreendimento,
        'titulo': 'Editar Tabela de Vendas',
        'obj': tabela,
    })


@login_required
def tabela_delete(request, pk):
    tabela = get_object_or_404(TabelaVendas.objects.select_related('empreendimento'), pk=pk)
    if request.method == 'POST':
        emp_pk = tabela.empreendimento.pk
        tabela.delete()
        messages.success(request, 'Tabela excluída.')
        return redirect('incorporadora:tabela_list', empreendimento_pk=emp_pk)
    return render(request, 'incorporadora/tabela_confirm_delete.html', {'obj': tabela})


@login_required
def tabela_detail(request, pk):
    tabela = get_object_or_404(
        TabelaVendas.objects.select_related('empreendimento__empresa').prefetch_related('series'),
        pk=pk,
    )
    itens = (ItemTabelaVendas.objects
             .filter(tabela=tabela)
             .select_related('unidade__bloco')
             .prefetch_related('valores__serie', 'unidade__vinculadas'))
    series = list(tabela.series.all())
    return render(request, 'incorporadora/tabela_detail.html', {
        'tabela': tabela,
        'series': series,
        'itens': itens,
    })


# ── Séries de Pagamento ───────────────────────────────────────────────────────

@login_required
def serie_create(request, tabela_pk):
    from decimal import Decimal
    tabela = get_object_or_404(TabelaVendas, pk=tabela_pk)
    form = SeriePagamentoForm(request.POST or None, tabela=tabela)
    if request.method == 'POST' and form.is_valid():
        serie = form.save(commit=False)
        serie.tabela = tabela
        serie.save()
        messages.success(request, 'Série adicionada.')
        return redirect('incorporadora:tabela_detail', pk=tabela.pk)
    soma = sum(s.percentual for s in tabela.series.all() if s.percentual) or Decimal('0')
    return render(request, 'incorporadora/serie_form.html', {
        'form': form, 'tabela': tabela, 'titulo': 'Nova Série de Pagamento',
        'pct_disponivel': f'{Decimal("100") - soma:.3f}',
    })


@login_required
def serie_edit(request, pk):
    from decimal import Decimal
    serie = get_object_or_404(SeriePagamento.objects.select_related('tabela'), pk=pk)
    form = SeriePagamentoForm(request.POST or None, instance=serie, tabela=serie.tabela)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Série atualizada.')
        return redirect('incorporadora:tabela_detail', pk=serie.tabela.pk)
    soma = sum(s.percentual for s in serie.tabela.series.exclude(pk=serie.pk) if s.percentual) or Decimal('0')
    return render(request, 'incorporadora/serie_form.html', {
        'form': form, 'tabela': serie.tabela, 'titulo': 'Editar Série', 'obj': serie,
        'pct_disponivel': f'{Decimal("100") - soma:.3f}',
    })


@login_required
def serie_delete(request, pk):
    serie = get_object_or_404(SeriePagamento.objects.select_related('tabela'), pk=pk)
    if request.method == 'POST':
        tabela_pk = serie.tabela.pk
        serie.delete()
        messages.success(request, 'Série excluída.')
        return redirect('incorporadora:tabela_detail', pk=tabela_pk)
    return render(request, 'incorporadora/serie_confirm_delete.html', {'obj': serie})


# ── Itens da Tabela (CSV) ─────────────────────────────────────────────────────

@login_required
def tabela_item_csv_template(request, pk):
    tabela = get_object_or_404(TabelaVendas.objects.prefetch_related('series'), pk=pk)
    series = list(tabela.series.all())
    import csv
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="tabela_{pk}_template.csv"'
    writer = csv.writer(response, delimiter=';')
    header = ['bloco', 'unidade', 'status', 'valor_venda'] + [s.tipo for s in series]
    writer.writerow(header)
    writer.writerow(['Torre A', '101', 'disponivel', '350000.00'] + ['0.00'] * len(series))
    return response


@login_required
def tabela_item_import_csv(request, pk):
    import csv, io, re
    tabela = get_object_or_404(TabelaVendas.objects.select_related('empreendimento').prefetch_related('series'), pk=pk)
    series = {s.tipo: s for s in tabela.series.all()}

    if request.method != 'POST':
        return redirect('incorporadora:tabela_detail', pk=pk)

    arquivo = request.FILES.get('arquivo')
    if not arquivo:
        messages.error(request, 'Nenhum arquivo enviado.')
        return redirect('incorporadora:tabela_detail', pk=pk)

    # Mapeamento de status do CV CRM para os valores internos
    STATUS_CV_MAP = {
        'disponível': 'disponivel',
        'disponivel': 'disponivel',
        'bloqueada':  'bloqueado',
        'bloqueado':  'bloqueado',
        'reservada':  'reservado',
        'reservado':  'reservado',
        'vendida':    'vendido',
        'vendido':    'vendido',
    }

    def _parse_valor_cv(raw):
        """Converte 'R$132.123,00' ou '132.123,00' para '132123.00'."""
        s = re.sub(r'[R$\s]', '', raw).strip()
        s = s.replace('.', '').replace(',', '.')
        return s or '0'

    try:
        texto = arquivo.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(texto), delimiter=';')
        # Normalize header keys stripping whitespace
        fieldnames = [f.strip() for f in (reader.fieldnames or [])]
        criados = atualizados = erros = 0

        # Detect CV CRM format by presence of 'SITUAÇÃO' column
        is_cv_format = 'SITUAÇÃO' in fieldnames

        for i, row in enumerate(reader, start=2):
            row = {k.strip(): v for k, v in row.items()}

            if is_cv_format:
                unidade_num = row.get('UNIDADE', '').strip()
                if not unidade_num:
                    continue

                situacao_raw = row.get('SITUAÇÃO', '').strip()
                status_interno = STATUS_CV_MAP.get(situacao_raw.lower(), 'disponivel')
                valor_raw = row.get('VALOR TOTAL', '0').strip()
                valor_venda = _parse_valor_cv(valor_raw)

                try:
                    unidade = Unidade.objects.get(
                        bloco__empreendimento=tabela.empreendimento,
                        numero=unidade_num,
                    )
                except Unidade.DoesNotExist:
                    messages.warning(request, f'Linha {i}: unidade {unidade_num} não encontrada.')
                    erros += 1
                    continue
                except Unidade.MultipleObjectsReturned:
                    messages.warning(request, f'Linha {i}: número {unidade_num} ambíguo (múltiplos blocos). Ignorado.')
                    erros += 1
                    continue

                # Update the Unidade itself
                Unidade.objects.filter(pk=unidade.pk).update(
                    status=status_interno,
                    valor_tabela=valor_venda,
                )

                item, criado = ItemTabelaVendas.objects.update_or_create(
                    tabela=tabela, unidade=unidade,
                    defaults={'status': status_interno, 'valor_venda': valor_venda},
                )
                if criado:
                    criados += 1
                else:
                    atualizados += 1

            else:
                # Legacy format: bloco;unidade;status;valor_venda[;series...]
                bloco_nome  = row.get('bloco', '').strip()
                unidade_num = row.get('unidade', '').strip()
                status      = row.get('status', 'disponivel').strip()
                valor_venda = row.get('valor_venda', '0').strip().replace(',', '.')

                if not bloco_nome or not unidade_num:
                    continue

                try:
                    unidade = Unidade.objects.get(
                        bloco__empreendimento=tabela.empreendimento,
                        bloco__nome__iexact=bloco_nome,
                        numero=unidade_num,
                    )
                except Unidade.DoesNotExist:
                    messages.warning(request, f'Linha {i}: unidade {bloco_nome}/{unidade_num} não encontrada.')
                    erros += 1
                    continue

                item, criado = ItemTabelaVendas.objects.update_or_create(
                    tabela=tabela, unidade=unidade,
                    defaults={'status': status, 'valor_venda': valor_venda},
                )
                if criado:
                    criados += 1
                else:
                    atualizados += 1

                # valores por série
                for col_name, valor_str in row.items():
                    col_key = col_name.strip()
                    if col_key in series:
                        ValorSerie.objects.update_or_create(
                            item=item, serie=series[col_key],
                            defaults={'valor_parcela': valor_str.strip().replace(',', '.') or '0'},
                        )

        fmt = 'CV CRM' if is_cv_format else 'padrão'
        messages.success(request, f'Importação ({fmt}) concluída: {criados} criados, {atualizados} atualizados, {erros} erro(s).')
    except Exception as e:
        messages.error(request, f'Erro ao processar arquivo: {e}')

    return redirect('incorporadora:tabela_detail', pk=pk)


@login_required
def tabela_gerar_itens(request, pk):
    tabela = get_object_or_404(TabelaVendas.objects.select_related('empreendimento'), pk=pk)
    if request.method != 'POST':
        return redirect('incorporadora:tabela_detail', pk=pk)

    unidades = Unidade.objects.filter(
        bloco__empreendimento=tabela.empreendimento,
        tipo__in=['apartamento', 'sala', 'loja'],
    )
    series = list(tabela.series.all())
    criados = atualizados = 0
    for u in unidades:
        item, criado = ItemTabelaVendas.objects.get_or_create(
            tabela=tabela,
            unidade=u,
            defaults={'status': 'disponivel', 'valor_venda': u.valor_tabela},
        )
        if criado:
            criados += 1
        else:
            if item.valor_venda != u.valor_tabela:
                item.valor_venda = u.valor_tabela
                item.save(update_fields=['valor_venda'])
            atualizados += 1

        for serie in series:
            if serie.percentual and item.valor_venda:
                from decimal import Decimal
                qtd = serie.quantidade or 1
                valor_parcela = (item.valor_venda * serie.percentual / Decimal('100') / qtd).quantize(Decimal('0.01'))
                ValorSerie.objects.update_or_create(
                    item=item, serie=serie,
                    defaults={'valor_parcela': valor_parcela},
                )
            else:
                ValorSerie.objects.get_or_create(item=item, serie=serie,
                                                 defaults={'valor_parcela': 0})

    messages.success(request, f'{criados} item(ns) criado(s), {atualizados} já existia(m).')
    return redirect('incorporadora:tabela_detail', pk=pk)


@login_required
def tabela_pdf(request, pk):
    from decimal import Decimal
    from io import BytesIO
    from itertools import groupby
    from collections import defaultdict
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, PageBreak
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_LEFT, TA_CENTER

    tabela = get_object_or_404(
        TabelaVendas.objects.select_related('empreendimento__empresa').prefetch_related('series'),
        pk=pk,
    )
    series = list(tabela.series.order_by('ordem'))

    apenas_disponiveis = request.GET.get('disponivel') == '1'
    itens_qs = (ItemTabelaVendas.objects
                .filter(tabela=tabela)
                .select_related('unidade__bloco')
                .prefetch_related('valores__serie', 'unidade__vinculadas')
                .order_by('unidade__bloco__nome', 'unidade__pagina', 'unidade__ordem', 'unidade__numero'))
    if apenas_disponiveis:
        itens_qs = itens_qs.filter(unidade__status='disponivel')

    cub = tabela.cub_referencia or Decimal('1')

    processed = []
    for item in itens_qs:
        u = item.unidade
        vinculadas = list(u.vinculadas.all())
        garagens   = [v for v in vinculadas if v.tipo == 'garagem']
        hbs        = [v for v in vinculadas if v.tipo == 'hobby_box']

        vagas_nums = ', '.join(v.numero_display for v in garagens + hbs) or '—'
        vagas_tipo = garagens[0].tipologia if garagens else (hbs[0].tipologia if hbs else '')

        area_vinc  = sum(v.area_privativa for v in vinculadas)
        area_total = u.area_privativa + u.area_privativa_acessoria + u.area_comum + area_vinc

        cub_priv  = (item.valor_venda / (cub * u.area_privativa)).quantize(Decimal('0.000001'))  if u.area_privativa else Decimal('0')
        cub_total = (item.valor_venda / (cub * area_total)).quantize(Decimal('0.000001'))         if area_total       else Decimal('0')

        processed.append({
            'item':       item,
            'unidade':    u,
            'bloco':      u.bloco.nome,
            'pagina':     u.pagina,
            'vagas_nums': vagas_nums,
            'vagas_tipo': vagas_tipo,
            'area_total': area_total,
            'cub_priv':   cub_priv,
            'cub_total':  cub_total,
            'valores':    {v.serie_id: v.valor_parcela for v in item.valores.all()},
        })

    STATUS_LABELS = dict(ItemTabelaVendas.STATUS_CHOICES)
    resumo = defaultdict(lambda: {'label': '', 'count': 0,
                                  'area_priv': Decimal('0'), 'area_total': Decimal('0'), 'valor': Decimal('0')})
    for p in processed:
        s = p['item'].status
        resumo[s]['label']       = STATUS_LABELS.get(s, s)
        resumo[s]['count']      += 1
        resumo[s]['area_priv']  += p['unidade'].area_privativa
        resumo[s]['area_total'] += p['area_total']
        resumo[s]['valor']      += p['item'].valor_venda

    # ── cores ──
    C_HDR   = colors.HexColor('#A7A3AB')
    C_ALT   = colors.HexColor('#f7f7f7')
    C_SUB   = colors.HexColor('#f0eff2')
    C_BORDA = colors.HexColor('#e0e0e0')
    C_DARK  = colors.HexColor('#1a1a2e')
    C_MUTED = colors.HexColor('#888888')

    def ps(name, font='Helvetica', size=7.5, color=colors.black, align=TA_LEFT, **kw):
        return ParagraphStyle(name, fontName=font, fontSize=size, textColor=color,
                              alignment=align, leading=kw.pop('leading', size * 1.15), **kw)

    sN   = ps('n')
    sC   = ps('c',  align=TA_CENTER)
    sR   = ps('r',  align=TA_RIGHT)
    sB   = ps('b',  font='Helvetica-Bold')
    sBR  = ps('br', font='Helvetica-Bold', align=TA_RIGHT)
    sBC  = ps('bc', font='Helvetica-Bold', align=TA_CENTER)
    sMR  = ps('mr', color=C_MUTED, align=TA_RIGHT)
    sMC  = ps('mc', color=C_MUTED, align=TA_CENTER)
    sST  = ps('st', font='Helvetica-Bold', size=6, color=C_MUTED, align=TA_CENTER)
    sH   = ps('h',  font='Helvetica-Bold', size=7, color=colors.white, align=TA_CENTER)
    sHR  = ps('hr', font='Helvetica-Bold', size=7, color=colors.white, align=TA_RIGHT)

    # ── larguras (A4 landscape 297mm − 2×15mm margem = 267mm) ──
    TOTAL_W  = 267
    FIXED_MM = [14, 12, 22, 30, 16, 16, 16, 26, 14, 14]   # 10 colunas fixas = 180mm
    n_series = len(series)
    if n_series:
        serie_w = max(16.0, min(35.0, (TOTAL_W - sum(FIXED_MM)) / n_series))
        avail   = TOTAL_W - serie_w * n_series
        ratio   = avail / sum(FIXED_MM)
    else:
        serie_w = 0
        ratio   = TOTAL_W / sum(FIXED_MM)
    CW = [f * ratio * mm for f in FIXED_MM[:8]] + [serie_w * mm] * n_series + [f * ratio * mm for f in FIXED_MM[8:]]

    def fmt(v, places=2):
        try:
            s = f'{float(v):,.{places}f}'
            return s.replace(',', 'X').replace('.', ',').replace('X', '.')
        except Exception:
            return str(v) if v else ''

    def hdr():
        row = [
            Paragraph('STATUS', sH), Paragraph('APTO', sH), Paragraph('TIPOLOGIA', sH),
            Paragraph('VAGAS', sH), Paragraph('TIPO', sH),
            Paragraph('Á. PRIV.<br/>(m²)', sHR), Paragraph('Á. TOTAL<br/>(m²)', sHR),
            Paragraph('VALOR<br/>VENDA', sHR),
        ]
        for s in series:
            row.append(Paragraph(s.label, sHR))
        row += [Paragraph('CUB<br/>Á.PRIV.', sHR), Paragraph('CUB<br/>Á.TOTAL', sHR)]
        return row

    def tbl_style(n_rows):
        st = TableStyle([
            ('BACKGROUND',   (0, 0), (-1, 0), C_HDR),
            ('VALIGN',       (0, 0), (-1,-1), 'MIDDLE'),
            ('GRID',         (0, 0), (-1, 0), 0.5, colors.HexColor('#888888')),
            ('LINEBELOW',    (0, 1), (-1,-1), 0.4, C_BORDA),
            ('TOPPADDING',   (0, 0), (-1,-1), 3),
            ('BOTTOMPADDING',(0, 0), (-1,-1), 3),
            ('LEFTPADDING',  (0, 0), (-1,-1), 3),
            ('RIGHTPADDING', (0, 0), (-1,-1), 3),
        ])
        for i in range(2, n_rows + 1, 2):
            st.add('BACKGROUND', (0, i), (-1, i), C_ALT)
        return st

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    els = []
    hoje = date.today().strftime('%d/%m/%Y')

    vigencia = ''
    if tabela.data_inicio:
        vigencia = f' | Vigência: {tabela.data_inicio.strftime("%d/%m/%Y")}'
        if tabela.data_fim:
            vigencia += f' a {tabela.data_fim.strftime("%d/%m/%Y")}'

    first_page = True
    for (bloco_nome, pagina), grupo_iter in groupby(processed, key=lambda x: (x['bloco'], x['pagina'])):
        if not first_page:
            els.append(PageBreak())
        first_page = False
        grupo_list = list(grupo_iter)

        els.append(Paragraph(str(tabela.empreendimento.empresa).upper(),
                             ps('emp', font='Helvetica-Bold', size=8, color=C_HDR)))
        els.append(Paragraph(
            f'TABELA DE VENDAS ({tabela.get_modalidade_display().upper()}) — {bloco_nome.upper()}',
            ps('tit', font='Helvetica-Bold', size=13, color=C_DARK, spaceBefore=3)))
        data_ref = tabela.data_referencia.strftime('%m/%Y') if tabela.data_referencia else ''
        els.append(Paragraph(
            f'{tabela.nome} | Data ref.: {data_ref} | CUB: R$ {tabela.cub_referencia}{vigencia} | Gerado em {hoje}',
            ps('sub', size=7.5, color=C_MUTED, spaceBefore=2)))
        els.append(HRFlowable(width='100%', thickness=2, color=C_HDR, spaceAfter=3*mm, spaceBefore=2*mm))

        data = [hdr()]
        for p in grupo_list:
            u = p['unidade']
            row = [
                Paragraph(u.get_status_display() if u.status != 'disponivel' else '', sST),
                Paragraph(u.numero, sBC),
                Paragraph(u.tipologia or '—', ps('tp', color=C_MUTED, size=7.5)),
                Paragraph(p['vagas_nums'], sMC),
                Paragraph(p['vagas_tipo'] or '—', sMC),
                Paragraph(fmt(u.area_privativa), sR),
                Paragraph(fmt(p['area_total']), sMR),
                Paragraph(fmt(p['item'].valor_venda), sBR),
            ]
            for s in series:
                v = p['valores'].get(s.pk)
                row.append(Paragraph(fmt(v) if v is not None else '', sMR))
            row += [Paragraph(fmt(p['cub_priv'], 4), sMR), Paragraph(fmt(p['cub_total'], 4), sMR)]
            data.append(row)

        n_data = len(data)
        ftr_row = [Paragraph('TOTAL', ps('ft', font='Helvetica-Bold', size=7.5))] + [''] * 4
        ftr_row.append(Paragraph(f'{len(grupo_list)} unid.',
                                  ps('ftr', font='Helvetica-Bold', size=7.5, align=TA_RIGHT)))
        ftr_row += ['', ''] + [''] * n_series + ['', '']
        data.append(ftr_row)

        st = tbl_style(n_data - 1)
        st.add('BACKGROUND', (0, n_data), (-1, n_data), C_SUB)
        st.add('LINEABOVE',  (0, n_data), (-1, n_data), 1, C_HDR)
        t = Table(data, colWidths=CW, repeatRows=1)
        t.setStyle(st)
        els.append(t)
        data_ref_long = tabela.data_referencia.strftime('%B/%Y') if tabela.data_referencia else ''
        els.append(Paragraph(
            f'Mês de referência: {data_ref_long} — A presente tabela poderá ser alterada sem aviso prévio.',
            ps('fn', size=6.5, color=C_MUTED, align=TA_RIGHT, spaceBefore=3)))

    # ── RESUMO ──
    els.append(PageBreak())
    els.append(Paragraph(str(tabela.empreendimento.empresa).upper(),
                         ps('emp2', font='Helvetica-Bold', size=8, color=C_HDR)))
    els.append(Paragraph(
        f'TABELA DE VENDAS ({tabela.get_modalidade_display().upper()}) — RESUMO',
        ps('tit2', font='Helvetica-Bold', size=13, color=C_DARK, spaceBefore=3)))
    els.append(Paragraph(f'{tabela.nome} | Gerado em {hoje}',
                         ps('sub2', size=7.5, color=C_MUTED, spaceBefore=2)))
    els.append(HRFlowable(width='100%', thickness=2, color=C_HDR, spaceAfter=4*mm, spaceBefore=2*mm))

    sRH  = ps('rh',  font='Helvetica-Bold', size=8, color=colors.white)
    sRHR = ps('rhr', font='Helvetica-Bold', size=8, color=colors.white, align=TA_RIGHT)
    sRN  = ps('rn',  size=8)
    sRR  = ps('rr',  size=8, align=TA_RIGHT)
    sRB  = ps('rb',  font='Helvetica-Bold', size=8)
    sRBR = ps('rbr', font='Helvetica-Bold', size=8, align=TA_RIGHT)

    tot = {'count': 0, 'area_priv': Decimal('0'), 'area_total': Decimal('0'), 'valor': Decimal('0')}
    rdata = [[Paragraph('Situação', sRH), Paragraph('Unidades', sRHR),
              Paragraph('Área Privativa (m²)', sRHR), Paragraph('Área Total (m²)', sRHR),
              Paragraph('Valor Total (R$)', sRHR)]]
    for dados in resumo.values():
        rdata.append([Paragraph(dados['label'], sRN), Paragraph(str(dados['count']), sRR),
                      Paragraph(fmt(dados['area_priv']), sRR), Paragraph(fmt(dados['area_total']), sRR),
                      Paragraph(fmt(dados['valor']), sRR)])
        tot['count']     += dados['count']
        tot['area_priv'] += dados['area_priv']
        tot['area_total']+= dados['area_total']
        tot['valor']     += dados['valor']
    rdata.append([Paragraph('TOTAL', sRB), Paragraph(str(tot['count']), sRBR),
                  Paragraph(fmt(tot['area_priv']), sRBR), Paragraph(fmt(tot['area_total']), sRBR),
                  Paragraph(fmt(tot['valor']), sRBR)])

    RCW = [80*mm, 30*mm, 52*mm, 52*mm, 53*mm]
    rt = Table(rdata, colWidths=RCW)
    rt.setStyle(TableStyle([
        ('BACKGROUND',   (0, 0), (-1, 0), C_HDR),
        ('VALIGN',       (0, 0), (-1,-1), 'MIDDLE'),
        ('GRID',         (0, 0), (-1, 0), 0.5, colors.HexColor('#888888')),
        ('LINEBELOW',    (0, 1), (-1,-2), 0.4, C_BORDA),
        ('BACKGROUND',   (0, len(rdata)-1), (-1,-1), C_SUB),
        ('LINEABOVE',    (0, len(rdata)-1), (-1,-1), 1, C_HDR),
        ('TOPPADDING',   (0, 0), (-1,-1), 4),
        ('BOTTOMPADDING',(0, 0), (-1,-1), 4),
        ('LEFTPADDING',  (0, 0), (-1,-1), 5),
        ('RIGHTPADDING', (0, 0), (-1,-1), 5),
    ]))
    els.append(rt)

    doc.build(els)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="tabela_{pk}.pdf"'
    return response


# ── Página de Importação por Empreendimento ───────────────────────────────────

def _decode_csv(file_obj):
    """Decodifica arquivo CSV respeitando BOM e encoding."""
    raw = file_obj.read()
    for enc in ('utf-8-sig', 'latin-1'):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode('utf-8', errors='replace')


def _csv_reader(texto, delimiter=None):
    import csv, io
    if delimiter is None:
        try:
            dialect = csv.Sniffer().sniff(texto[:2048], delimiters=';,|\t')
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ';'
    reader = csv.DictReader(io.StringIO(texto), delimiter=delimiter)
    return reader


STATUS_CV_MAP = {
    'disponível': 'disponivel', 'disponivel': 'disponivel',
    'bloqueada':  'bloqueado',  'bloqueado':  'bloqueado',
    'reservada':  'reservado',  'reservado':  'reservado',
    'vendida':    'vendido',    'vendido':    'vendido',
    'permuta':    'permuta',
    'qa':         'qa',
}


def _parse_valor_br(raw):
    import re
    s = re.sub(r'[R$\s]', '', str(raw or '')).strip()
    s = s.replace('.', '').replace(',', '.')
    try:
        from decimal import Decimal
        return Decimal(s) if s else Decimal('0')
    except Exception:
        return None


def _import_tabela_cv(arquivo, empreendimento):
    """UNIDADE ; SITUAÇÃO ; VALOR TOTAL — atualiza status e valor_tabela."""
    texto = _decode_csv(arquivo)
    reader = _csv_reader(texto, delimiter=';')
    fieldnames = [f.strip() for f in (reader.fieldnames or [])]
    atualizados = 0
    nao_encontrados = []
    for i, row in enumerate(reader, 2):
        row = {k.strip(): v for k, v in row.items()}
        num = row.get('UNIDADE', '').strip()
        if not num:
            continue
        status_raw = row.get('SITUAÇÃO', '').strip()
        valor_raw  = row.get('VALOR TOTAL', '').strip()
        status = STATUS_CV_MAP.get(status_raw.lower(), None)
        valor  = _parse_valor_br(valor_raw)
        try:
            u = Unidade.objects.get(bloco__empreendimento=empreendimento, numero=num)
            fields = {}
            if status:
                fields['status'] = status
            if valor is not None:
                fields['valor_tabela'] = valor
            if fields:
                Unidade.objects.filter(pk=u.pk).update(**fields)
            atualizados += 1
        except Unidade.DoesNotExist:
            nao_encontrados.append(num)
        except Unidade.MultipleObjectsReturned:
            nao_encontrados.append(num)
    return atualizados, nao_encontrados


def _import_unidades(arquivo, empreendimento):
    """CSV completo de unidades. Aceita tanto o formato interno (snake_case)
    quanto o formato de exportação do incorporadora (nomes com acentos e unidades).
    Também aceita blocos 'Garagens' e 'Hobby Boxes' buscando a unidade pelo
    número em todos os blocos do empreendimento."""
    import csv, io, unicodedata

    # Mapeamento de nomes de colunas do formato exportação → nome interno
    COL_MAP = {
        'numero': 'numero', 'número': 'numero',
        'numeros_adicionais': 'numeros_adicionais',
        'números adicionais': 'numeros_adicionais',
        'numeros adicionais': 'numeros_adicionais',
        'bloco': 'bloco', 'ordem': 'ordem', 'tipo': 'tipo',
        'tipologia': 'tipologia',
        'localizacao': 'localizacao', 'localização': 'localizacao',
        'area_privativa': 'area_privativa',
        'área privativa (m²)': 'area_privativa',
        'area privativa (m2)': 'area_privativa',
        'area_privativa_acessoria': 'area_privativa_acessoria',
        'área priv. acess. (m²)': 'area_privativa_acessoria',
        'area priv. acess. (m2)': 'area_privativa_acessoria',
        'area_comum': 'area_comum',
        'área comum (m²)': 'area_comum',
        'area comum (m2)': 'area_comum',
        'fracao_ideal': 'fracao_ideal', 'fração ideal': 'fracao_ideal',
        'fracao ideal': 'fracao_ideal',
        'valor_tabela': 'valor_tabela',
        'valor tabela (r$)': 'valor_tabela',
        'status': 'status',
        'cliente': 'cliente_nome', 'cliente_nome': 'cliente_nome',
        'descricao1': 'descricao1', 'descrição 1': 'descricao1', 'descricao 1': 'descricao1',
        'descricao2': 'descricao2', 'descrição 2': 'descricao2', 'descricao 2': 'descricao2',
        'descricao3': 'descricao3', 'descrição 3': 'descricao3', 'descricao 3': 'descricao3',
    }

    def strip_accents(s):
        return ''.join(c for c in unicodedata.normalize('NFD', s)
                       if unicodedata.category(c) != 'Mn')

    def norm_col(k):
        """Normaliza nome de coluna: minúsculo, sem acentos."""
        return strip_accents(k.strip().lower())

    texto = _decode_csv(arquivo)
    reader = csv.DictReader(io.StringIO(texto), delimiter=';')

    # Zera unidades existentes antes de importar
    ItemTabelaVendas.objects.filter(unidade__bloco__empreendimento=empreendimento).delete()
    Unidade.objects.filter(bloco__empreendimento=empreendimento).delete()

    TIPO_MAP = {
        'apartamento': 'apartamento', 'garagem': 'garagem',
        'hobby box': 'hobby_box', 'hobby_box': 'hobby_box',
        'loja': 'loja', 'sala': 'sala',
    }

    criados = atualizados = 0
    blocos_nao_encontrados = set()
    for i, row in enumerate(reader, 2):
        # Normaliza chaves usando COL_MAP (com fallback para a chave normalizada)
        row_n = {}
        for k, v in row.items():
            nk = norm_col(k)
            interno = COL_MAP.get(nk) or COL_MAP.get(k.strip().lower()) or nk
            row_n[interno] = (v or '').strip()

        bloco_nome = row_n.get('bloco', '').strip()
        num        = row_n.get('numero', '').strip()
        if not num:
            continue

        bloco = None
        if bloco_nome:
            try:
                bloco = Bloco.objects.get(empreendimento=empreendimento, nome__iexact=bloco_nome)
            except Bloco.DoesNotExist:
                pass

        defaults = {}
        tipo_raw = row_n.get('tipo', '').strip().lower()
        if tipo_raw:
            defaults['tipo'] = TIPO_MAP.get(tipo_raw, tipo_raw)

        for campo in ('tipologia', 'localizacao',
                      'descricao1', 'descricao2', 'descricao3',
                      'cliente_nome', 'numeros_adicionais'):
            if row_n.get(campo):
                defaults[campo] = row_n[campo]

        if row_n.get('ordem'):
            try:
                defaults['ordem'] = int(row_n['ordem'])
            except ValueError:
                pass

        for campo in ('area_privativa', 'area_privativa_acessoria', 'area_comum',
                      'fracao_ideal', 'valor_tabela'):
            if row_n.get(campo):
                v = _parse_valor_br(row_n[campo])
                if v is not None:
                    defaults[campo] = v

        status_raw = strip_accents(row_n.get('status', '').lower())
        if status_raw in STATUS_CV_MAP:
            defaults['status'] = STATUS_CV_MAP[status_raw]
        elif row_n.get('status'):
            # Tenta sem acentos no mapa também
            for k, v in STATUS_CV_MAP.items():
                if strip_accents(k) == status_raw:
                    defaults['status'] = v
                    break

        if bloco:
            _, criado = Unidade.objects.update_or_create(
                bloco=bloco, numero=num, defaults=defaults,
            )
            if criado: criados += 1
            else: atualizados += 1
        else:
            blocos_nao_encontrados.add(bloco_nome or '(sem bloco)')

    return criados + atualizados, sorted(blocos_nao_encontrados)


def _import_vinculos_emp(arquivo, empreendimento):
    """unidade ; vinculada ; numeros_adicionais — atualiza unidade_principal."""
    import csv, io
    texto = _decode_csv(arquivo)
    reader = csv.DictReader(io.StringIO(texto), delimiter=';')
    todas = list(Unidade.objects.filter(bloco__empreendimento=empreendimento).select_related('bloco'))
    por_numero = {}
    for u in todas:
        por_numero.setdefault(u.numero, []).append(u)
        for alias in [a.strip() for a in (u.numeros_adicionais or '').split(',') if a.strip()]:
            por_numero.setdefault(alias, []).append(u)

    def buscar(num):
        matches = por_numero.get(num, [])
        if len(matches) == 1:
            return matches[0], None
        if len(matches) > 1:
            return None, f'ambíguo'
        return None, f'não encontrado'

    Unidade.objects.filter(bloco__empreendimento=empreendimento, tipo__in=['garagem', 'hobby_box']).update(unidade_principal=None)
    vinculados = 0
    nao_encontrados = []
    for i, row in enumerate(reader, 2):
        row = {k.strip().lower(): (v or '').strip() for k, v in row.items()}
        num_p = row.get('unidade', '').strip()
        num_v = row.get('vinculada', '').strip()
        if not num_p or not num_v:
            continue
        principal, err = buscar(num_p)
        if err:
            nao_encontrados.append(num_p)
            continue
        vinculada, err = buscar(num_v)
        if err:
            nao_encontrados.append(num_v)
            continue
        vinculada.unidade_principal = principal
        update_fields = ['unidade_principal']
        if row.get('numeros_adicionais'):
            vinculada.numeros_adicionais = row['numeros_adicionais']
            update_fields.append('numeros_adicionais')
        vinculada.save(update_fields=update_fields)
        vinculados += 1
    return vinculados, nao_encontrados


def _import_atualizacao_mensal(arquivo, empreendimento):
    """bloco ; unidade ; situação ; valor total — atualiza status e valor_tabela por bloco+unidade."""
    import csv, io
    texto = _decode_csv(arquivo)
    try:
        import csv as _csv
        dialect = _csv.Sniffer().sniff(texto[:2048], delimiters=';,|\t')
        delim = dialect.delimiter
    except Exception:
        delim = ';'
    reader = csv.DictReader(io.StringIO(texto), delimiter=delim)
    atualizados = 0
    nao_encontrados = []
    for i, row in enumerate(reader, 2):
        row = {k.strip().lower(): (v or '').strip() for k, v in row.items()}
        bloco_nome = row.get('bloco', '').strip()
        num        = row.get('unidade', '').strip()
        if not bloco_nome or not num:
            continue
        status_raw = row.get('situação', row.get('situacao', '')).strip()
        valor_raw  = row.get('valor total', row.get('valor_total', '')).strip()
        status = STATUS_CV_MAP.get(status_raw.lower(), None)
        valor  = _parse_valor_br(valor_raw)
        try:
            u = Unidade.objects.get(bloco__empreendimento=empreendimento, bloco__nome__iexact=bloco_nome, numero=num)
            fields = {}
            if status:
                fields['status'] = status
            if valor is not None:
                fields['valor_tabela'] = valor
            if fields:
                Unidade.objects.filter(pk=u.pk).update(**fields)
            atualizados += 1
        except Unidade.DoesNotExist:
            nao_encontrados.append(f'{bloco_nome}/{num}')
    return atualizados, nao_encontrados


def _import_vendas_emp(arquivo, empreendimento):
    """Reserva ; Unidade ; Situação ; Cliente ; Imobiliária — atualiza cliente_nome e status."""
    import csv, io
    texto = _decode_csv(arquivo)
    reader = csv.DictReader(io.StringIO(texto), delimiter=';')
    atualizados = 0
    nao_encontrados = []
    for i, row in enumerate(reader, 2):
        row = {k.strip(): (v or '').strip() for k, v in row.items()}
        # busca case-insensitive pelas colunas-chave
        get = lambda *keys: next((row[k] for k in row if k.lower().strip() in [x.lower() for x in keys]), '')
        num        = get('Unidade', 'UNIDADE').strip()
        cliente    = get('Cliente', 'CLIENTE').strip()
        status_raw = get('Situação', 'Situacao', 'SITUAÇÃO').strip()
        email      = get('Email', 'E-mail', 'EMAIL').strip()
        if not num:
            continue
        status = STATUS_CV_MAP.get(status_raw.lower(), None)
        try:
            u = Unidade.objects.get(bloco__empreendimento=empreendimento, numero=num)
            fields = {}
            if cliente:
                fields['cliente_nome'] = cliente
            if email:
                fields['cliente_email'] = email
            if status:
                fields['status'] = status
            if fields:
                Unidade.objects.filter(pk=u.pk).update(**fields)
            atualizados += 1
        except Unidade.DoesNotExist:
            nao_encontrados.append(num)
        except Unidade.MultipleObjectsReturned:
            nao_encontrados.append(num)
    return atualizados, nao_encontrados


_IMPORTADORES = {
    'tabela_cv':          _import_tabela_cv,
    'unidades':           _import_unidades,
    'vinculos':           _import_vinculos_emp,
    'atualizacao_mensal': _import_atualizacao_mensal,
    'vendas':             _import_vendas_emp,
}


@login_required
def empreendimento_importar(request, pk):
    empreendimento = get_object_or_404(Empreendimento, pk=pk)

    if request.method == 'POST':
        tipo    = request.POST.get('tipo', '').strip()
        arquivo = request.FILES.get('arquivo')

        if tipo not in _IMPORTADORES:
            messages.error(request, 'Tipo de importação inválido.')
            return redirect('incorporadora:empreendimento_importar', pk=pk)
        if not arquivo:
            messages.error(request, 'Nenhum arquivo enviado.')
            return redirect('incorporadora:empreendimento_importar', pk=pk)

        try:
            total, nao_encontrados = _IMPORTADORES[tipo](arquivo, empreendimento)
            ImportLog.objects.create(
                empreendimento=empreendimento,
                tipo=tipo,
                nome_arquivo=arquivo.name,
                total_registros=total,
                importado_por=request.user,
            )
            msg = f'Importação concluída: {total} registro(s) processado(s).'
            if nao_encontrados:
                lista = ', '.join(nao_encontrados)
                msg += f' Bloco(s) não encontrado(s): {lista}.'
            messages.success(request, msg)
        except Exception as e:
            messages.error(request, f'Erro na importação: {e}')

        return redirect('incorporadora:empreendimento_importar', pk=pk)

    # Último log por tipo
    logs = {}
    for log in ImportLog.objects.filter(empreendimento=empreendimento):
        if log.tipo not in logs:
            logs[log.tipo] = log

    return render(request, 'incorporadora/empreendimento_importar.html', {
        'empreendimento': empreendimento,
        'logs': logs,
    })
