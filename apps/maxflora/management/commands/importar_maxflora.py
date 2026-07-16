"""
Importa a Tabela de Vendas do Max & Flora Shopping a partir de um arquivo Excel.

Uso:
    python manage.py importar_maxflora
    python manage.py importar_maxflora --arquivo=/caminho/para/arquivo.xlsx

O comando lê as abas "Tabela" e "Locatários", cruza os dados pelo campo EUC
e substitui completamente os registros anteriores.
"""

import re
from datetime import datetime, date
from pathlib import Path

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.maxflora.models import ImportacaoMaxFlora, UnidadeMaxFlora


def _to_float(v):
    if v is None:
        return None
    try:
        return float(str(v).replace(',', '.').strip())
    except (ValueError, TypeError):
        return None


def _to_date(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.date() if isinstance(v, datetime) else v
    s = str(v).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d/%m/%y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _find_xlsx():
    """Procura o arquivo Excel da MaxFlora mais recente na raiz do projeto."""
    candidatos = list(Path(settings.BASE_DIR).glob('MaxFlora*.xlsx'))
    if not candidatos:
        raise CommandError(
            'Arquivo MaxFlora*.xlsx não encontrado em BASE_DIR. '
            'Use --arquivo=/caminho/completo/arquivo.xlsx'
        )
    return str(max(candidatos, key=lambda p: p.stat().st_mtime))


class Command(BaseCommand):
    help = 'Importa Tabela de Vendas do Max & Flora a partir de Excel.'

    def add_arguments(self, parser):
        parser.add_argument('--arquivo', type=str, default=None,
                            help='Caminho do Excel (padrão: MaxFlora*.xlsx na raiz)')

    def handle(self, *args, **options):
        arquivo = options['arquivo'] or _find_xlsx()
        if not Path(arquivo).exists():
            raise CommandError(f'Arquivo não encontrado: {arquivo}')

        self.stdout.write(f'Lendo {Path(arquivo).name}…')
        wb = openpyxl.load_workbook(arquivo, data_only=True)

        # ── Aba Locatários ─────────────────────────────────────────────────
        ws_loc = next(
            (wb[name] for name in wb.sheetnames if 'locat' in name.lower()),
            None
        )

        locatarios = {}  # euc → nome
        if ws_loc:
            for row in ws_loc.iter_rows(min_row=3, values_only=True):
                euc, nome = row[0], row[1]
                if euc is not None:
                    locatarios[str(euc).strip()] = str(nome or '').strip()

        # ── Aba Tabela ─────────────────────────────────────────────────────
        ws = wb['Tabela']
        rows = list(ws.iter_rows(min_row=5, values_only=True))  # pula 4 linhas de cabeçalho

        unidades_data = []
        for ordem, row in enumerate(rows):
            euc = row[0]
            if euc is None:
                continue
            euc_str = str(euc).strip()
            if not euc_str:
                continue

            # Nova estrutura: A=Lojas B=Térreo C=Mez D=Total E=V.Vendas
            #                 F=Situação G=Aluguel H=Locado até I=Cond J=IPTU K=TCRS
            situacao_raw = str(row[5] or '').strip().upper()
            situacao = 'LOCADO' if situacao_raw == 'LOCADO' else 'DISPONIVEL'

            unidades_data.append({
                'loja':          euc_str,
                'locatario':     locatarios.get(euc_str, ''),
                'area_terreo':   _to_float(row[1]),
                'area_mezanino': _to_float(row[2]),
                'area_total':    _to_float(row[3]),
                'valor_vendas':  _to_float(row[4]),
                'situacao':      situacao,
                'valor_aluguel': _to_float(row[6]),
                'locado_ate':    _to_date(row[7]),
                'condominio':    _to_float(row[8]),
                'iptu':          _to_float(row[9]),
                'tcrs':          _to_float(row[10]),
                'ordem':         ordem,
            })

        self.stdout.write(f'  {len(unidades_data)} unidades encontradas.')

        with transaction.atomic():
            # Mantém histórico mas só exibe o mais recente (foreign key)
            imp = ImportacaoMaxFlora.objects.create(
                arquivo=Path(arquivo).name,
                total_unidades=len(unidades_data),
            )
            for d in unidades_data:
                UnidadeMaxFlora.objects.create(importacao=imp, **d)

            # Remove importações antigas (mantém apenas a última)
            ImportacaoMaxFlora.objects.exclude(pk=imp.pk).delete()

        self.stdout.write(self.style.SUCCESS(
            f'Importação concluída: {len(unidades_data)} unidades salvas '
            f'(importação #{imp.pk} em {imp.importado_em:%d/%m/%Y %H:%M}).'
        ))
