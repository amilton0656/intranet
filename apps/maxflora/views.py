import io
from datetime import datetime
from pathlib import Path

import openpyxl
from django.contrib import messages
from django.db import transaction
from django.db.models import Sum
from django.http import FileResponse, HttpResponse
from django.shortcuts import redirect, render

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

from .models import ImportacaoMaxFlora, UnidadeMaxFlora


# ── helpers ───────────────────────────────────────────────────────────────────

def _fmt_brl(v):
    if v is None:
        return ''
    return 'R$\xa0' + f'{float(v):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')


def _fmt_m2(v):
    if v is None:
        return ''
    return f'{float(v):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')


def _build_stats(importacao):
    qs = UnidadeMaxFlora.objects.filter(importacao=importacao)
    total = qs.count()
    locadas = qs.filter(situacao='LOCADO').count()
    disponiveis = total - locadas
    area = qs.aggregate(s=Sum('area_total'))['s'] or 0
    val = qs.exclude(euc='Estac.').aggregate(s=Sum('valor_vendas'))['s'] or 0
    return {
        'total': total,
        'locadas': locadas,
        'disponiveis': disponiveis,
        'pct_locado': round(locadas / total * 100) if total else 0,
        'area_total': _fmt_m2(area),
        'valor_total': _fmt_brl(val),
        'importado_em': importacao.importado_em,
        'arquivo': importacao.arquivo,
    }


def _parse_excel(fileobj):
    """Lê bytes/file-like do Excel e retorna lista de dicts prontos para salvar."""
    from datetime import datetime as dt, date

    wb = openpyxl.load_workbook(fileobj, data_only=True)

    # Locatários
    ws_loc = next((wb[n] for n in wb.sheetnames if 'locat' in n.lower()), None)
    locatarios = {}
    if ws_loc:
        for row in ws_loc.iter_rows(min_row=3, values_only=True):
            if row[0] is not None:
                locatarios[str(row[0]).strip()] = str(row[1] or '').strip()

    # Tabela
    ws = wb['Tabela']
    result = []
    for ordem, row in enumerate(ws.iter_rows(min_row=5, values_only=True)):
        euc = row[0]
        if euc is None:
            continue
        euc_str = str(euc).strip()
        if not euc_str:
            continue

        def _f(v):
            if v is None:
                return None
            try:
                return float(str(v).replace(',', '.').strip())
            except (ValueError, TypeError):
                return None

        def _d(v):
            if v is None:
                return None
            if isinstance(v, (dt, date)):
                return v.date() if isinstance(v, dt) else v
            s = str(v).strip()
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d/%m/%y'):
                try:
                    return dt.strptime(s, fmt).date()
                except ValueError:
                    pass
            return None

        sit = str(row[6] or '').strip().upper()
        result.append({
            'euc': euc_str,
            'espaco': int(row[1]) if row[1] is not None else None,
            'locatario': locatarios.get(euc_str, ''),
            'area_terreo': _f(row[2]),
            'area_mezanino': _f(row[3]),
            'area_total': _f(row[4]),
            'valor_vendas': _f(row[5]),
            'situacao': 'LOCADO' if sit == 'LOCADO' else 'DISPONIVEL',
            'valor_aluguel': _f(row[7]),
            'locado_ate': _d(row[8]),
            'condominio': _f(row[9]),
            'iptu_tcrs': _f(row[10]),
            'ordem': ordem,
        })
    return result


# ── views ─────────────────────────────────────────────────────────────────────

def tabela_vendas(request):
    importacao = ImportacaoMaxFlora.objects.first()
    unidades, stats = [], {}
    if importacao:
        unidades = list(UnidadeMaxFlora.objects.filter(importacao=importacao))
        stats = _build_stats(importacao)
    return render(request, 'maxflora/tabela.html', {
        'unidades': unidades,
        'stats': stats,
        'importacao': importacao,
    })


def importar_upload(request):
    if request.method != 'POST':
        return redirect('maxflora:tabela')

    f = request.FILES.get('arquivo')
    if not f:
        messages.error(request, 'Nenhum arquivo selecionado.')
        return redirect('maxflora:tabela')
    if not f.name.lower().endswith(('.xlsx', '.xls')):
        messages.error(request, 'O arquivo deve ser .xlsx ou .xls.')
        return redirect('maxflora:tabela')

    try:
        dados = _parse_excel(io.BytesIO(f.read()))
        if not dados:
            messages.error(request, 'Nenhuma unidade encontrada no arquivo.')
            return redirect('maxflora:tabela')

        with transaction.atomic():
            imp = ImportacaoMaxFlora.objects.create(
                arquivo=f.name,
                total_unidades=len(dados),
            )
            for d in dados:
                UnidadeMaxFlora.objects.create(importacao=imp, **d)
            ImportacaoMaxFlora.objects.exclude(pk=imp.pk).delete()

        messages.success(
            request,
            f'Importação concluída: {len(dados)} unidades carregadas de "{f.name}".'
        )
    except Exception as e:
        messages.error(request, f'Erro ao processar o arquivo: {e}')

    return redirect('maxflora:tabela')


# ── PDF ───────────────────────────────────────────────────────────────────────

_C_NAVY    = colors.HexColor('#1a1a2e')
_C_GREEN   = colors.HexColor('#1a7a4a')
_C_GREEN2  = colors.HexColor('#2e8f60')
_C_LOCADO  = colors.HexColor('#d4edda')
_C_DISP    = colors.HexColor('#dbeafe')
_C_ESTAC   = colors.HexColor('#f8f8e8')
_C_STRIPE  = colors.HexColor('#f4f9f6')
_C_WHITE   = colors.white
_C_GOLD    = colors.HexColor('#c8a951')


def exportar_pdf(request):
    importacao = ImportacaoMaxFlora.objects.first()
    if not importacao:
        return HttpResponse('Nenhum dado importado.', status=404)

    unidades = list(UnidadeMaxFlora.objects.filter(importacao=importacao))
    stats = _build_stats(importacao)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.4*cm, rightMargin=1.4*cm,
        topMargin=0.72*cm, bottomMargin=1.2*cm,
    )
    W = doc.width
    styles = getSampleStyleSheet()

    def ps(name, **kw):
        return ParagraphStyle(name, parent=styles['Normal'], **kw)

    hdr_s   = ps('mfh', fontSize=6,   leading=8,  alignment=1, textColor=colors.HexColor('#333333'),
                 fontName='Helvetica-Bold')
    cell_s  = ps('mfc', fontSize=7,   leading=9)
    cell_r  = ps('mfcr', fontSize=7,  leading=9,  alignment=2)
    cell_c  = ps('mfcc', fontSize=7,  leading=9,  alignment=1)
    euc_s   = ps('mfe', fontSize=7,   leading=9,  fontName='Helvetica-Bold', alignment=1)
    brl_s   = ps('mfb', fontSize=7,   leading=9,  alignment=2, fontName='Helvetica-Bold')
    loc_s   = ps('mfl', fontSize=6.5, leading=8.5)
    sit_loc = ps('mfsl', fontSize=7,  leading=9,  alignment=1, fontName='Helvetica-Bold',
                 textColor=colors.HexColor('#155724'))
    sit_dis = ps('mfsd', fontSize=7,  leading=9,  alignment=1, fontName='Helvetica-Bold',
                 textColor=colors.HexColor('#0a58ca'))

    # ── Logo + cabeçalho ────────────────────────────────────────────────────
    logo_path = Path(__file__).parent / 'static' / 'maxflora' / 'img' / 'logo.jpg'
    story = []

    # Linha de cabeçalho: logo | título
    if logo_path.exists():
        logo_img = Image(str(logo_path), width=3.5*cm, height=1.8*cm, kind='proportional')
    else:
        logo_img = Paragraph('Max & Flora', ps('lp', fontSize=14, fontName='Helvetica-Bold',
                                                textColor=_C_GREEN))

    title_p  = Paragraph(
        '<b><font size="16" color="#1a7a4a">TABELA DE VENDAS</font></b>',
        ps('tp', alignment=0, leading=18),
    )
    sub_p    = Paragraph(
        '<font size="8" color="#6c757d">Max &amp; Flora Center Administradora de Shoppings Ltda.</font>',
        ps('sp', alignment=0, leading=12),
    )
    date_p   = Paragraph(
        f'<font size="7" color="#aaaaaa">Gerado em {datetime.now():%d/%m/%Y %H:%M}</font>',
        ps('dp', alignment=0, leading=10),
    )

    hdr_table = Table(
        [[logo_img, [title_p, sub_p, date_p]]],
        colWidths=[3.8*cm, W - 3.8*cm],
    )
    hdr_table.setStyle(TableStyle([
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING',   (1, 0), (1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(hdr_table)

    story.append(Spacer(1, 0.3*cm))

    # ── Tabela de dados ──────────────────────────────────────────────────────
    # Larguras (total = W ≈ 26.7cm)
    # EUC | Esp | Locatário | Térreo | Mez | Total | V.Vendas | Sit | Aluguel | Loc.até | Cond | IPTU
    COL_W = [
        1.5*cm, 1.2*cm, 5.0*cm,
        1.8*cm, 1.8*cm, 1.8*cm,
        3.0*cm, 2.0*cm,
        2.5*cm, 1.8*cm, 2.3*cm, 2.3*cm,
    ]
    # Ajusta última coluna para preencher a largura exata
    COL_W[-1] = W - sum(COL_W[:-1])

    def h(txt):
        return Paragraph(f'<b>{txt}</b>', hdr_s)

    HEADERS = [
        [h('EUC'), h('ESP.\nCOM.'), h('LOCATÁRIO'),
         h('ÁREA PRIV. (m²)'), h(''), h(''),
         h('VALOR DE\nVENDAS'), h('SITUAÇÃO'),
         h('VALOR DO\nALUGUEL'), h('LOCADO\nATÉ'), h('VALOR DO\nCONDOMÍNIO'), h('VALOR DO\nIPTU / TCRS')],
        [h(''), h(''), h(''),
         h('Térreo'), h('Mezanino'), h('Total'),
         h(''), h(''), h(''), h(''), h(''), h('')],
    ]

    def make_row(u):
        sit_par = Paragraph('LOCADO',    sit_loc) if u.locado else Paragraph('DISPONÍVEL', sit_dis)
        return [
            Paragraph(str(u.euc), euc_s),
            Paragraph(str(u.espaco or ''), cell_c),
            Paragraph(u.locatario or '', loc_s),
            Paragraph(_fmt_m2(u.area_terreo),   cell_r),
            Paragraph(_fmt_m2(u.area_mezanino),  cell_r),
            Paragraph(_fmt_m2(u.area_total),     ps('atot', fontSize=7, leading=9,
                                                   alignment=2, fontName='Helvetica-Bold')),
            Paragraph(_fmt_brl(u.valor_vendas) if u.euc != 'Estac.' else '', brl_s),
            sit_par,
            Paragraph(_fmt_brl(u.valor_aluguel), cell_r),
            Paragraph(u.locado_ate.strftime('%d/%m/%Y') if u.locado_ate else '', cell_c),
            Paragraph(_fmt_brl(u.condominio),    cell_r),
            Paragraph(_fmt_brl(u.iptu_tcrs),     cell_r),
        ]

    table_data = HEADERS + [make_row(u) for u in unidades]
    tbl = Table(table_data, colWidths=COL_W, repeatRows=2)

    # Estilos base
    tbl_cmds = [
        # Cabeçalhos — cinza bem claro
        ('BACKGROUND',    (0, 0), (-1, 1), colors.HexColor('#e8e8e8')),
        ('TEXTCOLOR',     (0, 0), (-1, 1), colors.HexColor('#333333')),
        # Span de "ÁREA PRIV." nas 3 colunas
        ('SPAN',          (3, 0), (5, 0)),
        # Células vazias do segundo header mergeadas
        ('SPAN',          (0, 0), (0, 1)),
        ('SPAN',          (1, 0), (1, 1)),
        ('SPAN',          (2, 0), (2, 1)),
        ('SPAN',          (6, 0), (6, 1)),
        ('SPAN',          (7, 0), (7, 1)),
        ('SPAN',          (8, 0), (8, 1)),
        ('SPAN',          (9, 0), (9, 1)),
        ('SPAN',         (10, 0), (10, 1)),
        ('SPAN',         (11, 0), (11, 1)),
        # Alinhamento cabeçalho
        ('VALIGN',        (0, 0), (-1, 1), 'MIDDLE'),
        # Linhas de dados
        ('GRID',          (0, 0), (-1, -1), 0.3, colors.HexColor('#c8d8cf')),
        ('TOPPADDING',    (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING',   (0, 0), (-1, -1), 4),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
        ('VALIGN',        (0, 2), (-1, -1), 'MIDDLE'),
        # Sub-header: mesma cor da primeira linha
        ('BACKGROUND',    (0, 1), (-1, 1), colors.HexColor('#e8e8e8')),
    ]

    # Zebra sutil nas linhas de dados (sem cor por situação)
    for i in range(len(unidades)):
        if i % 2 == 1:
            tbl_cmds.append(('BACKGROUND', (0, i + 2), (-1, i + 2), colors.HexColor('#f8f9fa')))

    tbl.setStyle(TableStyle(tbl_cmds))
    story.append(tbl)

    doc.build(story)
    buf.seek(0)
    resp = FileResponse(buf, content_type='application/pdf')
    resp['Content-Disposition'] = 'inline; filename="MaxFlora_TabelaVendas.pdf"'
    return resp
